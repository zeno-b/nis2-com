"""
NIS2 Belgian Companies – Scanner + Contact Intelligence

Combines two tools in one pipeline:

1. NIS2 scanner  – extracts NIS2-regulated Belgian companies from KBO Open
Data CSVs, resolves their websites, runs passive Nuclei security checks,
and produces a per-host coverage report.
1. Contact intel – for every scanned company (or any single target via
–contact-only), runs 9 OSINT sources to rank who to call/email:
KBO mandataries → website staff pages → Belgisch Staatsblad →
Google / Bing / DuckDuckGo SERPs → LinkedIn public profiles →
Hunter.io API (optional) → SMTP RCPT-TO verification → score ranking

Bug fixed vs original nis2_scan.py:
-no-mhe removed from the nuclei command.  In nuclei v3+ this flag
("no-multi-host-execution") stops template iteration after the first host
batch, so only a fraction of targets were ever scanned and only the first
matcher in multi-matcher templates appeared to fire.

CBE Open Data (https://kbopub.economie.fgov.be/kbo-open-data/):
activity.csv       NACE / sector per company
contact.csv        website URLs per company
denomination.csv   company names (optional but recommended)

Template requirements for full coverage reporting:

- Each matcher MUST have a unique `name:` field.
- Templates should use `matchers-condition: or` (nuclei default).

Usage modes:

# Standard NIS2 scan

python nis2_combined.py –sector Health –resolve-dns –export-xlsx

# Scan + auto-enrich top 20 affected companies with contact intelligence

python nis2_combined.py –sector Health –enrich-contacts –contact-limit 20

# Contact intelligence only (no scan needed)

python nis2_combined.py –contact-only –kbo 0419649912 –domain bhak.be

# Full pipeline with Hunter.io verification

python nis2_combined.py –sector Health –enrich-contacts –hunter-key KEY
"""

import argparse
import base64
import concurrent.futures
import csv
import functools
import html as html_lib
import json
import mimetypes
import os
import re

# FIX 4: removed unused `import signal`

import socket
import shutil
import subprocess
import sys
import threading
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse
import urllib.request
import urllib.error

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Prerequisite check

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _check_prerequisites():
    import importlib.util, shutil, platform, subprocess as _sp

    REQUIRED = {"pandas": "pandas"}
    OPTIONAL = {
        "tqdm":       ("tqdm",       "progress bars"),
        "colorama":   ("colorama",   "coloured terminal output"),
        "yaml":       ("pyyaml",     "--config YAML support + template parsing"),
        "tldextract": ("tldextract", "accurate root-domain deduplication"),
        "openpyxl":   ("openpyxl",   "Excel report export (--export-xlsx)"),
        "orjson":     ("orjson",     "fast JSON parsing for large result files"),
        # Contact intel deps
        "requests":   ("requests",   "contact intel HTTP client"),
        "bs4":        ("beautifulsoup4", "contact intel HTML parser"),
        "dns":        ("dnspython",  "contact intel SMTP MX lookup"),
    }

    print(f"Python {platform.python_version()}  |  "
          f"{platform.system()} {platform.release()}")
    print(f"Script : {Path(__file__).resolve()}\n")

    def _pip(pip_name):
        print(f"  Installing {pip_name} …", end=" ", flush=True)
        r = _sp.run([sys.executable, "-m", "pip", "install", pip_name, "--quiet"],
                    capture_output=True, text=True)
        if r.returncode == 0:
            print("✓"); return True
        print("✗ FAILED"); print(f"    {r.stderr.strip()}"); return False

    failed = []
    for imp, pip in REQUIRED.items():
        if importlib.util.find_spec(imp) is None:
            print(f"[ Required: {imp} missing ]")
            if not _pip(pip): failed.append(pip)
    if failed:
        print(f"\n[ FATAL ] Run: pip install {' '.join(failed)}")
        sys.exit(1)

    any_opt = False
    for imp, (pip, purpose) in OPTIONAL.items():
        if importlib.util.find_spec(imp) is None:
            if not any_opt:
                print("[ Optional packages – attempting auto-install ]")
                any_opt = True
            print(f"  {imp:<14} ({purpose})", end=" ")
            if not _pip(pip):
                print(f"    → continuing without {imp}")
    if any_opt:
        print()

    # FIX 8: replace deprecated urlretrieve with urlopen-based download
    def _install_nuclei_linux():
        print("  Attempting nuclei auto-install on Linux …")
        try:
            import zipfile, stat as _st
            api = "https://api.github.com/repos/projectdiscovery/nuclei/releases/latest"
            with urllib.request.urlopen(api, timeout=15) as r:
                rel = json.loads(r.read())
            ver     = rel["tag_name"]
            zip_url = next(a["browser_download_url"] for a in rel["assets"]
                           if "linux_amd64" in a["name"] and a["name"].endswith(".zip"))
            zp = Path("/tmp/nuclei_dl.zip")
            print(f"  Downloading nuclei {ver} …", end=" ", flush=True)
            with urllib.request.urlopen(zip_url, timeout=120) as resp, \
                 open(zp, "wb") as fout:
                fout.write(resp.read())
            print("✓")
            idir = Path("/usr/local/bin")
            if not os.access(idir, os.W_OK):
                idir = Path.home() / ".local" / "bin"
                idir.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zp) as z:
                z.extract("nuclei", idir)
            nb = idir / "nuclei"
            nb.chmod(nb.stat().st_mode | _st.S_IEXEC)
            os.environ["PATH"] = str(idir) + os.pathsep + os.environ.get("PATH", "")
            print(f"  ✓  nuclei installed at {nb}")
            return True
        except Exception as e:
            print(f"  ✗  auto-install failed: {e}"); return False

    nuclei_path = (shutil.which("nuclei") or
                   shutil.which(str(Path(__file__).parent / "nuclei")) or
                   shutil.which(str(Path(__file__).parent / "nuclei.exe")))
    if nuclei_path:
        print(f"  ✓  nuclei found at {nuclei_path}")
        os.environ["PATH"] = (str(Path(__file__).parent) + os.pathsep +
                              os.environ.get("PATH", ""))
    elif platform.system() == "Linux":
        print("  ✗  nuclei not found — attempting auto-install …")
        if not _install_nuclei_linux():
            print("     Manual: https://github.com/projectdiscovery/nuclei/releases")
    else:
        print("  ✗  nuclei not found → dry-run mode")
        print("     Place nuclei.exe next to this script, or add to PATH")
        print("     Download: https://github.com/projectdiscovery/nuclei/releases")
    print()

# Allow deterministic imports in tests/automation without network/package side effects.
if os.environ.get("SCANNER_SKIP_PREREQ_CHECK", "").strip().lower() not in (
    "1", "true", "yes", "on",
):
    _check_prerequisites()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Deferred imports

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import pandas as pd

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    def tqdm(iterable=None, *a, **k):
        return iterable if iterable is not None else iter([])

def _pbar(iterable, desc: str, total=None, unit: str = "it"):
    """tqdm wrapper that fits the current terminal width and clears when done.

    dynamic_ncols re-measures columns on every redraw so a window narrower
    than the bar no longer wraps and leaves a garbled trail; leave=False
    erases the bar on completion, since the caller prints its own summary
    line. A compact bar_format degrades cleanly on very narrow terminals.
    """
    if not HAS_TQDM:
        return iterable
    return tqdm(
        iterable,
        desc=desc,
        total=total,
        unit=unit,
        dynamic_ncols=True,
        leave=False,
        mininterval=0.1,
    )

try:
    from colorama import Fore, Style, init as _cinit
    # strip=None / convert=None lets colorama auto-detect: it passes ANSI
    # through on capable terminals (Windows Terminal, VS Code, macOS, Linux)
    # and converts on legacy Windows consoles. Forcing strip=True on Windows
    # removed colour even where the terminal supported it.
    _cinit(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False

    class _StubColor:
        def __getattr__(self, _):
            return ""

    Fore = _StubColor()
    Style = _StubColor()

try:
    import yaml;       HAS_YAML = True
except ImportError:    HAS_YAML = False

try:
    import tldextract; HAS_TLDEXTRACT = True
except ImportError:    HAS_TLDEXTRACT = False

try:
    import orjson;     HAS_ORJSON = True
except ImportError:    HAS_ORJSON = False

try:
    import openpyxl; HAS_OPENPYXL = True   # noqa: F401 (probe: sets availability flag)
except ImportError:    HAS_OPENPYXL = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Defaults

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ACTIVITY_FILE           = "activity.csv"
CONTACT_FILE            = "contact.csv"
DENOMINATION_FILE       = "denomination.csv"
OUTPUT_DIR              = "nis2_scan_output"
DEFAULT_TEMPLATES       = ["t.yaml"]
DEFAULT_RATE            = 150
DEFAULT_CONCUR          = 25
DEFAULT_TIMEOUT         = 10
DEFAULT_SEVERITY        = "low,medium,high,critical"
DEFAULT_CONTACT_WORKERS = 4
CHECKPOINT_FILE         = "checkpoint.json"
RETRY_FILE              = "retry_targets.txt"
DEAD_TARGETS_FILE       = "dead_targets.txt"
URL_LOOKUP_FILE         = "url_company_lookup.json"
COVERAGE_CSV            = "full_coverage_report.csv"
SCAN_RESULTS_JSON       = "scan_results.json"
SCAN_RESULTS_CSV        = "scan_results.csv"
SCAN_RESULTS_HTML       = "scan_results.html"
TIMINGS_FILE            = "step_timings.json"
TARGETS_MAX_AGE_MINUTES = 60
DNS_WORKERS             = 100
DNS_TIMEOUT             = 3
URL_WORKERS             = 50
URL_TIMEOUT             = 8
TABLE_MAX_COL           = 48
KBO_BASE                = ("https://kbopub.economie.fgov.be/kbopub/"
"toonondernemingps.html?ondernemingsnummer=")

def default_output_dir(base_dir: str = OUTPUT_DIR) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_dir}_{stamp}"

SEV_ORDER  = ["critical", "high", "medium", "low", "info", "unknown"]
SEV_COLORS = {
"critical": Fore.RED,    "high":    Fore.RED,
"medium":   Fore.YELLOW, "low":     Fore.CYAN,
"info":     Fore.WHITE,  "unknown": Fore.WHITE,
}

FINDING_MATCHER_DESCRIPTIONS: Dict[str, str] = {
    "missing-hsts":                   "Strict-Transport-Security not set - no HTTPS enforcement",
    "weak-hsts-max-age":              "HSTS max-age < 1 year - short expiry reduces protection",
    "hsts-missing-includesubdomains": "HSTS present but missing includeSubDomains - subdomains unprotected",
    "missing-csp":                    "No Content-Security-Policy - XSS protection policy absent",
    "weak-csp-unsafe-inline":         "CSP allows unsafe-inline - inline script execution permitted",
    "weak-csp-unsafe-eval":           "CSP allows unsafe-eval - eval() and similar not blocked",
    "missing-x-frame-options":        "No X-Frame-Options and no CSP frame-ancestors - clickjacking possible",
    "missing-x-content-type-options": "X-Content-Type-Options: nosniff not set - MIME sniffing enabled",
    "missing-referrer-policy":        "No Referrer-Policy - full URLs may leak to third parties",
    "missing-permissions-policy":     "No Permissions-Policy - camera/mic/geo access unrestricted",
    "missing-cache-control":          "No Cache-Control - responses may be cached by proxies",
    "missing-coop":                   "No Cross-Origin-Opener-Policy - Spectre-class attacks possible",
    "missing-corp":                   "No Cross-Origin-Resource-Policy - resources readable cross-origin",
    "cookie-missing-secure":          "Cookie set without Secure flag - transmitted over plain HTTP",
    "cookie-missing-httponly":        "Cookie set without HttpOnly - readable by JavaScript (XSS risk)",
    "cookie-missing-samesite":        "Cookie set without SameSite - CSRF attack surface",
    "cors-wildcard-origin":           "CORS wildcard (*) - any origin can read responses",
    "cors-reflects-origin":           "CORS reflects arbitrary Origin header - credentials leakable",
    "server-version-disclosure":      "Server header exposes software version - aids targeted attacks",
    "x-powered-by-disclosure":        "X-Powered-By header reveals backend technology stack",
    "aspnet-version-disclosure":      "ASP.NET version header present - exact .NET runtime exposed",
    "internal-ip-in-headers":         "RFC1918 internal IP address found in response headers",
    "generator-meta-disclosure":      "HTML meta generator tag reveals CMS name and version",
    "tech-wordpress":                 "WordPress CMS detected - check for outdated plugins/themes",
    "tech-joomla":                    "Joomla CMS detected - check version and extensions",
    "tech-drupal":                    "Drupal CMS detected - check version and modules",
    "tech-sharepoint":                "Microsoft SharePoint detected",
    "tech-jquery-version":            "jQuery version detectable - verify it is not end-of-life",
    "mixed-content-http-resource":    "HTTPS page loads resources over HTTP - mixed content warning",
    "external-script-detected":       "External script loaded without Subresource Integrity (SRI)",
    "stack-trace-disclosure":         "Stack trace or exception detail visible in response body",
    "debug-mode-indicators":          "Debug mode or development environment indicators in response",
    "sensitive-paths-disclosed":      "robots.txt lists sensitive paths (admin/api/backup/config...)",
    "ssl-certificate-expired":        "TLS certificate has passed its expiry date",
    "ssl-certificate-expiring-30d":   "TLS certificate expires within 30 days - renew urgently",
    "ssl-certificate-expiring-90d":   "TLS certificate expires within 90 days - plan renewal",
    "ssl-self-signed":                "Certificate is self-signed - not trusted by browsers",
    "ssl-weak-protocol":              "Server negotiated deprecated TLS 1.0 or TLS 1.1",
    "ssl-hostname-mismatch":          "Certificate CN/SAN does not match the hostname",
    # Matcher names emitted by the bundled default template (t.yaml). Keeping
    # these in sync with the template is what makes the report risk/remediation
    # finding-specific instead of falling back to generic severity text.
    "missing-strict-transport-security":        "Strict-Transport-Security header absent - HTTPS is not enforced, enabling SSL-strip/downgrade attacks",
    "hsts-without-includesubdomains":           "HSTS present but missing includeSubDomains - subdomains are not protected against downgrade",
    "hsts-without-preload":                      "HSTS present but missing preload - first-visit requests are unprotected until the policy is cached",
    "missing-content-security-policy":           "No Content-Security-Policy header - the primary defence against XSS and content injection is absent",
    "content-security-policy-unsafe-inline":     "CSP allows 'unsafe-inline' - inline scripts/styles can execute, weakening XSS protection",
    "content-security-policy-unsafe-eval":       "CSP allows 'unsafe-eval' - eval()-style dynamic code execution is permitted",
    "content-security-policy-wildcard-source":   "CSP uses a wildcard (*) source - any origin may supply content, defeating the policy",
    "content-security-policy-allows-http":       "CSP allows plain http: sources - mixed/insecure content can be loaded",
    "x-frame-options-allowall":                  "X-Frame-Options set to ALLOWALL - framing is effectively unrestricted, enabling clickjacking",
    "x-content-type-options-not-nosniff":        "X-Content-Type-Options present but not set to nosniff - MIME sniffing is still possible",
    "referrer-policy-unsafe-url":                "Referrer-Policy: unsafe-url - full URLs are sent to every destination, leaking sensitive paths",
    "permissions-policy-wildcard":               "Permissions-Policy grants a feature to all origins (=*) - powerful browser APIs are broadly exposed",
    "missing-cross-origin-opener-policy":        "No Cross-Origin-Opener-Policy - cross-origin window references enable Spectre-class attacks",
    "missing-cross-origin-resource-policy":      "No Cross-Origin-Resource-Policy - resources may be embedded/read by other origins",
    "missing-cross-origin-embedder-policy":      "No Cross-Origin-Embedder-Policy - cross-origin isolation is not enforced",
    "missing-x-permitted-cross-domain-policies": "No X-Permitted-Cross-Domain-Policies - Adobe clients may load cross-domain policy files",
    "permissive-cors-wildcard-origin":           "Access-Control-Allow-Origin: * - any website can read responses cross-origin",
    "invalid-cors-credentials-with-wildcard":    "CORS wildcard origin combined with credentials:true - a misconfiguration exposing authenticated data",
    "server-header-disclosure":                  "Server header exposes the web-server product (and possibly version), aiding targeted attacks",
    "x-aspnet-version-disclosure":               "X-AspNet(Mvc)-Version header exposes the exact ASP.NET runtime version",
    "via-header-disclosure":                     "Via header reveals proxy/cache infrastructure between the client and origin",
    "cookie-without-secure-attribute":           "Cookie set without the Secure attribute - it can be transmitted over plain HTTP",
    "cookie-without-httponly-attribute":         "Cookie set without HttpOnly - it is readable by JavaScript (credential theft via XSS)",
    "cookie-without-samesite-attribute":         "Cookie set without SameSite - this broadens the CSRF attack surface",
    "html-generator-meta-disclosure":            "HTML <meta generator> tag reveals the CMS/framework name and version",
    "directory-listing-index-page":              "Directory listing enabled ('Index of /') - file and structure enumeration is possible",
    "default-nginx-page":                        "Default nginx welcome page served - indicates an unconfigured or unused host",
    "default-apache-page":                       "Default Apache page served - indicates an unconfigured or unused host",
    "default-iis-page":                          "Default IIS page served - indicates an unconfigured or unused host",
}

FINDING_MATCHER_REMEDIATIONS: Dict[str, str] = {
    "missing-hsts":                   "Enable HSTS with a long max-age (>=31536000) and include subdomains.",
    "weak-hsts-max-age":              "Increase HSTS max-age to at least one year and preload when applicable.",
    "hsts-missing-includesubdomains": "Add includeSubDomains to HSTS after validating all subdomains support HTTPS.",
    "missing-csp":                    "Deploy a strict CSP using nonces/hashes and remove inline script allowances.",
    "weak-csp-unsafe-inline":         "Remove unsafe-inline and migrate scripts/styles to nonce/hash-based policies.",
    "weak-csp-unsafe-eval":           "Remove unsafe-eval and refactor code paths that rely on dynamic eval behavior.",
    "missing-x-frame-options":        "Block framing with X-Frame-Options DENY or CSP frame-ancestors.",
    "missing-x-content-type-options": "Set X-Content-Type-Options to nosniff on all HTTP responses.",
    "missing-referrer-policy":        "Set a restrictive Referrer-Policy such as strict-origin-when-cross-origin.",
    "missing-permissions-policy":     "Define a restrictive Permissions-Policy and disable unused browser features.",
    "missing-cache-control":          "Set Cache-Control directives, especially no-store for sensitive pages/data.",
    "missing-coop":                   "Set Cross-Origin-Opener-Policy to same-origin unless isolation requires otherwise.",
    "missing-corp":                   "Set Cross-Origin-Resource-Policy to same-site or same-origin where feasible.",
    "cookie-missing-secure":          "Mark cookies Secure and enforce HTTPS across every authenticated endpoint.",
    "cookie-missing-httponly":        "Mark session cookies HttpOnly to reduce credential theft via XSS.",
    "cookie-missing-samesite":        "Set SameSite=Lax/Strict for session cookies and review CSRF protections.",
    "cors-wildcard-origin":           "Replace wildcard CORS with an explicit allow-list of trusted origins.",
    "cors-reflects-origin":           "Stop reflecting arbitrary Origin values and validate against a strict allow-list.",
    "server-version-disclosure":      "Suppress version headers and keep the stack patched to current stable versions.",
    "x-powered-by-disclosure":        "Remove X-Powered-By and other framework disclosure headers in production.",
    "aspnet-version-disclosure":      "Disable ASP.NET version disclosure headers and verify hardened defaults.",
    "internal-ip-in-headers":         "Remove internal addressing from responses and sanitize upstream proxy headers.",
    "generator-meta-disclosure":      "Remove generator/version metadata from HTML and update CMS regularly.",
    "tech-wordpress":                 "Patch WordPress core, themes, and plugins; remove unused extensions.",
    "tech-joomla":                    "Update Joomla core/extensions and remove unsupported or unused modules.",
    "tech-drupal":                    "Update Drupal core/modules and apply vendor security advisories promptly.",
    "tech-sharepoint":                "Apply latest SharePoint security updates and harden internet-facing services.",
    "tech-jquery-version":            "Upgrade to a supported jQuery release and retest compatibility.",
    "mixed-content-http-resource":    "Serve all resources over HTTPS and enforce upgrades via CSP upgrade-insecure-requests.",
    "external-script-detected":       "Pin external scripts with Subresource Integrity and restrict via CSP.",
    "stack-trace-disclosure":         "Disable verbose errors in production and route details to protected logs.",
    "debug-mode-indicators":          "Disable debug/development modes in production configuration.",
    "sensitive-paths-disclosed":      "Remove sensitive paths from robots.txt and protect endpoints with auth/ACLs.",
    "ssl-certificate-expired":        "Replace expired certificates immediately and validate full certificate chain.",
    "ssl-certificate-expiring-30d":   "Renew certificates now and verify automated renewal monitoring.",
    "ssl-certificate-expiring-90d":   "Schedule certificate renewal and enable expiry alerting.",
    "ssl-self-signed":                "Use a trusted CA-issued certificate for public-facing services.",
    "ssl-weak-protocol":              "Disable TLS 1.0/1.1 and enforce TLS 1.2+ with modern ciphers.",
    "ssl-hostname-mismatch":          "Reissue certificate with correct SAN/CN entries for served hostnames.",
    # Matcher names emitted by the bundled default template (t.yaml).
    "missing-strict-transport-security":        "Enable HSTS with max-age>=31536000, includeSubDomains, and preload.",
    "hsts-without-includesubdomains":           "Add includeSubDomains to the HSTS header after confirming all subdomains support HTTPS.",
    "hsts-without-preload":                      "Add the preload directive (and submit to the preload list) once includeSubDomains is set.",
    "missing-content-security-policy":           "Deploy a strict Content-Security-Policy using nonces/hashes and avoid inline scripts.",
    "content-security-policy-unsafe-inline":     "Remove 'unsafe-inline' and migrate to nonce/hash-based script and style sources.",
    "content-security-policy-unsafe-eval":       "Remove 'unsafe-eval' and refactor code that relies on eval()/new Function().",
    "content-security-policy-wildcard-source":   "Replace wildcard (*) CSP sources with an explicit allow-list of trusted origins.",
    "content-security-policy-allows-http":       "Restrict CSP to https: sources and add upgrade-insecure-requests.",
    "x-frame-options-allowall":                  "Change X-Frame-Options to DENY/SAMEORIGIN and define CSP frame-ancestors.",
    "x-content-type-options-not-nosniff":        "Set X-Content-Type-Options strictly to 'nosniff'.",
    "referrer-policy-unsafe-url":                "Replace 'unsafe-url' with strict-origin-when-cross-origin or no-referrer.",
    "permissions-policy-wildcard":               "Scope Permissions-Policy features to specific trusted origins instead of '*'.",
    "missing-cross-origin-opener-policy":        "Set Cross-Origin-Opener-Policy: same-origin to enable origin isolation.",
    "missing-cross-origin-resource-policy":      "Set Cross-Origin-Resource-Policy: same-site (or same-origin) where feasible.",
    "missing-cross-origin-embedder-policy":      "Set Cross-Origin-Embedder-Policy: require-corp to complete cross-origin isolation.",
    "missing-x-permitted-cross-domain-policies": "Set X-Permitted-Cross-Domain-Policies: none unless cross-domain policy files are required.",
    "permissive-cors-wildcard-origin":           "Replace the wildcard CORS origin with an explicit allow-list of trusted origins.",
    "invalid-cors-credentials-with-wildcard":    "Never combine Access-Control-Allow-Credentials:true with a wildcard origin; echo only vetted origins.",
    "server-header-disclosure":                  "Suppress or normalise the Server header and keep the server patched to current stable versions.",
    "x-aspnet-version-disclosure":               "Disable ASP.NET version headers (enableVersionHeader=false) and the MVC version header.",
    "via-header-disclosure":                     "Strip or anonymise the Via header at the proxy/CDN layer.",
    "cookie-without-secure-attribute":           "Mark all cookies Secure and serve the site exclusively over HTTPS.",
    "cookie-without-httponly-attribute":         "Mark session cookies HttpOnly to block JavaScript access.",
    "cookie-without-samesite-attribute":         "Set SameSite=Lax or Strict on cookies and review CSRF defences.",
    "html-generator-meta-disclosure":            "Remove the generator meta tag and keep the CMS/framework updated.",
    "directory-listing-index-page":              "Disable automatic directory listing (e.g. 'autoindex off' / 'Options -Indexes').",
    "default-nginx-page":                        "Replace the default nginx page with the real site or remove the exposed host.",
    "default-apache-page":                       "Replace the default Apache page with the real site or remove the exposed host.",
    "default-iis-page":                          "Replace the default IIS page with the real site or remove the exposed host.",
}

SEVERITY_RISK_TEXT: Dict[str, str] = {
    "critical": "Critical risk: immediate exposure could enable compromise, data loss, or operational disruption.",
    "high":     "High risk: the weakness materially increases attack likelihood or impact and should be prioritized.",
    "medium":   "Medium risk: the weakness increases attack surface and should be addressed in planned hardening work.",
    "low":      "Low risk: the issue is primarily hardening debt but can help attackers when combined with other gaps.",
    "info":     "Informational risk: validate the exposure and use it to guide security hygiene improvements.",
    "unknown":  "Unrated risk: review the finding manually to confirm business impact and priority.",
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# NIS2 NACE codes

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

NIS2_NACE_PREFIXES = {
"Energy":                          ["0610","0620","1920","3511","3512","3513","3514","3515","3516","3521","3522","3523","3530","4950","4941"],
"Transport":                       ["4910","4920","4931","4932","4939","4941","4942","4950","5010","5020","5030","5040","5110","5121","5122","5223","5224"],
"Banking":                         ["6411","6419","6491","6492","6499"],
"Financial market infrastructure": ["6611","6612","6619","6621","6622","6629","6630"],
"Health":                          ["8610","8621","8622","8623","8690","8710","8720","8730","8790","8811","8812","8891","8899","2110","2120"],
"Drinking water":                  ["3600"],
"Wastewater":                      ["3700"],
"Digital infrastructure":          ["6110","6120","6130","6190","6201","6202","6209","6311","6312","6391","6399"],
"Public administration":           ["8411","8412","8413","8421","8422","8423","8430"],
"Space":                           ["5121","5122","7490"],
"Postal & courier":                ["5310","5320"],
"Waste management":                ["3811","3812","3821","3822","3831","3832","3900"],
"Chemicals":                       ["2011","2012","2013","2014","2015","2016","2017","2020","2030","2041","2042","2051","2052","2059","2060"],
"Food":                            ["1011","1012","1013","1020","1031","1032","1039","1041","1042","1051","1052","1061","1062","1071","1072","1073","1081","1082","1083","1084","1085","1086","1089","1091","1092","1101","1102","1103","1104","1105","1106","1107"],
"Manufacturing – medical devices": ["2651","2660","3250"],
"Manufacturing – computers":       ["2611","2612","2620","2630","2640","2680"],
"Manufacturing – electrical":      ["2711","2712","2720","2731","2732","2733","2740","2751","2752","2790"],
"Manufacturing – machinery":       ["2811","2812","2813","2814","2815","2821","2822","2823","2824","2825","2829","2830","2841","2849","2891","2892","2893","2894","2895","2896","2899"],
"Manufacturing – motor vehicles":  ["2910","2920","2931","2932"],
"Manufacturing – other transport": ["3011","3012","3020","3030","3040","3091","3092","3099"],
"Digital providers":               ["6201","6202","6209","6311","6312","6391","6399","6110"],
"Research":                        ["7210","7220"],
}

ANNEX_I_SECTORS = list(NIS2_NACE_PREFIXES.keys())[:10]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Step timer

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_STEP_TIMINGS: List[Dict] = []
_step_start_ts: float = 0.0
_step_counter: int = 0

def step_start(name: str) -> None:
    global _step_start_ts, _step_counter
    _step_start_ts = time.monotonic()
    _step_counter += 1
    _STEP_TIMINGS.append({"step": name, "elapsed": None})
    if not QUIET:
        print(_c(f"\n[{_step_counter}] {name}", Fore.CYAN + Style.BRIGHT))

def step_end() -> None:
    elapsed = time.monotonic() - _step_start_ts
    if _STEP_TIMINGS:
        _STEP_TIMINGS[-1]["elapsed"] = round(elapsed, 2)
    if not QUIET:
        print(_c(f"    done in {elapsed:.1f}s", Style.DIM))

def save_timings(output_dir: str) -> None:
    path = Path(output_dir) / TIMINGS_FILE
    try:
        with open(path, "w") as f:
            json.dump(_STEP_TIMINGS, f, indent=2)
    except OSError:
        pass  # non-fatal

def print_timings() -> None:
    if not _STEP_TIMINGS:
        return
    subhead("Step Timings")
    rows = [(t["step"],
    f"{t['elapsed']:.1f}s" if t["elapsed"] is not None else "—")
    for t in _STEP_TIMINGS]
    summary_table(rows, ["Step", "Elapsed"])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Logging helpers

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

USE_COLOR = True
QUIET     = False

def _c(t: str, col: str) -> str:
    return f"{col}{t}{Style.RESET_ALL}" if (HAS_COLOR and USE_COLOR and col) else t

def _trunc(s, n: int = TABLE_MAX_COL) -> str:
    s = str(s)
    return s if len(s) <= n else s[:n - 1] + "…"

def info(m: str)    -> None:
    if not QUIET: print(_c(f"[•] {m}", Fore.CYAN))

def ok(m: str)      -> None: print(_c(f"[✓] {m}", Fore.GREEN))
def warn(m: str)    -> None: print(_c(f"[!] {m}", Fore.YELLOW))
def error(m: str)   -> None: print(_c(f"[✗] {m}", Fore.RED))

def header(m: str) -> None:
    if QUIET:
        return
    # "[CI] ..." markers fire once per company across 9 sub-steps, so a full
    # banner each time floods the log. Render them as a single indented line.
    if m.startswith("[CI]"):
        print(_c(f"  › {m[4:].strip()}", Fore.CYAN))
        return
    # A "STEP N – ..." banner now duplicates the bright step line printed by
    # step_start(), so render it as a quiet rule instead of a full banner.
    if m.startswith("STEP "):
        print(_c(f"  {'─'*66}", Style.DIM))
        return
    print(_c(f"\n{'━'*68}\n    {m}\n{'━'*68}", Fore.MAGENTA + Style.BRIGHT))

def bullet(m: str) -> None:
    if not QUIET: print(_c(f"    {m}", Fore.WHITE))

def detail(m: str) -> None:
    if not QUIET: print(_c(f"      {m}", Style.DIM))

def _emit_progress(text: str) -> None:
    """Redraw a single-line \\r progress counter that fits the terminal width.

    Truncates to the current column count and pads to overwrite any residue
    from a longer previous line, so a window narrower than the counter no
    longer wraps and leaves a garbled trail. Pass plain text (no ANSI/colour),
    since the length has to be measured to be truncated correctly.
    """
    if QUIET:
        return
    try:
        cols = shutil.get_terminal_size((80, 20)).columns
    except Exception:
        cols = 80
    limit = max(cols - 1, 20)
    if len(text) > limit:
        text = text[:limit]
    sys.stdout.write("\r" + text.ljust(limit))
    sys.stdout.flush()

def subhead(m: str) -> None:
    if not QUIET: print(_c(f"\n  ── {m} ──", Fore.CYAN))

def summary_table(rows, cols, row_colors=None) -> None:
    """
    Print a left-aligned table.
    row_colors: optional list of colorama colour strings, one per row.
    All cells are truncated to TABLE_MAX_COL chars.
    Rows with fewer columns than cols are right-padded with empty strings.
    """
    if not rows:
        return
    n_cols = len(cols)
    trunc_rows = [[_trunc(r[i] if i < len(r) else "") for i in range(n_cols)]
    for r in rows]
    trunc_cols = [_trunc(c, 60) for c in cols]
    ws = [len(trunc_cols[i]) for i in range(n_cols)]
    for r in trunc_rows:
        for i in range(n_cols):
            ws[i] = max(ws[i], len(r[i]))
    fmt = "  " + "  ".join(f"{{:<{w}}}" for w in ws)
    print(_c(fmt.format(*trunc_cols), Fore.WHITE))
    print(_c("  " + "  ".join("─" * w for w in ws), Fore.WHITE))
    for i, r in enumerate(trunc_rows):
        line = fmt.format(*r)
        col  = row_colors[i] if (row_colors and i < len(row_colors)) else None
        print(_c(line, col) if col else line)

def _sev_bar(count: int, width: int = 40) -> str:
    return "█" * min(count, width)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Fast JSON helpers

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _json_loads(line: str) -> dict:
    """Use orjson when available (3-5× faster on large result files)."""
    if HAS_ORJSON:
        return orjson.loads(line)
    return json.loads(line)

def stream_findings(path: Path):
    """
    Generator: yield parsed finding dicts from a nuclei JSONL file.
    Skips error-type entries and unparseable lines without aborting.
    """
    errs = 0
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = _json_loads(line)
                    if isinstance(obj, dict) and obj.get("type") != "error":
                        yield obj
                except (ValueError, KeyError):
                    errs += 1
    except OSError as e:
        warn(f"Could not read results file: {e}")
        if errs:
            warn(f"{errs} result line(s) could not be parsed (skipped).")

def validate_results_file(path: Path) -> Tuple[bool, str]:
    """
    Sanity-check the nuclei output file.
    Returns (is_valid, message).
    """
    if not path.exists():
        return False, "File does not exist."
    if path.stat().st_size == 0:
        return False, "File is empty."
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    _json_loads(line)
                    return True, "OK"
                except (ValueError, KeyError) as e:
                    return False, f"First non-blank line is not valid JSON: {e}"
    except OSError as e:
        return False, f"Cannot open file: {e}"
    return False, "File contains no non-blank lines."

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Targets cache helpers

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def targets_are_fresh(targets_file: str,
    max_age_minutes: int = TARGETS_MAX_AGE_MINUTES
    ) -> Tuple[bool, Optional[str]]:
    p = Path(targets_file)
    if not p.exists() or p.stat().st_size == 0:
        return False, None
    age_sec = time.time() - p.stat().st_mtime
    age_min = age_sec / 60
    age_str = f"{int(age_min)}m ago" if age_min < 60 else f"{age_min/60:.1f}h ago"
    return age_min <= max_age_minutes, age_str

def load_targets_from_file(targets_file: str) -> List[str]:
    try:
        with open(targets_file) as f:
            return [l.strip() for l in f if l.strip()]
    except OSError as e:
        warn(f"Could not read targets file: {e}")
        return []

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Template check parser

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def resolve_template_path(t: str) -> Optional[Path]:
    """Return local filesystem path for template string, or None if not found."""
    local = Path(__file__).parent / t
    if local.exists():
        return local
    p = Path(t)
    if p.exists():
        return p
    return None

def parse_template_checks(template_paths: List[str]) -> Dict[str, dict]:
    """
    Parse each template YAML and extract every named matcher.
    Handles both single-document and multi-document (—) YAML files.

    Returns:
    {template_id: {"path": str, "checks": [{"name", "severity"}]}}
    """
    if not HAS_YAML:
        warn("pyyaml not installed – template check parsing unavailable. "
             "Run: pip install pyyaml")
        return {}

    result: Dict[str, dict] = {}
    for tp in template_paths:
        fpath = resolve_template_path(tp)
        if fpath is None:
            warn(f"Template '{tp}' not found locally; checks will be "
                 f"inferred from findings only.")
            result[Path(tp).stem] = {"path": tp, "checks": []}
            continue
        try:
            with open(fpath, encoding="utf-8") as f:
                docs = [d for d in yaml.safe_load_all(f)
                        if isinstance(d, dict)]
            if not docs:
                warn(f"Template '{tp}' contains no valid YAML documents – skipped.")
                continue

            for tmpl in docs:
                tid      = tmpl.get("id", fpath.stem)
                base_sev = tmpl.get("info", {}).get("severity", "unknown").lower()
                checks: List[dict] = []
                unnamed = 0

                # http and ssl blocks both carry matchers
                for block_key in ("http", "ssl"):
                    for req in tmpl.get(block_key, []):
                        cond = req.get("matchers-condition", "or").lower()
                        if cond not in ("or", "and"):
                            warn(f"Template '{tid}': unknown matchers-condition={cond!r}")
                        for matcher in req.get("matchers", []):
                            name = matcher.get("name")
                            if not name:
                                unnamed += 1
                                name = f"(unnamed-{unnamed})"
                                warn(f"Template '{tid}' has an unnamed matcher "
                                     f"→ reported as '{name}'. Add 'name:' for tracking.")
                            checks.append({
                                "name":     name,
                                "severity": base_sev,
                            })

                if checks:
                    ok(f"Template '{tid}': {len(checks)} check(s) parsed "
                       f"from {fpath.name}")
                else:
                    warn(f"Template '{tid}': no named matchers found — "
                         f"coverage inferred from findings only.")

                result[tid] = {"path": str(fpath), "checks": checks}

        except yaml.YAMLError as e:
            warn(f"Template '{tp}' YAML parse error: {e} – skipped.")
        except OSError as e:
            warn(f"Template '{tp}' could not be read: {e} – skipped.")

    return result

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# YAML config

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_yaml_config(path: str) -> dict:
    if not HAS_YAML:
        error("pyyaml not installed. Run: pip install pyyaml"); sys.exit(1)
    p = Path(path)
    if not p.exists():
        error(f"Config file not found: {path}"); sys.exit(1)
    try:
        with open(p, encoding="utf-8", errors="replace") as f:
            cfg = yaml.safe_load(f)
            if not isinstance(cfg, dict):
                error(f"{path} must be a YAML mapping."); sys.exit(1)
            ok(f"Loaded config: {path}")
            return cfg
    except yaml.YAMLError as e:
        error(f"YAML parse error in {path}: {e}"); sys.exit(1)
    except OSError as e:
        error(f"Cannot read config {path}: {e}"); sys.exit(1)

    # FIX 2: rewritten to correctly apply ALL value types from config,

    # including booleans (resolve_dns, resolve_urls, per_sector_dirs, no_retry, schedule).

    # The original inner if/elif only handled list-vs-None and silently dropped

    # boolean False values and list values when the arg was None.

def apply_config_to_args(cfg: dict, args) -> None:
    mapping = {
    "sector":"sector", "nace":"nace", "limit":"limit",
    "templates":"templates", "severity":"severity",
    "rate":"rate", "concur":"concur", "timeout":"timeout",
    "proxy":"proxy", "exclude":"exclude", "output_dir":"output_dir",
    "run_mode":"run_mode",
    "activity":"activity", "contact":"contact",
    "denomination":"denomination",
    "resolve_dns":"resolve_dns", "resolve_urls":"resolve_urls",
    "per_sector_dirs":"per_sector_dirs", "schedule":"schedule",
    "no_retry":"no_retry",
    "power_automate_webhook":"power_automate_webhook",
    "outlook_to":"outlook_to",
    "outlook_subject":"outlook_subject",
    "power_automate_timeout":"power_automate_timeout",
    "attach_report_files":"attach_report_files",
    "sharepoint_upload":"sharepoint_upload",
    "sharepoint_folder":"sharepoint_folder",
    "annex1_only":"annex1_only",
    "enrich_contacts":"enrich_contacts",
    "contact_only":"contact_only",
    "kbo":"kbo", "domain":"domain",
    "hunter_key":"hunter_key", "apollo_key":"apollo_key",
    "no_smtp":"no_smtp", "contact_proxy":"contact_proxy",
    "export_xlsx":"export_xlsx", "no_color":"no_color", "quiet":"quiet",
    "force_refresh":"force_refresh", "verbose_nuclei":"verbose_nuclei",
    }
    for ck, ak in mapping.items():
        if ck not in cfg:
            continue
        current = getattr(args, ak, None)
        # Only apply if the arg is still at its "unset" default:
        # None, empty list, or False (store_true flags start at False).
        if current in (None, [], False, "") or (ak == "run_mode" and current == "auto"):
            setattr(args, ak, cfg[ck])

def write_example_config(path: str) -> None:
    example = """\
# NIS2 Scanner – example config file
# All keys are optional. Explicit CLI flags always override config values.

# sector: NIS2 sector name(s) – run `--list-sectors` for the valid names.
sector:
  - Health
limit: 500
# run_mode: dryrun        # auto | dryrun | dryrun-skip-selected | run-from-dryrun
denomination: denomination.csv
resolve_dns: true
resolve_urls: true
per_sector_dirs: true
# templates: defaults to the bundled `t.yaml` when omitted.
templates:
  - t.yaml
severity: low,medium,high,critical
rate: 10
concur: 3
timeout: 15

# proxy: http://127.0.0.1:8080
# exclude: exclude_domains.txt
# output_dir: campaigns/my_campaign

# Power Automate / Outlook delivery
# power_automate_webhook: "https://prod-00.westeurope.logic.azure.com:443/workflows/..."
# outlook_to: "soc@example.com;cto@example.com"
# outlook_subject: "Weekly NIS2 Report"
# power_automate_timeout: 20
# attach_report_files: true
"""
    try:
        os.makedirs(Path(path).parent, exist_ok=True)
        with open(path, "w") as f:
            f.write(example)
            ok(f"Example config written: {path}")
    except OSError as e:
        error(f"Could not write example config to {path}: {e}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# NACE helpers

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def codes_for_filter(sectors: List[str], naces: List[str]) -> set:
    result: set = set()
    if sectors:
        for s in sectors:
            m = next((k for k in NIS2_NACE_PREFIXES if k.lower() == s.lower()), None)
            if m is None:
                error(f"Unknown sector: '{s}'")
                bullet("Run –list-sectors to see valid names.")
                sys.exit(1)
            result.update(NIS2_NACE_PREFIXES[m])

    if naces:
        for n in naces:
            if not re.fullmatch(r"\d{4,5}", str(n).strip()):
                error(f"Invalid NACE code '{n}' – must be 4 or 5 digits.")
                sys.exit(1)
        # The reference table and the activity.csv filter both operate at
        # 4-digit NACE granularity (the filter truncates NaceCode to [:4]),
        # so a 5-digit code must be narrowed to 4 or it can never match.
        result.update([str(n).strip()[:4] for n in naces])

    if not result:
        for codes in NIS2_NACE_PREFIXES.values():
            result.update(codes)
    return result

def get_sector_label(nace: str) -> str:
    n = str(nace).strip().zfill(4)
    for sector, codes in NIS2_NACE_PREFIXES.items():
        if any(n == c or n.startswith(c) or c.startswith(n) for c in codes):
            return sector
    return "Unknown"

def kbo_url(entity_number: str) -> str:
    num = entity_number.replace(".", "").replace(" ", "").strip()
    return KBO_BASE + num

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# CSV chunked loader

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _load_csv(path, required_cols, label, filter_fn=None, chunksize=200_000):
    info(f"Loading {label}: {path}")
    p = Path(path)
    if not p.exists():
        error(f"File not found: {path}")
        bullet("Download: https://kbopub.economie.fgov.be/kbo-open-data/")
        sys.exit(1)
    if p.stat().st_size == 0:
        error(f"File is empty: {path}"); sys.exit(1)

    mb = p.stat().st_size / 1024 ** 2
    bullet(f"File size: {mb:,.1f} MB")

    enc = "utf-8"
    try:
        with open(p, encoding="utf-8", errors="strict") as f:
            first = f.readline()
    except UnicodeDecodeError:
        enc = "latin-1"
        with open(p, encoding="latin-1") as f:
            first = f.readline()
        warn("UTF-8 decode failed – using latin-1")
    delim = ";" if first.count(";") > first.count(",") else ","
    bullet(f"Delimiter: '{delim}'  |  Encoding: {enc}")

    header_cols = {c.strip().strip('"').strip("'") for c in first.split(delim)}
    missing = required_cols - header_cols
    if missing:
        error(f"Missing columns in {path}: {missing}")
        bullet(f"Found: {sorted(header_cols)}"); sys.exit(1)

    chunks, rows_read, rows_kept, start = [], 0, 0, time.time()
    try:
        for chunk in pd.read_csv(p, dtype=str, encoding=enc, sep=delim,
                                  chunksize=chunksize, low_memory=False,
                                  on_bad_lines="warn"):
            chunk.columns = chunk.columns.str.strip().str.strip('"').str.strip("'")
            rows_read += len(chunk)
            if filter_fn:
                chunk = filter_fn(chunk)
            rows_kept += len(chunk)
            if not chunk.empty:
                chunks.append(chunk)
            elapsed = time.time() - start
            speed   = rows_read / elapsed if elapsed > 0 else 0
            _emit_progress(
                f"  ► Read {rows_read:>12,}  |  "
                f"Kept {rows_kept:>8,}  |  "
                f"{speed:>8,.0f} rows/s  |  {elapsed:>5.1f}s"
            )
    except MemoryError:
        print(); error("Out of RAM while loading CSV.")
        bullet("Try --limit or increase available memory."); sys.exit(1)
    except pd.errors.ParserError as e:
        print(); error(f"CSV parse error in {path}: {e}"); sys.exit(1)
    except OSError as e:
        print(); error(f"I/O error reading {path}: {e}"); sys.exit(1)
    except KeyboardInterrupt:
        print(); warn("Interrupted during CSV load."); sys.exit(130)

    print()
    elapsed = time.time() - start
    if not chunks:
        error(f"No rows matched in {path}."); sys.exit(1)
    df = pd.concat(chunks, ignore_index=True)
    ok(f"Done – {rows_read:,} read → {len(df):,} kept  "
       f"({elapsed:.1f}s  |  {mb / max(elapsed, 0.01):.1f} MB/s)")
    return df

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Data loaders

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_nis2_entities(
    activity_file: str,
    allowed_codes: set,
    stop_after_entities: Optional[int] = None,
    chunksize: int = 200_000,
) -> Tuple[pd.DataFrame, bool]:
    """
    Load and filter activity.csv to MAIN NIS2 rows.

    Returns:
        (df, reached_eof)
        reached_eof=False means early-stop triggered after collecting enough
        unique entities (used for --limit optimization).
    """
    p = Path(activity_file)
    info(f"Loading activity file: {activity_file}")
    if not p.exists():
        error(f"File not found: {activity_file}")
        bullet("Download: https://kbopub.economie.fgov.be/kbo-open-data/")
        sys.exit(1)
    if p.stat().st_size == 0:
        error(f"File is empty: {activity_file}")
        sys.exit(1)

    mb = p.stat().st_size / 1024 ** 2
    bullet(f"File size: {mb:,.1f} MB")

    enc, delim = _sniff_csv_format(p)
    bullet(f"Delimiter: '{delim}'  |  Encoding: {enc}")

    if stop_after_entities is not None:
        stop_after_entities = max(1, int(stop_after_entities))
        bullet(f"Early-stop goal: first {stop_after_entities:,} matched entities")

    required = {"EntityNumber", "NaceCode", "Classification"}
    chunks: List[pd.DataFrame] = []
    rows_read, rows_kept = 0, 0
    unique_entities: set = set()
    reached_eof = True
    start = time.time()

    try:
        for chunk in pd.read_csv(
            p,
            dtype=str,
            encoding=enc,
            sep=delim,
            chunksize=chunksize,
            low_memory=False,
            on_bad_lines="warn",
            usecols=list(required),
        ):
            chunk.columns = chunk.columns.str.strip().str.strip('"').str.strip("'")
            rows_read += len(chunk)
            chunk = chunk[
                chunk["Classification"].fillna("").astype(str).str.strip().str.upper() == "MAIN"
            ]
            nace4 = chunk["NaceCode"].fillna("").astype(str).str.strip().str[:4]
            chunk = chunk[nace4.isin(allowed_codes)]
            if not chunk.empty:
                chunk["EntityNumber"] = chunk["EntityNumber"].fillna("").astype(str).str.strip()
                chunk = chunk[chunk["EntityNumber"] != ""]
                if not chunk.empty:
                    rows_kept += len(chunk)
                    chunks.append(chunk)
                    if stop_after_entities is not None:
                        unique_entities.update(chunk["EntityNumber"].tolist())
                        if len(unique_entities) >= stop_after_entities:
                            reached_eof = False
                            elapsed = time.time() - start
                            speed = rows_read / elapsed if elapsed > 0 else 0
                            _emit_progress(
                                f"  ► Read {rows_read:>12,}  |  "
                                f"Kept {rows_kept:>8,}  |  "
                                f"Unique {len(unique_entities):>8,}  |  "
                                f"{speed:>8,.0f} rows/s  |  {elapsed:>5.1f}s"
                            )
                            break
            elapsed = time.time() - start
            speed = rows_read / elapsed if elapsed > 0 else 0
            unique_part = (
                f"  |  Unique {len(unique_entities):>8,}"
                if stop_after_entities is not None
                else ""
            )
            _emit_progress(
                f"  ► Read {rows_read:>12,}  |  "
                f"Kept {rows_kept:>8,}{unique_part}  |  "
                f"{speed:>8,.0f} rows/s  |  {elapsed:>5.1f}s"
            )
    except ValueError as e:
        msg = str(e)
        if "Usecols do not match columns" in msg:
            with open(p, encoding=enc, errors="ignore") as f:
                first = f.readline()
            header_cols = {c.strip().strip('"').strip("'") for c in first.split(delim)}
            missing = required - header_cols
            error(f"Missing columns in {activity_file}: {missing}")
            bullet(f"Found: {sorted(header_cols)}")
            sys.exit(1)
        error(f"CSV schema error in {activity_file}: {e}")
        sys.exit(1)
    except MemoryError:
        print()
        error("Out of RAM while loading CSV.")
        bullet("Try --limit or increase available memory.")
        sys.exit(1)
    except pd.errors.ParserError as e:
        print()
        error(f"CSV parse error in {activity_file}: {e}")
        sys.exit(1)
    except OSError as e:
        print()
        error(f"I/O error reading {activity_file}: {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        print()
        warn("Interrupted during CSV load.")
        sys.exit(130)

    print()
    if not chunks:
        error(f"No rows matched in {activity_file}.")
        sys.exit(1)

    df = pd.concat(chunks, ignore_index=True)
    if stop_after_entities is not None and not reached_eof:
        bullet(f"Early-stop reached after {len(set(df['EntityNumber'])):,} entities.")

    null_nace = df["NaceCode"].isna().sum()
    if null_nace:
        warn(f"{null_nace:,} rows with missing NaceCode skipped.")
    bullet(f"NIS2 companies found: {len(df):,}")
    return df, reached_eof

def _sniff_csv_format(path: Path) -> Tuple[str, str]:
    """Detect CSV encoding and delimiter."""
    enc = "utf-8"
    try:
        with open(path, encoding="utf-8", errors="strict") as f:
            first = f.readline()
    except UnicodeDecodeError:
        enc = "latin-1"
        with open(path, encoding="latin-1") as f:
            first = f.readline()
        warn("UTF-8 decode failed – using latin-1")
    delim = ";" if first.count(";") > first.count(",") else ","
    return enc, delim

def load_websites_for_entities(
    contact_file: str,
    entity_numbers: Optional[set] = None,
    stop_after: Optional[int] = None,
    chunksize: int = 200_000,
) -> dict:
    """
    Stream contact.csv and return first WEB URL per entity.

    If `entity_numbers` is given, only those entities are considered.
    If `stop_after` is set, stop once enough entities were collected.
    """
    p = Path(contact_file)
    if not p.exists():
        error(f"File not found: {contact_file}")
        bullet("Download: https://kbopub.economie.fgov.be/kbo-open-data/")
        sys.exit(1)
    if p.stat().st_size == 0:
        error(f"File is empty: {contact_file}")
        sys.exit(1)

    target_set = set(entity_numbers) if entity_numbers else None
    target_goal = len(target_set) if target_set is not None else None
    if stop_after is not None:
        stop_after = max(1, int(stop_after))
        target_goal = min(target_goal, stop_after) if target_goal else stop_after

    enc, delim = _sniff_csv_format(p)
    info(f"Loading contact file: {contact_file}")
    bullet(f"Delimiter: '{delim}'  |  Encoding: {enc}")
    if target_set is not None:
        bullet(f"Entity filter: {len(target_set):,} target entities")
    if target_goal:
        bullet(f"Early-stop goal: first {target_goal:,} entities with WEB URL")

    websites: dict = {}
    rows_read = 0
    start = time.time()
    try:
        for chunk in pd.read_csv(
            p,
            dtype=str,
            encoding=enc,
            sep=delim,
            chunksize=chunksize,
            low_memory=False,
            on_bad_lines="warn",
            usecols=["EntityNumber", "ContactType", "Value"],
        ):
            rows_read += len(chunk)
            chunk.columns = chunk.columns.str.strip().str.strip('"').str.strip("'")
            chunk = chunk[
                (chunk["ContactType"].str.strip().str.upper() == "WEB")
                & chunk["Value"].notna()
            ]
            if chunk.empty:
                continue

            chunk["EntityNumber"] = chunk["EntityNumber"].astype(str).str.strip()
            chunk["Value"] = chunk["Value"].astype(str).str.strip()
            chunk = chunk[chunk["Value"] != ""]
            if target_set is not None:
                chunk = chunk[chunk["EntityNumber"].isin(target_set)]
            if chunk.empty:
                continue

            for ent, val in zip(chunk["EntityNumber"], chunk["Value"]):
                if ent and ent not in websites:
                    websites[ent] = val

            if target_goal and len(websites) >= target_goal:
                break
    except pd.errors.ParserError as e:
        error(f"CSV parse error in {contact_file}: {e}")
        sys.exit(1)
    except OSError as e:
        error(f"I/O error reading {contact_file}: {e}")
        sys.exit(1)

    elapsed = time.time() - start
    ok(f"WEB entries loaded: {len(websites):,} entities "
       f"(rows scanned: {rows_read:,}  |  {elapsed:.1f}s)")
    return websites

def load_websites_all(contact_file: str) -> dict:
    """Backwards-compatible wrapper: load all WEB entries."""
    return load_websites_for_entities(contact_file)

def load_denominations(denomination_file: str,
    entity_numbers) -> dict:
    p = Path(denomination_file)
    if not p.exists():
        warn(f"denomination.csv not found: {denomination_file} – names omitted.")
        bullet("Download from: https://kbopub.economie.fgov.be/kbo-open-data/")
        return {}
    ent_set = set(entity_numbers)
    def filt(chunk):
        return chunk[chunk["EntityNumber"].str.strip().isin(ent_set)]
    df = _load_csv(denomination_file,
    {"EntityNumber", "TypeOfDenomination", "Denomination"},
    "denomination file", filter_fn=filt)
    df["EntityNumber"] = df["EntityNumber"].str.strip()
    official = df[df["TypeOfDenomination"].str.strip() == "001"]
    names    = official.groupby("EntityNumber")["Denomination"].first().to_dict()
    fallback = df.groupby("EntityNumber")["Denomination"].first().to_dict()
    for ent in ent_set:
        if ent not in names and ent in fallback:
            names[ent] = fallback[ent]
    ok(f"Company names loaded: {len(names):,}")
    return names

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# URL helpers

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

URL_REJECT_REASONS: Counter = Counter()

def normalise_url(raw) -> Optional[str]:
    if not raw:
        return None
    raw = str(raw).strip().rstrip("/")
    if len(raw) < 5:
        URL_REJECT_REASONS["too short"] += 1; return None
    if re.search(r"\s", raw):
        URL_REJECT_REASONS["whitespace"] += 1; return None
    if "." not in raw:
        URL_REJECT_REASONS["no dot"] += 1; return None
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    if re.search(r"^https?://(n/?a|onbekend|inconnu|unknown)\b", raw, re.I):
        URL_REJECT_REASONS["placeholder"] += 1
        return None
    return raw

def _resolve_one_url(raw_url: str) -> Tuple[str, str, bool]:
    for scheme in ("https://", "http://"):
        url = re.sub(r"^https?://", scheme, raw_url)
        try:
            req = urllib.request.Request(url, method="HEAD")
            req.add_header("User-Agent",
            "Mozilla/5.0 (compatible; NIS2-Scanner/1.0)")
            with urllib.request.urlopen(req, timeout=URL_TIMEOUT) as r:
                return raw_url, r.url.rstrip("/"), True
        except urllib.error.HTTPError as e:
            if e.code < 500:
                return raw_url, url.rstrip("/"), True
        except Exception:
            continue
    return raw_url, raw_url, False

def resolve_urls_batch(urls: List[str]) -> Tuple[List[str], int]:
    info(f"Resolving {len(urls):,} URLs "
    f"({URL_WORKERS} threads, {URL_TIMEOUT}s each) …")
    resolved, dead = [], 0
    try:
        with concurrent.futures.ThreadPoolExecutor(
        max_workers=URL_WORKERS) as ex:
            futures = {ex.submit(_resolve_one_url, u): u for u in urls}
            for f in _pbar(concurrent.futures.as_completed(futures),
            total=len(futures), unit="url",
            desc="HTTP/S resolve"):
                try:
                    _, final, success = f.result()
                    if success: resolved.append(final)
                    else: dead += 1
                except Exception as e:
                    warn(f"URL resolution task error: {e}")
                    dead += 1
    except KeyboardInterrupt:
        warn("URL resolution interrupted – using partial results.")
        ok(f"Resolved: {len(resolved):,}   Unreachable: {dead:,}")
        return resolved, dead
    ok(f"Resolved: {len(resolved):,}   Unreachable: {dead:,}")
    return sorted(set(resolved)), dead

def preflight_check(urls: List[str], timeout: int = 3) -> Tuple[List[str], List[str]]:
    """
    Quick HTTP HEAD reachability check before handing targets to nuclei.
    Drops any host that doesn't respond within `timeout` seconds.
    Runs automatically when –limit is set; keeps the scan from blocking
    on dead targets.
    """
    if not urls:
        return urls, []
    info(f"Preflight reachability check  ({len(urls):,} targets, "
    f"{min(URL_WORKERS, len(urls))} threads, {timeout}s timeout) …")

    def _check(url: str) -> Tuple[str, bool]:
        for scheme in ("https://", "http://"):
            target = re.sub(r"^https?://", scheme, url)
            try:
                req = urllib.request.Request(target, method="HEAD")
                req.add_header("User-Agent",
                               "Mozilla/5.0 (compatible; NIS2-Scanner/1.0)")
                with urllib.request.urlopen(req, timeout=timeout):
                    return url, True
            except urllib.error.HTTPError as e:
                if e.code < 500:          # 4xx = host alive, just refusing
                    return url, True
            except Exception:
                continue
        return url, False

    live, dead = [], 0
    dead_urls: List[str] = []
    try:
        with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(URL_WORKERS, len(urls))) as ex:
            futures = {ex.submit(_check, u): u for u in urls}
            for fut in _pbar(concurrent.futures.as_completed(futures),
                            total=len(futures), unit="host",
                            desc="Preflight"):
                try:
                    url, ok_flag = fut.result()
                    if ok_flag:
                        live.append(url)
                    else:
                        dead += 1
                        dead_urls.append(url)
                        bullet(f"  Dead target dropped: {url}")
                except Exception as e:
                    warn(f"Preflight task error: {e}")
                    dead += 1
                    failed_url = futures.get(fut)
                    if failed_url:
                        dead_urls.append(failed_url)
    except KeyboardInterrupt:
        warn("Preflight interrupted – using partial results.")
    ok(f"Preflight: {len(live):,} live  |  {dead:,} dead/unreachable (dropped)")
    return live, dead_urls

@functools.lru_cache(maxsize=8192)
def root_domain(url: str) -> str:
    """Cached root-domain extraction."""
    host = urlparse(url).hostname or url
    if HAS_TLDEXTRACT:
        ext = tldextract.extract(host)
        return f"{ext.domain}.{ext.suffix}".lower() if ext.suffix else host.lower()
    parts = host.lower().split(".")
    return ".".join(parts[-2:]) if len(parts) >= 2 else host.lower()

def dedup_by_root_domain(urls: List[str]) -> List[str]:
    seen, result, dupes = set(), [], 0
    for url in urls:
        rd = root_domain(url)
        if rd in seen:
            dupes += 1
        else:
            seen.add(rd); result.append(url)
    if dupes:
        bullet(f"Deduped {dupes:,} URLs → {len(result):,} unique root domains")
    return result

def load_exclude_list(path: str) -> set:
    p = Path(path)
    if not p.exists():
        warn(f"Exclude list not found: {path} – skipping exclusions.")
        return set()
    excludes: set = set()
    try:
        with open(p) as f:
            for line in f:
                line = line.strip().lower()
                if line and not line.startswith("#"):
                    excludes.add(line)
    except OSError as e:
        warn(f"Could not read exclude list {path}: {e}")
        return excludes
    ok(f"Exclude list loaded: {len(excludes):,} entries")
    return excludes

def apply_excludes(urls: List[str], excludes: set) -> List[str]:
    before = len(urls)
    result = [u for u in urls
    if root_domain(u) not in excludes
    and urlparse(u).hostname not in excludes]
    removed = before - len(result)
    if removed:
        bullet(f"Excluded {removed:,} URLs → {len(result):,} remaining")
    return result

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# DNS pre-resolution  (thread-safe; no global socket timeout mutation)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _resolve_one_dns(url: str) -> Tuple[str, bool, Optional[str]]:
    """
    Per-call DNS check using a daemon thread + join timeout.
    Avoids socket.setdefaulttimeout() which mutates global state
    and causes races when called from many threads simultaneously.
    """
    host = urlparse(url).hostname
    if not host:
        return url, False, "no hostname"

    result:     List = [None]
    exc_holder: List = [None]

    def _lookup():
        try:
            result[0] = socket.getaddrinfo(host, None)
        except Exception as e:
            exc_holder[0] = e

    t = threading.Thread(target=_lookup, daemon=True)
    t.start()
    t.join(DNS_TIMEOUT)
    if t.is_alive():
        return url, False, "DNS timeout"
    if exc_holder[0]:
        return url, False, str(exc_holder[0])
    return url, True, None

def filter_resolvable(urls: List[str]) -> List[str]:
    info(f"Pre-resolving {len(urls):,} domains "
    f"({DNS_WORKERS} threads, {DNS_TIMEOUT}s each) …")
    resolved, errs = [], Counter()
    try:
        with concurrent.futures.ThreadPoolExecutor(
        max_workers=DNS_WORKERS) as ex:
            futures = {ex.submit(_resolve_one_dns, u): u for u in urls}
            for fut in _pbar(concurrent.futures.as_completed(futures),
            total=len(futures), unit="host",
            desc="DNS resolving"):
                try:
                    url, ok_flag, reason = fut.result()
                    if ok_flag:
                        resolved.append(url)
                    else:
                        errs[reason or "unknown"] += 1
                except Exception as e:
                    errs[str(e)] += 1
    except KeyboardInterrupt:
        warn("DNS resolution interrupted – using partial results.")
    ok(f"Resolvable: {len(resolved):,}   Dead: {len(urls)-len(resolved):,}")
    for reason, cnt in errs.most_common(5):
        bullet(f"  {cnt:>5}  {reason}")
    return resolved

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Checkpoint

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_checkpoint(output_dir: str) -> set:
    path = Path(output_dir) / CHECKPOINT_FILE
    if not path.exists():
        return set()
    try:
        with open(path) as f:
            data = json.load(f)
            scanned = set(data.get("scanned_urls", []))
            ok(f"Checkpoint: {len(scanned):,} previously scanned "
            f"(last: {data.get('timestamp', '')})")
            return scanned
    except (json.JSONDecodeError, OSError) as e:
        warn(f"Checkpoint unreadable ({e}) – starting fresh.")
        return set()

def load_subdir_targets(output_dir: str) -> set:
    """
    Load URLs from nested */targets.txt files under output_dir.

    Used as an additional skip source so re-runs can avoid re-targeting URLs
    already materialized in per-sector subdirectories, including dry-run flows.
    """
    root = Path(output_dir)
    if not root.exists():
        return set()

    skip_dirs = {
        ".git", ".hg", ".svn", "__pycache__", ".pytest_cache", ".mypy_cache",
        ".venv", "venv", "node_modules",
    }
    root_targets = root / "targets.txt"
    try:
        root_targets_resolved = root_targets.resolve()
    except OSError:
        root_targets_resolved = root_targets

    source_files = 0
    urls: set = set()
    for cur_root, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs if d not in skip_dirs and not d.startswith(".")]
        if "targets.txt" not in files:
            continue
        candidate = Path(cur_root) / "targets.txt"
        try:
            candidate_resolved = candidate.resolve()
        except OSError:
            candidate_resolved = candidate
        if candidate_resolved == root_targets_resolved:
            continue

        source_files += 1
        try:
            with open(candidate, encoding="utf-8") as f:
                for line in f:
                    raw = line.strip()
                    if not raw:
                        continue
                    norm = normalise_url(raw)
                    if norm:
                        urls.add(norm)
        except OSError as e:
            warn(f"Could not read subdir targets file '{candidate}': {e}")

    if urls:
        ok(f"Subdir target cache: {len(urls):,} URLs from {source_files:,} target file(s)")
    return urls

def load_dead_targets(output_dir: str) -> set:
    path = Path(output_dir) / DEAD_TARGETS_FILE
    if not path.exists():
        return set()
    try:
        with open(path, encoding="utf-8") as f:
            dead = {line.strip() for line in f if line.strip()}
            if dead:
                ok(f"Dead-target cache: {len(dead):,} known unreachable targets")
            return dead
    except OSError as e:
        warn(f"Could not read dead-target cache: {e}")
        return set()

def save_dead_targets(output_dir: str, dead_urls: set) -> None:
    path = Path(output_dir) / DEAD_TARGETS_FILE
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(sorted(dead_urls)))
    except OSError as e:
        warn(f"Could not write dead-target cache: {e}")

def save_checkpoint(output_dir: str, urls) -> None:
    path = Path(output_dir) / CHECKPOINT_FILE
    tmp  = path.with_suffix(".tmp")
    try:
        with open(tmp, "w") as f:
            json.dump({"timestamp": datetime.now().isoformat(timespec="seconds"),
            "scanned_urls": list(urls)}, f, indent=2)
            tmp.replace(path)
            ok(f"Checkpoint saved: {len(urls):,} URLs → {path}")
    except OSError as e:
        warn(f"Could not save checkpoint: {e}")
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass

def clear_checkpoint(output_dir: str) -> None:
    path = Path(output_dir) / CHECKPOINT_FILE
    if path.exists():
        try:
            path.unlink(); ok("Checkpoint cleared.")
        except OSError as e:
            warn(f"Could not clear checkpoint: {e}")

def apply_resume(urls: List[str], already_scanned: set) -> List[str]:
    before  = len(urls)
    urls    = [u for u in urls if u not in already_scanned]
    skipped = before - len(urls)
    if skipped:
        bullet(f"Resume: skipping {skipped:,} already-scanned "
        f"→ {len(urls):,} remaining")
    return urls

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# URL → Company lookup

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_url_company_lookup(
    nis2_df, websites: dict, denominations: dict
    ) -> Tuple[dict, dict]:
    """
    Returns (lookup, hostname_index).
    lookup         : {normalised_url_no_slash: record}
    hostname_index : {hostname: record}   — used for O(1) fallback matching.
    """
    lookup:         dict = {}
    hostname_index: dict = {}

    ents  = nis2_df["EntityNumber"].str.strip()
    naces = nis2_df["NaceCode"].str.strip()
    sects = nis2_df["NIS2_Sector"]

    for ent, nace, sector in zip(ents, naces, sects):
        raw_w = websites.get(ent, "")
        url   = normalise_url(raw_w) if raw_w else None
        if not url:
            continue
        record = {
            "entity":  ent,
            "name":    denominations.get(ent, ""),
            "nace":    str(nace),
            "sector":  str(sector),
            "kbo_url": kbo_url(ent),
        }
        key = url.rstrip("/")
        lookup[key] = record
        host = (urlparse(key).hostname or "").lower()
        if host:
            # Index the exact hostname plus www/non-www variants and the root
            # domain, so a finding/scan URL that was redirected (e.g. to www.* or
            # a deeper path) still resolves back to the originating KBO company.
            hostname_index.setdefault(host, record)
            bare = host[4:] if host.startswith("www.") else host
            hostname_index.setdefault(bare, record)
            hostname_index.setdefault("www." + bare, record)
            rd = root_domain(key)
            if rd:
                hostname_index.setdefault(rd, record)

    return lookup, hostname_index

def save_url_lookup(lookup: dict, hostname_index: dict,
    output_dir: str) -> None:
    path = Path(output_dir) / URL_LOOKUP_FILE
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"lookup": lookup, "hostname_index": hostname_index},
            f, ensure_ascii=False, indent=2)
            ok(f"URL→company lookup: {path}  ({len(lookup):,} entries)")
    except OSError as e:
        warn(f"Could not save URL lookup: {e}")

def load_url_lookup(output_dir: str) -> Tuple[dict, dict]:
    """Returns (lookup, hostname_index). Supports both old and new file format."""
    path = Path(output_dir) / URL_LOOKUP_FILE
    if not path.exists():
        return {}, {}
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and "lookup" in data:
                return data["lookup"], data.get("hostname_index", {})
            return data, {}
    except (json.JSONDecodeError, OSError) as e:
        warn(f"Could not load URL lookup ({e}).")
        return {}, {}

def resolve_company(host_url: str, lookup: dict,
    hostname_index: dict) -> Optional[dict]:
    """
    O(1) company resolution.

    Priority: exact URL → scheme-swapped URL → hostname → www/non-www variant
    → root domain. The fallbacks matter because scanned/finding URLs are often
    redirected (www.*, deeper paths, :443 ports) and would otherwise fail to map
    back to the KBO company that seeded them, showing up as "no company".
    """
    if not host_url:
        return None
    key = host_url.rstrip("/")
    if key in lookup:
        return lookup[key]
    if key.startswith("https://"):
        alt = "http://" + key[len("https://"):]
    elif key.startswith("http://"):
        alt = "https://" + key[len("http://"):]
    else:
        alt = None
    if alt and alt in lookup:
        return lookup[alt]

    parse_target = host_url if "://" in host_url else "https://" + host_url
    host = (urlparse(parse_target).hostname or "").lower()
    if not host:
        return None

    candidates = [host]
    if host.startswith("www."):
        candidates.append(host[4:])
    else:
        candidates.append("www." + host)
    rd = root_domain(parse_target)
    if rd and rd not in candidates:
        candidates.append(rd)

    for cand in candidates:
        record = hostname_index.get(cand)
        if record:
            return record
    return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Save outputs

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_outputs(nis2_df, websites, denominations, targets_file, manifest_file,
    excludes, resolve_dns, resolve_urls, resume_urls,
    per_sector_dirs, output_dir, limit=None, dead_urls_cache=None):

        nis2_df = nis2_df.copy()
        nis2_df["NIS2_Sector"] = nis2_df["NaceCode"].apply(get_sector_label)
        nis2_df["Website"]     = nis2_df["EntityNumber"].map(websites).fillna("")
        nis2_df["CompanyName"] = nis2_df["EntityNumber"].map(denominations).fillna("")
        nis2_df["KBO_URL"]     = nis2_df["EntityNumber"].apply(kbo_url)

        sector_counts = Counter(nis2_df["NIS2_Sector"].tolist())

        urls = list(dict.fromkeys(
            u for u in (normalise_url(websites.get(e, ""))
                        for e in nis2_df["EntityNumber"]) if u
        ))

        if URL_REJECT_REASONS:
            warn(f"Rejected {sum(URL_REJECT_REASONS.values()):,} malformed URLs:")
            for reason, cnt in URL_REJECT_REASONS.most_common():
                bullet(f"  {cnt:>5}  {reason}")

        header("STEP 5b – DEDUPLICATION")
        urls = dedup_by_root_domain(urls)

        if excludes:
            header("STEP 5c – EXCLUDE LIST")
            urls = apply_excludes(urls, excludes)

        if resolve_dns:
            header("STEP 5d – DNS PRE-RESOLUTION")
            urls = filter_resolvable(urls)

        if resolve_urls:
            header("STEP 5e – HTTP/HTTPS NORMALISATION")
            urls, _ = resolve_urls_batch(urls)

        if resume_urls:
            header("STEP 5f – RESUME FILTER")
            urls = apply_resume(urls, resume_urls)

        if dead_urls_cache:
            before_dead = len(urls)
            urls = [u for u in urls if u not in dead_urls_cache]
            skipped_dead = before_dead - len(urls)
            if skipped_dead:
                bullet(f"Skipping {skipped_dead:,} known dead target(s) from cache")

        # Preflight: auto-run when --limit is set so dead targets are dropped
        # before nuclei gets them. Always 3s timeout — fast enough to check
        # a small candidate pool without noticeably slowing down large scans.
        if limit:
            header("STEP 5f.5 – PREFLIGHT REACHABILITY CHECK")
            urls, dead_urls = preflight_check(urls, timeout=3)
            if dead_urls and dead_urls_cache is not None:
                before_cache = len(dead_urls_cache)
                dead_urls_cache.update(dead_urls)
                added = len(dead_urls_cache) - before_cache
                if added:
                    save_dead_targets(output_dir, dead_urls_cache)
                    bullet(f"Dead-target cache updated: +{added:,} "
                           f"(total {len(dead_urls_cache):,})")

        if limit and len(urls) > limit:
            urls = urls[:limit]
            ok(f"After --limit {limit}: {len(urls):,} live targets")

        urls = sorted(set(urls))

        # Write manifest scoped to the final URL set so --contact-limit respects
        # --limit. Writing before the limit slice caused enrichment to see all
        # prefetch candidates (limit×10) instead of the actual scanned targets.
        final_url_set = set(urls)
        final_entities = {
            e for e in nis2_df["EntityNumber"].str.strip()
            if normalise_url(websites.get(e, "")) in final_url_set
        }
        manifest_df = nis2_df[nis2_df["EntityNumber"].str.strip().isin(final_entities)]
        try:
            manifest_df[["EntityNumber", "CompanyName", "NaceCode",
                         "NIS2_Sector", "Website", "KBO_URL"]
                       ].to_csv(manifest_file, index=False)
            ok(f"Manifest: {manifest_file}  ({len(manifest_df):,} companies)")
        except OSError as e:
            warn(f"Could not write manifest: {e}")

        try:
            with open(targets_file, "w") as f:
                f.write("\n".join(urls))
            ok(f"Targets: {targets_file}  ({len(urls):,} URLs)")
        except OSError as e:
            error(f"Could not write targets file: {e}"); sys.exit(1)

        lookup, hostname_index = build_url_company_lookup(
            nis2_df, websites, denominations)
        save_url_lookup(lookup, hostname_index, output_dir)

        sector_files: dict = {}
        if per_sector_dirs:
            header("STEP 5g – PER-SECTOR TARGET FILES")
            url_set    = set(urls)
            ent_to_url = {e: normalise_url(w) for e, w in websites.items()}
            for sector in nis2_df["NIS2_Sector"].unique():
                s_df   = nis2_df[nis2_df["NIS2_Sector"] == sector]
                s_urls = sorted(set(
                    u for e in s_df["EntityNumber"]
                    for u in [ent_to_url.get(e.strip())]
                    if u and u in url_set
                ))
                if not s_urls:
                    continue
                safe  = re.sub(r'[^\w\-]', '_', sector)
                s_dir = Path(output_dir) / safe
                try:
                    s_dir.mkdir(parents=True, exist_ok=True)
                    s_file = s_dir / "targets.txt"
                    with open(s_file, "w") as f:
                        f.write("\n".join(s_urls))
                    sector_files[sector] = str(s_file)
                    bullet(f"  {sector:<40} {len(s_urls):>5} URLs → {s_file}")
                except OSError as e:
                    warn(f"Could not write sector file for '{sector}': {e}")

        return len(urls), sector_counts, urls, sector_files

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Nuclei

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# FIX 3: guard against empty/missing nuclei version output to prevent IndexError

def check_nuclei() -> bool:
    try:
        r   = subprocess.run(["nuclei", "-version"],
        capture_output=True, text=True, timeout=10)
        lines = (r.stdout + r.stderr).strip().splitlines()
        ver   = lines[0] if lines else "(unknown version)"
        ok(f"nuclei: {ver}"); return True
    except FileNotFoundError:
        error("nuclei not found.")
        bullet("Place nuclei.exe next to this script, or add to PATH.")
        return False
    except subprocess.TimeoutExpired:
        error("nuclei -version timed out."); return False

def update_nuclei_templates() -> None:
    info("Updating nuclei templates …")
    try:
        r = subprocess.run(["nuclei", "-update-templates"],
        capture_output=True, text=True, timeout=120)
        for line in (r.stdout + r.stderr).strip().splitlines()[-5:]:
            bullet(line)
        if r.returncode == 0:
            ok("Templates updated.")
        else:
            warn(f"nuclei -update-templates exited with code {r.returncode}.")
    except subprocess.TimeoutExpired:
        warn("Template update timed out.")
    except FileNotFoundError:
        error("nuclei not found."); sys.exit(1)

def build_nuclei_cmd(targets_file, output_file, templates,
    rate, concur, timeout, severity, proxy,
    verbose: bool = False) -> List[str]:
    resolved = []
    for t in templates:
        fpath = resolve_template_path(t)
        resolved.append(str(fpath.resolve()) if fpath else t)

    cmd = ["nuclei", "-l", targets_file, "-o", output_file, "-jsonl",
           "-rate-limit", str(rate), "-concurrency", str(concur),
           "-timeout", str(timeout), "-severity", severity,
           "-stats", "-no-color",
           "-max-host-error", "3",       # skip host after 3 consecutive errors
           "-disable-update-check",      # never hang on startup version check
           "-retries", "0"]             # no retries — fail fast on dead hosts
    if verbose:
        cmd.append("-v")
    for rt in resolved:
        cmd += ["-t", rt]
    if proxy:
        cmd += ["-proxy", proxy]
    return cmd

def run_nuclei(targets_file, output_file, templates,
    rate, concur, timeout, severity, proxy,
    verbose: bool = False) -> int:
    cmd = build_nuclei_cmd(targets_file, output_file, templates,
    rate, concur, timeout, severity, proxy,
    verbose=verbose)
    info(f"Templates ({len(templates)}): {', '.join(templates)}")
    info(f"Command: {' '.join(cmd)}\n")
    print("─" * 68)
    start = datetime.now()
    rc    = 1
    try:
        proc = subprocess.Popen(cmd,
        stdout=sys.stdout,
        stderr=sys.stderr)
        proc.wait()
        rc = proc.returncode if proc.returncode is not None else 1
    except FileNotFoundError:
        error("nuclei binary not found."); rc = 1
    except KeyboardInterrupt:
        try:
            proc.terminate()
            proc.wait(timeout=5)
        except Exception:
            pass
        warn("Interrupted by user."); rc = 130
    except OSError as e:
        error(f"Could not launch nuclei: {e}"); rc = 1
    finally:
        elapsed = datetime.now() - start
        print("─" * 68)
        if rc == 0:
            ok(f"Nuclei finished in {elapsed}.")
        elif rc == 130:
            warn(f"Nuclei interrupted after {elapsed}.")
        else:
            warn(f"Nuclei exited with code {rc} after {elapsed}.")
            rf = Path(output_file)
            if rf.exists():
                errs = []
                try:
                    with open(rf, encoding="utf-8", errors="replace") as f:
                        for line in f:
                            try:
                                obj = _json_loads(line.strip())
                                if obj.get("type") == "error":
                                    errs.append(
                                    obj.get("error", "(no message)"))
                            except Exception:
                                pass
                except OSError:
                    pass
                if errs:
                    warn(f"nuclei reported {len(errs)} error(s); first 5:")
                    for e in errs[:5]: bullet(e)
    return rc

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Retry

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def find_failed_targets(nuclei_output: str) -> List[str]:
    p = Path(nuclei_output)
    if not p.exists() or p.stat().st_size == 0:
        return []
    failed: set = set()
    try:
        with open(p, encoding="utf-8", errors="replace") as f:
            for line in f:
                try:
                    entry = _json_loads(line.strip())
                    if entry.get("type") == "error":
                        host = entry.get("host") or entry.get("input", "")
                        if host: failed.add(host.rstrip("/"))
                except Exception:
                    continue
    except OSError as e:
        warn(f"Could not read nuclei output for retry analysis: {e}")
    return list(failed)

def retry_failed_targets(output_dir, nuclei_output, templates,
    rate, concur, timeout, severity, proxy) -> None:
    header("RETRY – FAILED / TIMED-OUT TARGETS")
    failed = find_failed_targets(nuclei_output)
    if not failed:
        info("No failed targets found."); return
    retry_file    = Path(output_dir) / RETRY_FILE
    retry_output  = Path(output_dir) / "nuclei_retry_results.json"
    retry_timeout = min(timeout * 3, 60)
    try:
        with open(retry_file, "w") as f:
            f.write("\n".join(failed))
    except OSError as e:
        warn(f"Could not write retry file: {e}"); return
    warn(f"{len(failed):,} failed targets → retrying "
         f"(timeout={retry_timeout}s)")
    run_nuclei(str(retry_file), str(retry_output), templates,
               max(rate // 2, 5), max(concur // 2, 1),
               retry_timeout, severity, proxy)
    ok("Retry complete.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Scheduler

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# FIX 5: guard against batch_size <= 0 to prevent ValueError in range()

def run_scheduled(urls, output_dir, nuclei_output, templates,
    rate, concur, timeout, severity, proxy,
    batch_size, no_retry, already_scanned) -> None:
    if batch_size <= 0:
        warn(f"–schedule value must be > 0 (got {batch_size}); defaulting to 1.")
        batch_size = 1
    total   = len(urls)
    batches = [urls[i:i + batch_size] for i in range(0, total, batch_size)]
    info(f"Scheduled: {total:,} targets in "
         f"{len(batches)} batch(es) of {batch_size}/day")
    all_scanned = set(already_scanned)
    for idx, batch in enumerate(batches, 1):
        header(f"BATCH {idx}/{len(batches)}  ({len(batch):,} targets  |  "
               f"{datetime.now().strftime('%Y-%m-%d %H:%M')})")
        batch_file   = Path(output_dir) / f"batch_{idx:03d}_targets.txt"
        batch_output = Path(output_dir) / f"batch_{idx:03d}_results.json"
        try:
            with open(batch_file, "w") as f:
                f.write("\n".join(batch))
        except OSError as e:
            warn(f"Could not write batch file: {e}"); continue
        run_nuclei(str(batch_file), str(batch_output), templates,
                   rate, concur, timeout, severity, proxy)
        all_scanned.update(batch)
        save_checkpoint(output_dir, list(all_scanned))
        if not no_retry:
            retry_failed_targets(output_dir, str(batch_output), templates,
                                 rate, concur, timeout, severity, proxy)
        if idx < len(batches):
            wake = datetime.now() + timedelta(hours=24)
            ok(f"Next batch at {wake.strftime('%Y-%m-%d %H:%M')} "
               f"– sleeping 24h …")
            try:
                time.sleep(86400)
            except KeyboardInterrupt:
                warn("Scheduler interrupted."); break
    ok(f"Scheduled run complete. "
       f"{len(all_scanned):,} total targets processed.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Coverage matrix

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_coverage_matrix(findings: List[dict],
    scanned_hosts: List[str],
    template_checks: Dict[str, dict]
    ) -> Tuple[dict, List[dict]]:
    """
    Returns (matrix, all_checks).

    matrix[host_no_slash][check_key] = {
    "status":   "FINDING" | "CLEAN",
    "severity": str,
    "template": str,
    }
    all_checks = [{"key", "template", "name", "severity"}, ...]
    """
    all_checks: List[dict] = []
    seen_keys: set = set()
    for tid, tdata in template_checks.items():
        for ch in tdata["checks"]:
            key = f"{tid}/{ch['name']}"
            if key not in seen_keys:
                all_checks.append({"key": key, "template": tid,
                                   "name": ch["name"],
                                   "severity": ch["severity"]})
                seen_keys.add(key)

    fired: Dict[str, dict] = defaultdict(dict)
    for finding in findings:
        host  = _finding_host(finding)
        if not host:
            continue
        tid   = finding.get("template-id", "")
        mname = finding.get("matcher-name", "")
        sev   = finding.get("info", {}).get("severity", "unknown").lower()
        key   = f"{tid}/{mname}" if mname else f"{tid}/(no-matcher-name)"
        fired[host][key] = {"severity": sev, "template": tid}
        if key not in seen_keys:
            all_checks.append({"key": key, "template": tid,
                               "name": mname or "(no-matcher-name)",
                               "severity": sev})
            seen_keys.add(key)

    matrix: dict = {}
    host_order = list(dict.fromkeys(
        [h.rstrip("/") for h in scanned_hosts if h] +
        [h for h in fired.keys() if h]
    ))
    for host in host_order:
        h = host.rstrip("/")
        matrix[h] = {}
        for ch in all_checks:
            if ch["key"] in fired[h]:
                matrix[h][ch["key"]] = {
                    "status":   "FINDING",
                    "severity": fired[h][ch["key"]]["severity"],
                    "template": ch["template"],
                }
            else:
                matrix[h][ch["key"]] = {
                    "status":   "CLEAN",
                    "severity": ch["severity"],
                    "template": ch["template"],
                }
    return matrix, all_checks

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Coverage export (CSV + optional XLSX)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_coverage_csv(matrix: dict, all_checks: List[dict],
    lookup: dict, hostname_index: dict,
    output_dir: str) -> Path:
    path = Path(output_dir) / COVERAGE_CSV
    try:
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Host", "Company", "Entity", "NACE", "Sector",
            "KBO_URL", "Template", "Check", "Severity", "Status"])
            for host, checks in sorted(matrix.items()):
                co = resolve_company(host, lookup, hostname_index) or {}
                for ch in all_checks:
                    cell = checks.get(ch["key"],
                    {"status":   "CLEAN",
                    "severity": ch["severity"],
                    "template": ch["template"]})
                    w.writerow([host,
                    co.get("name", ""), co.get("entity", ""),
                    co.get("nace", ""), co.get("sector", ""),
                    co.get("kbo_url", ""),
                    ch["template"], ch["name"],
                    cell["severity"], cell["status"]])
        ok(f"Coverage CSV: {path}  "
           f"({len(matrix):,} hosts × {len(all_checks):,} checks)")
    except OSError as e:
        warn(f"Could not write coverage CSV: {e}")
    return path

    # FIX 7: use only local openpyxl imports inside the function; the module-level

    # `import openpyxl` may not have succeeded if HAS_OPENPYXL is False.

def save_coverage_xlsx(matrix: dict, all_checks: List[dict],
    lookup: dict, hostname_index: dict,
    output_dir: str) -> None:
    if not HAS_OPENPYXL:
        warn("openpyxl not installed – skipping XLSX export. "
        "Run: pip install openpyxl")
        return

    import openpyxl as _xl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    RED   = PatternFill("solid", fgColor="FFCCCC")
    GREEN = PatternFill("solid", fgColor="CCFFCC")
    GREY  = PatternFill("solid", fgColor="EEEEEE")
    BOLD  = Font(bold=True)

    wb  = _xl.Workbook()

    # ── Sheet 1: Full coverage ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Coverage"
    hdrs = ["Host", "Company", "Entity", "NACE", "Sector", "KBO_URL",
            "Template", "Check", "Severity", "Status"]
    ws1.append(hdrs)
    for cell in ws1[1]:
        cell.font = BOLD; cell.fill = GREY

    for host, checks in sorted(matrix.items()):
        co = resolve_company(host, lookup, hostname_index) or {}
        for ch in all_checks:
            cell = checks.get(ch["key"],
                              {"status":   "CLEAN",
                               "severity": ch["severity"],
                               "template": ch["template"]})
            ws1.append([host,
                        co.get("name", ""), co.get("entity", ""),
                        co.get("nace", ""), co.get("sector", ""),
                        co.get("kbo_url", ""),
                        ch["template"], ch["name"],
                        cell["severity"], cell["status"]])
            fill = RED if cell["status"] == "FINDING" else GREEN
            for c in ws1[ws1.max_row]:
                c.fill = fill

    # ── Sheet 2: Check summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Check Summary")
    ws2.append(["Template", "Check", "Severity",
                "Findings", "Clean", "% Clean"])
    for c in ws2[1]:
        c.font = BOLD; c.fill = GREY

    fired_by_check: Counter = Counter()
    for hm in matrix.values():
        for key, v in hm.items():
            if v["status"] == "FINDING":
                fired_by_check[key] += 1

    n_hosts = len(matrix) or 1
    for ch in sorted(all_checks,
                     key=lambda x: -fired_by_check.get(x["key"], 0)):
        f_cnt = fired_by_check.get(ch["key"], 0)
        c_cnt = max(n_hosts - f_cnt, 0)
        pct   = round(c_cnt / n_hosts * 100, 1)
        ws2.append([ch["template"], ch["name"], ch["severity"],
                    f_cnt, c_cnt, pct])
        fill = RED if f_cnt else GREEN
        for c in ws2[ws2.max_row]:
            c.fill = fill

    # ── Auto-width all sheets ─────────────────────────────────────────
    for ws in (ws1, ws2):
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_len + 2, 60)

    path = Path(output_dir) / COVERAGE_CSV.replace(".csv", ".xlsx")
    try:
        wb.save(path)
        ok(f"Coverage XLSX: {path}")
    except OSError as e:
        warn(f"Could not write XLSX: {e}")

def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value)

def _html(value) -> str:
    return html_lib.escape(_safe_text(value), quote=True)

def _finding_host(finding: dict) -> str:
    """Return the URL/host key used to attach a nuclei finding to report rows."""
    return _safe_text(
        finding.get("host")
        or finding.get("matched-at")
        or finding.get("url")
        or finding.get("input")
    ).rstrip("/")

def _finding_check_key(finding: dict) -> str:
    tid   = finding.get("template-id", "")
    mname = finding.get("matcher-name", "")
    return f"{tid}/{mname}" if mname else f"{tid}/(no-matcher-name)"

def _finding_evidence(finding: dict) -> str:
    parts: List[str] = []
    matched_at = (finding.get("matched-at")
                  or finding.get("url")
                  or finding.get("host", ""))
    if matched_at:
        parts.append(f"Matched at: {_safe_text(matched_at)}")

    extracted = (finding.get("extracted-results")
                 or finding.get("extracted_results")
                 or finding.get("extractor-results")
                 or [])
    if isinstance(extracted, list):
        observed = "; ".join(_safe_text(x) for x in extracted[:5] if x is not None)
    else:
        observed = _safe_text(extracted)
    if observed:
        parts.append(f"Observed: {observed}")

    matcher = finding.get("matcher-name") or finding.get("template-id")
    if matcher:
        parts.append(f"Triggered check: {_safe_text(matcher)}")
    ip_addr = finding.get("ip")
    if ip_addr:
        parts.append(f"IP: {_safe_text(ip_addr)}")
    return " | ".join(parts)

def _finding_lookup(table: Dict[str, str], check: str, template: str = "") -> str:
    """Resolve a per-finding description/remediation.

    Tries the specific matcher-name first, then the template-id, so the report
    stays finding-specific even when a template does not name its matchers.
    """
    return table.get(check) or (table.get(template, "") if template else "")

def _finding_remediation(check: str, sev: str, template: str = "") -> str:
    action = _finding_lookup(FINDING_MATCHER_REMEDIATIONS, check, template)
    if action:
        return action
    if sev in ("critical", "high"):
        return ("Apply vendor fixes or configuration changes immediately, "
                "reduce external exposure where possible, and validate with a rescan.")
    if sev == "medium":
        return ("Plan remediation in the next maintenance cycle and verify the "
                "control with regression testing or a targeted rescan.")
    return ("Track as hardening debt, implement the recommended configuration "
            "improvement, and monitor for recurrence.")

def _finding_risk(check: str, sev: str, template: str = "") -> str:
    risk = SEVERITY_RISK_TEXT.get(sev, SEVERITY_RISK_TEXT["unknown"])
    desc = _finding_lookup(FINDING_MATCHER_DESCRIPTIONS, check, template)
    if desc:
        return f"{risk} Finding impact: {desc}"
    return risk

def _finding_exec_summary(host: str,
    check: str,
    sev: str,
    matched_at: str,
    description: str) -> str:
    target = matched_at or host or "the scanned target"
    detail = description or FINDING_MATCHER_DESCRIPTIONS.get(check) or check
    severity = sev.capitalize() if sev else "Unknown"
    return f"{severity} finding on {target}: {detail}"

def _finding_narrative(summary: str,
    evidence: str,
    risk: str,
    remediation: str) -> str:
    lines = [
        ("Executive summary", summary),
        ("Evidence", evidence),
        ("Risk", risk),
        ("Remediation", remediation),
    ]
    return "\n".join(f"{label}: {value}" for label, value in lines if value)

def _finding_summary(finding: dict) -> dict:
    info_data = finding.get("info") or {}
    check = finding.get("matcher-name") or "(no-matcher-name)"
    template = _safe_text(finding.get("template-id", ""))
    severity = _safe_text(info_data.get("severity", "unknown")).lower()
    matched_at = (finding.get("matched-at")
                  or finding.get("url")
                  or finding.get("host", ""))
    description = (_finding_lookup(FINDING_MATCHER_DESCRIPTIONS, check, template)
                   or info_data.get("description", ""))
    evidence = _finding_evidence(finding)
    executive_summary = _finding_exec_summary(
        _finding_host(finding), check, severity, matched_at, description)
    risk = _finding_risk(check, severity, template)
    remediation = _finding_remediation(check, severity, template)
    return {
        "template": finding.get("template-id", ""),
        "check": check,
        "severity": severity,
        "name": info_data.get("name", ""),
        "description": _finding_narrative(
            executive_summary, evidence, risk, remediation),
        "matched_at": matched_at,
        "evidence": evidence,
        "executive_summary": executive_summary,
        "risk": risk,
        "remediation": remediation,
    }

def _blank_company() -> dict:
    return {"name": "", "entity": "", "nace": "", "sector": "", "kbo_url": ""}

_SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3,
              "info": 4, "unknown": 5}

def _slug(text: str, fallback: str = "company") -> str:
    s = re.sub(r"[^\w\- ]", "", str(text or "")).strip().replace(" ", "_")
    s = re.sub(r"_+", "_", s)
    return s[:80] or fallback

def write_ccb_disclosure_reports(nuclei_output: str,
    output_dir: str,
    lookup: dict,
    hostname_index: dict,
    researcher: Optional[dict] = None) -> List[str]:
    """Write one CCB-style coordinated vulnerability disclosure report per company.

    Structured for good-faith reporting under the Belgian CVDP framework
    (art. 550bis §5): addressed to the CCB (CSIRT) and the organisation's
    security/abuse mailbox, factual, no commercial content, no withholding.
    Every finding the scan produced is listed in full — the safe harbour is
    conditioned on complete, prompt disclosure, so nothing is held back.
    """
    results_path = Path(nuclei_output)
    if not results_path.exists() or results_path.stat().st_size == 0:
        return []

    researcher = researcher or _sender_config()
    functional_by_domain = _load_functional_mailboxes(output_dir)
    security_by_domain = _load_security_contacts(output_dir)

    companies: Dict[str, dict] = {}
    for finding in stream_findings(results_path):
        host = _finding_host(finding)
        if not host:
            continue
        co = _company_record(host, lookup, hostname_index)
        key = co.get("entity") or co.get("name") or host
        bucket = companies.setdefault(key, {"company": co, "hosts": {}})
        bucket["hosts"].setdefault(host, []).append(_finding_summary(finding))

    if not companies:
        return []

    report_root = Path(output_dir) / "by_company"
    report_root.mkdir(parents=True, exist_ok=True)
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    written: List[str] = []

    for key, data in companies.items():
        co = data["company"]
        name = co.get("name") or key
        # Determine the security/abuse mailbox for this org's domain.
        first_host = next(iter(data["hosts"]), "")
        netloc = urlparse(first_host if "://" in first_host
                          else "http://" + first_host).netloc or first_host
        domain = netloc.replace("www.", "").strip("/")
        discovered = functional_by_domain.get(domain, [])
        sec_mailbox = (security_by_domain.get(domain)
                       or next((e for e in discovered
                                if _local_part(e) in
                                ("security", "abuse", "soc", "cert", "psirt")), "")
                       or (f"security@{domain}" if domain else "[security mailbox]"))
        org_to = sec_mailbox

        total = sum(len(v) for v in data["hosts"].values())
        lines: List[str] = []
        lines.append("=" * 74)
        lines.append("COORDINATED VULNERABILITY DISCLOSURE")
        lines.append("Belgisch kader voor gecoördineerde bekendmaking / "
                     "Belgian CVDP framework")
        lines.append("=" * 74)
        lines.append(f"Datum / Date        : {generated}")
        lines.append(f"Betrokken organisatie / Affected organisation : {name}")
        if co.get("entity"):
            lines.append(f"Ondernemingsnummer / Entity : {co['entity']}")
        lines.append("")
        lines.append("AAN / TO:")
        lines.append("  1. Centre for Cybersecurity Belgium (CCB) — CSIRT")
        lines.append("     vulnerabilityreport@ccb.belgium.be  |  cert@cert.be")
        lines.append(f"  2. {name}  <{org_to}>")
        lines.append("")
        lines.append("MELDER / REPORTER:")
        lines.append(f"  {researcher['name']}, {researcher['company']}")
        lines.append(f"  {researcher['email']}  |  {researcher['phone']}")
        lines.append("")
        lines.append("-" * 74)
        lines.append("AARD VAN DE MELDING / NATURE OF THIS REPORT")
        lines.append("-" * 74)
        lines.append(
            "Deze melding gebeurt te goeder trouw onder het Belgische kader voor")
        lines.append(
            "gecoördineerde bekendmaking van kwetsbaarheden (art. 550bis §5 Sw.).")
        lines.append(
            "Er is geen toegang genomen tot systemen of gegevens, geen data")
        lines.append(
            "gewijzigd of gekopieerd, en niet verder gegaan dan nodig om de")
        lines.append(
            "kwetsbaarheid vast te stellen. Deze bevindingen worden niet publiek")
        lines.append(
            "gemaakt en niet met derden gedeeld zonder akkoord van het CCB.")
        lines.append("")
        lines.append(
            "This is a good-faith report under the Belgian coordinated "
            "vulnerability")
        lines.append(
            "disclosure framework. No systems or data were accessed, altered or")
        lines.append(
            "copied; testing did not go beyond confirming each issue. Findings "
            "are")
        lines.append(
            "shared only with the CCB and the affected organisation.")
        lines.append("")
        lines.append(f"Aantal bevindingen / Findings : {total}  "
                     f"over {len(data['hosts'])} host(s)")
        lines.append("")

        for host in sorted(data["hosts"]):
            findings = sorted(
                data["hosts"][host],
                key=lambda f: (_SEV_ORDER.get(f.get("severity", "unknown"), 5),
                               f.get("name", "")))
            lines.append("-" * 74)
            lines.append(f"HOST: {host}")
            lines.append("-" * 74)
            for i, f in enumerate(findings, 1):
                title = f.get("name") or f.get("check") or f.get("template") or "Finding"
                lines.append(f"[{i}] {title}   ({f.get('severity', 'unknown').upper()})")
                if f.get("matched_at"):
                    lines.append(f"    Locatie / Location : {f['matched_at']}")
                if f.get("evidence"):
                    lines.append(f"    Bewijs / Evidence  : {f['evidence']}")
                if f.get("risk"):
                    lines.append(f"    Risico / Risk      : {f['risk']}")
                if f.get("remediation"):
                    lines.append(f"    Herstel / Remediation : {f['remediation']}")
                lines.append("")

        lines.append("-" * 74)
        lines.append("VOLGENDE STAPPEN / NEXT STEPS")
        lines.append("-" * 74)
        lines.append(
            "Het CCB coördineert de verdere opvolging. De organisatie wordt")
        lines.append(
            "verzocht de bevindingen te verifiëren en te herstellen. Er is geen")
        lines.append(
            "actie of betaling vereist tegenover de melder; dit rapport dient")
        lines.append(
            "uitsluitend ter beveiliging.")
        lines.append("")
        lines.append(
            "The CCB coordinates follow-up. No action toward or payment to the")
        lines.append(
            "reporter is required; this report exists solely to improve security.")

        slug = _slug(name)
        company_dir = report_root / slug
        company_dir.mkdir(parents=True, exist_ok=True)
        fpath = company_dir / f"ccb_disclosure_{slug}.txt"
        n = 2
        while fpath.exists():
            fpath = company_dir / f"ccb_disclosure_{slug}_{n}.txt"
            n += 1
        fpath.write_text("\n".join(lines), encoding="utf-8")
        written.append(str(fpath))

    if written:
        ok(f"Wrote {len(written)} CCB disclosure report(s) → {report_root}")
    return written

def write_company_findings_reports(nuclei_output: str,
    output_dir: str,
    lookup: dict,
    hostname_index: dict) -> List[str]:
    """Write one plain-text findings report per company.

    Each report lists that company's findings grouped by host and ordered by
    severity, with the risk and remediation the tool already resolves per
    finding. Internal analyst output — not addressed to any individual.
    Returns the list of file paths written.
    """
    results_path = Path(nuclei_output)
    if not results_path.exists() or results_path.stat().st_size == 0:
        return []

    # company_key -> {"company": {...}, "hosts": {host: [finding_summary, ...]}}
    companies: Dict[str, dict] = {}
    for finding in stream_findings(results_path):
        host = _finding_host(finding)
        if not host:
            continue
        co = _company_record(host, lookup, hostname_index)
        key = co.get("entity") or co.get("name") or host
        bucket = companies.setdefault(key, {"company": co, "hosts": {}})
        summ = _finding_summary(finding)
        bucket["hosts"].setdefault(host, []).append(summ)

    if not companies:
        return []

    report_root = Path(output_dir) / "by_company"
    report_root.mkdir(parents=True, exist_ok=True)
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    written: List[str] = []

    for key, data in companies.items():
        co = data["company"]
        name = co.get("name") or key
        sev_counts: Counter = Counter()
        for findings in data["hosts"].values():
            for f in findings:
                sev_counts[f.get("severity", "unknown")] += 1
        total = sum(sev_counts.values())

        lines: List[str] = []
        lines.append("=" * 70)
        lines.append(f"SECURITY FINDINGS REPORT — {name}")
        lines.append("=" * 70)
        lines.append(f"Generated : {generated}")
        if co.get("entity"):
            lines.append(f"Entity    : {co['entity']}")
        if co.get("sector"):
            lines.append(f"NIS2 sector: {co['sector']}"
                         + (f"  (NACE {co['nace']})" if co.get("nace") else ""))
        if co.get("kbo_url"):
            lines.append(f"KBO       : {co['kbo_url']}")
        lines.append(f"Hosts scanned : {len(data['hosts'])}")
        sev_line = ", ".join(
            f"{sev_counts[s]} {s}" for s in
            ("critical", "high", "medium", "low", "info", "unknown")
            if sev_counts.get(s))
        lines.append(f"Findings  : {total}" + (f"  ({sev_line})" if sev_line else ""))
        lines.append("")
        lines.append("This is an internal technical findings report for analyst "
                     "review. It is\nnot correspondence and is not addressed to "
                     "any individual.")
        lines.append("")

        for host in sorted(data["hosts"]):
            findings = sorted(
                data["hosts"][host],
                key=lambda f: (_SEV_ORDER.get(f.get("severity", "unknown"), 5),
                               f.get("name", "")))
            lines.append("-" * 70)
            lines.append(f"HOST: {host}")
            lines.append("-" * 70)
            if not findings:
                lines.append("  No findings.")
                lines.append("")
                continue
            for i, f in enumerate(findings, 1):
                title = f.get("name") or f.get("check") or f.get("template") or "Finding"
                lines.append(f"[{i}] {title}   ({f.get('severity', 'unknown').upper()})")
                if f.get("matched_at"):
                    lines.append(f"    Location    : {f['matched_at']}")
                if f.get("evidence"):
                    lines.append(f"    Evidence    : {f['evidence']}")
                if f.get("risk"):
                    lines.append(f"    Risk        : {f['risk']}")
                if f.get("remediation"):
                    lines.append(f"    Remediation : {f['remediation']}")
                lines.append("")

        slug = _slug(name)
        company_dir = report_root / slug
        company_dir.mkdir(parents=True, exist_ok=True)
        fpath = company_dir / f"findings_{slug}.txt"
        # Avoid collisions when two companies slugify to the same name.
        n = 2
        while fpath.exists():
            fpath = company_dir / f"findings_{slug}_{n}.txt"
            n += 1
        fpath.write_text("\n".join(lines), encoding="utf-8")
        written.append(str(fpath))

    if written:
        ok(f"Wrote {len(written)} per-company findings report(s) → {report_root}")
    return written

def _company_record(host: str, lookup: dict, hostname_index: dict) -> dict:
    co = resolve_company(host, lookup, hostname_index) or {}
    out = _blank_company()
    for key in out:
        out[key] = co.get(key, "")
    return out

def _sender_config() -> dict:
    """Sender identity for outbound intro emails, read from the environment."""
    return {
        "name":    os.environ.get("SENDER_NAME", "").strip()    or "[uw naam]",
        "company": os.environ.get("SENDER_COMPANY", "").strip() or "[uw bedrijf]",
        "email":   os.environ.get("SENDER_EMAIL", "").strip()   or "[uw e-mailadres]",
        "phone":   os.environ.get("SENDER_PHONE", "").strip()   or "[uw telefoon]",
        "unsub":   os.environ.get("UNSUB_URL", "").strip()      or "[afmeldlink]",
    }

def _intro_email_text(company_name: str, to_addr: str, sender: dict) -> str:
    """A neutral Dutch B2B service-introduction email.

    Contains no scan findings, no statement that the recipient's systems were
    examined, and no penalty/liability pressure. It introduces the sender's
    NIS2 services and asks whether a conversation is relevant, with an opt-out.
    """
    naam = company_name or "uw organisatie"
    lines = [
        f"Aan: {to_addr or '[algemeen e-mailadres]'}",
        f"Onderwerp: NIS2-ondersteuning voor {naam}",
        "",
        "Beste,",
        "",
        "NIS2 is van kracht en raakt steeds meer Belgische organisaties, "
        "mogelijk ook uw sector. Wij helpen bedrijven om hun beveiliging en "
        "documentatie in lijn te brengen met de nieuwe verplichtingen, zodat "
        "u goed voorbereid bent op een eventuele audit.",
        "",
        "Wij bieden onder meer:",
        "- een security-assessment met heldere prioriteiten en aanbevelingen;",
        "- concrete remediatie-adviezen, afgestemd op NIS2;",
        "- ondersteuning bij de documentatie die auditoren opvragen.",
        "",
        f"Als dit relevant is voor {naam}, plan ik graag een vrijblijvend "
        "gesprek van twintig minuten om te bekijken waar wij kunnen helpen.",
        "",
        "Met vriendelijke groet,",
        sender["name"],
        f"{sender['company']} · {sender['phone']} · {sender['email']}",
        "",
        "―",
        "Wilt u geen e-mails meer van ons ontvangen? Antwoord met \"stop\" "
        f"of gebruik deze afmeldlink: {sender['unsub']}",
    ]
    return "\n".join(lines)

def _load_functional_mailboxes(output_dir: str) -> Dict[str, List[str]]:
    """domain -> ranked functional mailboxes, from contact_enrichment.json.

    Returns an empty map if enrichment hasn't run, so intro emails still work.
    """
    out: Dict[str, List[str]] = {}
    p = Path(output_dir) / "contact_enrichment.json"
    if not p.exists():
        return out
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return out
    orgs = data if isinstance(data, list) else [data]
    for entry in orgs:
        org = entry.get("org", entry) if isinstance(entry, dict) else {}
        dom = str(org.get("domain", "")).strip().lower().replace("www.", "")
        fem = org.get("functional_emails") or []
        if dom and fem:
            out[dom] = fem
    return out

def _load_security_contacts(output_dir: str) -> Dict[str, str]:
    """domain -> declared/best security contact, from contact_enrichment.json."""
    out: Dict[str, str] = {}
    p = Path(output_dir) / "contact_enrichment.json"
    if not p.exists():
        return out
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return out
    orgs = data if isinstance(data, list) else [data]
    for entry in orgs:
        org = entry.get("org", entry) if isinstance(entry, dict) else {}
        dom = str(org.get("domain", "")).strip().lower().replace("www.", "")
        sc = str(org.get("security_contact", "")).strip()
        if dom and sc:
            out[dom] = sc
    return out

def write_intro_emails(scanned_hosts: List[str],
    output_dir: str,
    lookup: dict,
    hostname_index: dict,
    sender: Optional[dict] = None) -> List[str]:
    """Write one Dutch B2B intro email per scanned company (general mailbox)."""
    sender = sender or _sender_config()
    functional_by_domain = _load_functional_mailboxes(output_dir)
    companies: Dict[str, dict] = {}
    for host in (scanned_hosts or []):
        co = _company_record(host, lookup, hostname_index)
        key = co.get("entity") or co.get("name") or host
        companies.setdefault(key, {"co": co, "host": host})
    if not companies:
        return []

    report_root = Path(output_dir) / "by_company"
    report_root.mkdir(parents=True, exist_ok=True)
    written: List[str] = []
    for key, data in companies.items():
        co = data["co"]
        host = data["host"]
        name = co.get("name") or "uw organisatie"
        netloc = urlparse(host if "://" in host else "http://" + host).netloc or host
        domain = netloc.replace("www.", "").strip("/")
        # Prefer a discovered functional mailbox (contact@/info@, not security@
        # or a person) for a service introduction; fall back to info@domain.
        discovered = functional_by_domain.get(domain, [])
        intro_pref = [e for e in discovered
                      if _local_part(e) in ("contact", "info", "hello", "onthaal")]
        to_addr = (intro_pref[0] if intro_pref
                   else (f"info@{domain}" if domain else ""))
        body = _intro_email_text(name, to_addr, sender)
        slug = _slug(name)
        company_dir = report_root / slug
        company_dir.mkdir(parents=True, exist_ok=True)
        fpath = company_dir / f"intro_{slug}.txt"
        n = 2
        while fpath.exists():
            fpath = company_dir / f"intro_{slug}_{n}.txt"
            n += 1
        fpath.write_text(body, encoding="utf-8")
        written.append(str(fpath))
    if written:
        ok(f"Wrote {len(written)} intro email(s) → {report_root}")
    return written

# ── Split aggregate outputs into one file set per company ──────────────
_SPLIT_NAME_COLS = ["Company", "org_name", "company", "Denomination",
                    "denomination", "Naam", "name"]
_SPLIT_HOST_COLS = ["Host", "host", "url", "URL", "domain", "Domain",
                    "website", "Website"]

def _row_company(row: dict, lookup: dict, hostname_index: dict) -> str:
    for col in _SPLIT_NAME_COLS:
        v = str(row.get(col, "") or "").strip()
        if v:
            return v
    for col in _SPLIT_HOST_COLS:
        v = str(row.get(col, "") or "").strip()
        if v:
            co = _company_record(v, lookup, hostname_index)
            if co.get("name"):
                return co["name"]
    return ""

def _split_csv(path: Path, root: Path, lookup, hostname_index) -> List[str]:
    import pandas as pd
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    if df.empty:
        return []
    df["__company"] = df.apply(
        lambda r: _row_company(r.to_dict(), lookup, hostname_index)
        or "_unassigned", axis=1)
    written = []
    for co, sub in df.groupby("__company"):
        d = root / _slug(co, "unassigned")
        d.mkdir(parents=True, exist_ok=True)
        fp = d / path.name
        sub.drop(columns="__company").to_csv(fp, index=False)
        written.append(str(fp))
    return written

def _split_xlsx(path: Path, root: Path, lookup, hostname_index) -> List[str]:
    import pandas as pd
    df = pd.read_excel(path, dtype=str)
    df = df.fillna("")
    if df.empty:
        return []
    df["__company"] = df.apply(
        lambda r: _row_company(r.to_dict(), lookup, hostname_index)
        or "_unassigned", axis=1)
    written = []
    for co, sub in df.groupby("__company"):
        d = root / _slug(co, "unassigned")
        d.mkdir(parents=True, exist_ok=True)
        fp = d / path.name
        sub.drop(columns="__company").to_excel(fp, index=False)
        written.append(str(fp))
    return written

def _split_json_list(path: Path, root: Path, lookup, hostname_index) -> List[str]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(data, list) or not data:
        return []  # nested report objects are covered by the CSV split
    groups: Dict[str, list] = {}
    for item in data:
        if not isinstance(item, dict):
            continue
        co = _row_company(item, lookup, hostname_index) or "_unassigned"
        groups.setdefault(_slug(co, "unassigned"), []).append(item)
    written = []
    for slug, items in groups.items():
        d = root / slug
        d.mkdir(parents=True, exist_ok=True)
        fp = d / path.name
        fp.write_text(json.dumps(items, indent=2, ensure_ascii=False),
                      encoding="utf-8")
        written.append(str(fp))
    return written

def _split_nuclei_jsonl(path: Path, root: Path, lookup, hostname_index) -> List[str]:
    groups: Dict[str, list] = {}
    for finding in stream_findings(path):
        host = _finding_host(finding)
        co = _company_record(host, lookup, hostname_index).get("name") or "_unassigned"
        groups.setdefault(_slug(co, "unassigned"), []).append(finding)
    written = []
    for slug, items in groups.items():
        d = root / slug
        d.mkdir(parents=True, exist_ok=True)
        fp = d / path.name
        fp.write_text("\n".join(json.dumps(x, ensure_ascii=False) for x in items),
                      encoding="utf-8")
        written.append(str(fp))
    return written

def split_outputs_by_company(output_dir: str,
    nuclei_output: str,
    lookup: dict,
    hostname_index: dict) -> List[str]:
    """Split every aggregate output into one file per company under by_company/.

    After this runs, each company's folder holds its own scan CSV, coverage
    CSV/XLSX, contacts, manifest row, and findings JSONL, so no delivered file
    spans more than one company. Rows that can't be attributed land in an
    _unassigned/ folder rather than being dropped.
    """
    out = Path(output_dir)
    root = out / "by_company"
    root.mkdir(parents=True, exist_ok=True)
    written: List[str] = []

    csv_files = [SCAN_RESULTS_CSV, COVERAGE_CSV, "contact_enrichment.csv",
                 "combined_contacts.csv", "nis2_companies_manifest.csv"]
    for name in csv_files:
        p = out / name
        if p.exists() and p.stat().st_size > 0:
            try:
                written += _split_csv(p, root, lookup, hostname_index)
            except Exception as e:  # noqa: BLE001
                warn(f"split {name}: {e}")

    xlsx_p = out / COVERAGE_CSV.replace(".csv", ".xlsx")
    if xlsx_p.exists() and xlsx_p.stat().st_size > 0:
        try:
            written += _split_xlsx(xlsx_p, root, lookup, hostname_index)
        except Exception as e:  # noqa: BLE001
            warn(f"split {xlsx_p.name}: {e}")

    for name in ("contact_enrichment.json", "combined_contacts.json"):
        p = out / name
        if p.exists() and p.stat().st_size > 0:
            try:
                written += _split_json_list(p, root, lookup, hostname_index)
            except Exception as e:  # noqa: BLE001
                warn(f"split {name}: {e}")

    if nuclei_output and Path(nuclei_output).exists() \
            and Path(nuclei_output).stat().st_size > 0:
        try:
            written += _split_nuclei_jsonl(Path(nuclei_output), root,
                                           lookup, hostname_index)
        except Exception as e:  # noqa: BLE001
            warn(f"split {Path(nuclei_output).name}: {e}")

    if written:
        n_co = len({Path(p).parent.name for p in written})
        ok(f"Split aggregates into {n_co} company folder(s) → {root}")
    return written

def _build_readable_scan_report(matrix: dict,
    all_checks: List[dict],
    findings: List[dict],
    scanned_hosts: List[str],
    lookup: dict,
    hostname_index: dict,
    nuclei_output: str) -> Tuple[dict, List[dict]]:
    scanned_hosts = [h.rstrip("/") for h in scanned_hosts if h]
    host_order = list(dict.fromkeys(scanned_hosts + sorted(matrix.keys())))

    findings_by_host: Dict[str, List[dict]] = defaultdict(list)
    findings_by_host_check: Dict[Tuple[str, str], List[dict]] = defaultdict(list)
    for finding in findings:
        host = _finding_host(finding)
        if not host:
            continue
        item = _finding_summary(finding)
        key = _finding_check_key(finding)
        findings_by_host[host].append(item)
        findings_by_host_check[(host, key)].append(item)
        if host not in host_order:
            host_order.append(host)

    severity_counts = Counter(
        _safe_text((f.get("info") or {}).get("severity", "unknown")).lower()
        for f in findings
    )

    check_rows = []
    for ch in all_checks:
        f_count = sum(
            1 for host in host_order
            if matrix.get(host.rstrip("/"), {}).get(ch["key"], {}).get("status") == "FINDING"
        )
        c_count = max(len(host_order) - f_count, 0)
        check_rows.append({
            "template": ch["template"],
            "check": ch["name"],
            "severity": ch["severity"],
            "findings": f_count,
            "clean": c_count,
            "percent_clean": round(c_count / max(len(host_order), 1) * 100, 1),
        })
    check_rows.sort(key=lambda r: (-r["findings"], r["template"], r["check"]))

    hosts = []
    findings_by_url = []
    csv_rows = []
    total_clean = 0
    affected = 0

    for host in host_order:
        host_key = host.rstrip("/")
        co = _company_record(host_key, lookup, hostname_index)
        host_matrix = matrix.get(host_key, {})
        checks = []
        finding_count = 0
        clean_count = 0
        for ch in all_checks:
            cell = host_matrix.get(ch["key"], {
                "status": "CLEAN",
                "severity": ch["severity"],
                "template": ch["template"],
            })
            instances = findings_by_host_check.get((host_key, ch["key"]), [])
            status = cell.get("status", "CLEAN")
            if status == "FINDING":
                finding_count += 1
            else:
                clean_count += 1
                total_clean += 1
            first = instances[0] if instances else {}
            row = {
                "host": host_key,
                "company": co["name"],
                "entity": co["entity"],
                "nace": co["nace"],
                "sector": co["sector"],
                "kbo_url": co["kbo_url"],
                "template": ch["template"],
                "check": ch["name"],
                "severity": cell.get("severity", ch["severity"]),
                "result": status,
                "finding_count": len(instances),
                "matched_at": first.get("matched_at", ""),
                "evidence": first.get("evidence", ""),
                "description": first.get("description", ""),
                "executive_summary": first.get("executive_summary", ""),
                "risk": first.get("risk", ""),
                "remediation": first.get("remediation", ""),
            }
            checks.append({
                "template": row["template"],
                "check": row["check"],
                "severity": row["severity"],
                "result": row["result"],
                "finding_count": row["finding_count"],
                "matched_at": row["matched_at"],
                "evidence": row["evidence"],
                "description": row["description"],
                "executive_summary": row["executive_summary"],
                "risk": row["risk"],
                "remediation": row["remediation"],
            })
            csv_rows.append(row)

        if not all_checks:
            csv_rows.append({
                "host": host_key,
                "company": co["name"],
                "entity": co["entity"],
                "nace": co["nace"],
                "sector": co["sector"],
                "kbo_url": co["kbo_url"],
                "template": "",
                "check": "",
                "severity": "",
                "result": "NO_CHECKS_DEFINED",
                "finding_count": len(findings_by_host.get(host_key, [])),
                "matched_at": "",
                "evidence": "",
                "description": "No named template checks were parsed.",
                "executive_summary": "",
                "risk": "",
                "remediation": "",
            })

        raw_findings = findings_by_host.get(host_key, [])
        finding_rows = [
            c for c in checks
            if c.get("result") == "FINDING"
        ]
        if not finding_rows and raw_findings:
            finding_rows = [
                {
                    "template": f.get("template", ""),
                    "check": f.get("check", ""),
                    "severity": f.get("severity", ""),
                    "result": "FINDING",
                    "finding_count": 1,
                    "matched_at": f.get("matched_at", ""),
                    "evidence": f.get("evidence", ""),
                    "description": f.get("description", ""),
                    "executive_summary": f.get("executive_summary", ""),
                    "risk": f.get("risk", ""),
                    "remediation": f.get("remediation", ""),
                }
                for f in raw_findings
            ]

        if finding_count or raw_findings:
            affected += 1
            findings_by_url.append({
                "host": host_key,
                "company": co,
                "affected_checks": finding_count or len(finding_rows),
                "finding_events": len(raw_findings),
                "findings": finding_rows,
            })
        hosts.append({
            "host": host_key,
            "company": co,
            "status": "affected" if finding_count or raw_findings else "clean",
            "findings": finding_count,
            "affected_checks": finding_count,
            "finding_events": len(raw_findings),
            "clean_checks": clean_count,
            "raw_findings": raw_findings,
            "checks": checks,
        })

    hosts.sort(key=lambda h: (-h["affected_checks"], -h["finding_events"], h["host"]))
    findings_by_url.sort(key=lambda h: (
        -h["affected_checks"], -h["finding_events"], h["host"]
    ))
    csv_rows.sort(key=lambda r: (r["host"], r["template"], r["check"]))

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "results_file": nuclei_output,
        "hosts_scanned": len(scanned_hosts) or len(host_order),
        "urls_reported": len(host_order),
        "hosts_affected": affected,
        "hosts_clean": max(len(host_order) - affected, 0),
        "templates": len({ch["template"] for ch in all_checks}),
        "checks_defined": len(all_checks),
        "check_evaluations": len(host_order) * len(all_checks),
        "findings_total": len(findings),
        "finding_checks": sum(h["affected_checks"] for h in hosts),
        "clean_checks": total_clean,
        "severity": {sev: severity_counts.get(sev, 0) for sev in SEV_ORDER},
    }
    report = {
        "summary": summary,
        "checks": check_rows,
        "findings_by_url": findings_by_url,
        "hosts": hosts,
    }
    return report, csv_rows

def _write_readable_scan_json(report: dict, path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

def _write_readable_scan_csv(rows: List[dict], path: Path) -> None:
    columns = [
        "Host", "Company", "Entity", "NACE", "Sector", "KBO_URL",
        "Template", "Check", "Severity", "Result", "Finding_Count",
        "Matched_At", "Evidence", "Executive_Summary", "Risk",
        "Remediation", "Description",
    ]
    keys = [
        "host", "company", "entity", "nace", "sector", "kbo_url",
        "template", "check", "severity", "result", "finding_count",
        "matched_at", "evidence", "executive_summary", "risk",
        "remediation", "description",
    ]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(columns)
        for row in rows:
            w.writerow([row.get(k, "") for k in keys])

def _write_readable_scan_html(report: dict, path: Path) -> None:
    summary = report.get("summary", {})
    sev = summary.get("severity", {})

    severity_rows = ""
    for name in SEV_ORDER:
        severity_rows += (
            f"<tr><td>{_html(name.capitalize())}</td>"
            f"<td class='num'>{_html(sev.get(name, 0))}</td></tr>"
        )

    check_rows = ""
    for check in report.get("checks", []):
        cls = "finding" if check.get("findings", 0) else "clean"
        check_rows += (
            f"<tr class='{cls}'><td>{_html(check.get('template'))}</td>"
            f"<td>{_html(check.get('check'))}</td>"
            f"<td>{_html(check.get('severity'))}</td>"
            f"<td class='num'>{_html(check.get('findings'))}</td>"
            f"<td class='num'>{_html(check.get('clean'))}</td>"
            f"<td class='num'>{_html(check.get('percent_clean'))}%</td></tr>"
        )
    if not check_rows:
        check_rows = "<tr><td colspan='6' class='muted'>No named checks parsed.</td></tr>"

    finding_url_sections = ""
    for url_row in report.get("findings_by_url", []):
        co = url_row.get("company") or {}
        finding_rows = ""
        for finding in url_row.get("findings", []):
            finding_rows += (
                "<tr class='finding'>"
                f"<td>{_html(finding.get('severity'))}</td>"
                f"<td>{_html(finding.get('template'))}</td>"
                f"<td>{_html(finding.get('check'))}</td>"
                f"<td class='num'>{_html(finding.get('finding_count'))}</td>"
                f"<td>{_html(finding.get('matched_at'))}</td>"
                f"<td class='narrative'>{_html(finding.get('evidence'))}</td>"
                f"<td class='narrative'>{_html(finding.get('description'))}</td>"
                f"<td class='narrative'>{_html(finding.get('risk'))}</td>"
                f"<td class='narrative'>{_html(finding.get('remediation'))}</td>"
                "</tr>"
            )
        if not finding_rows:
            finding_rows = (
                "<tr><td colspan='9' class='muted'>"
                "No individual finding rows were available for this URL."
                "</td></tr>"
            )

        company = co.get("name") or "Unknown company"
        finding_url_sections += f"""
<details class="host-card" open>
  <summary>
    <span class="host-title">{_html(url_row.get('host'))}</span>
    <span class="badge bad">{_html(url_row.get('affected_checks', 0))} affected checks</span>
    <span class="badge neutral">{_html(url_row.get('finding_events', 0))} events</span>
    <span class="muted">{_html(company)}</span>
  </summary>
  <div class="company">
    <div><strong>Entity:</strong> {_html(co.get('entity'))}</div>
    <div><strong>NACE:</strong> {_html(co.get('nace'))}</div>
    <div><strong>Sector:</strong> {_html(co.get('sector'))}</div>
    <div><strong>KBO:</strong> <a href="{_html(co.get('kbo_url'))}">{_html(co.get('kbo_url'))}</a></div>
  </div>
  <table>
    <thead><tr><th>Severity</th><th>Template</th><th>Check</th><th>Events</th><th>Matched at</th><th>Evidence</th><th>Description / Executive Summary</th><th>Risk</th><th>Remediation</th></tr></thead>
    <tbody>{finding_rows}</tbody>
  </table>
</details>"""
    if not finding_url_sections:
        finding_url_sections = (
            "<p class='muted'>No findings were reported for any URL.</p>"
        )

    host_sections = ""
    for host in report.get("hosts", []):
        co = host.get("company") or {}
        finding_checks = [
            c for c in host.get("checks", [])
            if c.get("result") == "FINDING"
        ]
        clean_checks = [
            c for c in host.get("checks", [])
            if c.get("result") != "FINDING"
        ]
        detail_rows = ""
        for check in finding_checks + clean_checks:
            cls = "finding" if check.get("result") == "FINDING" else "clean"
            detail_rows += (
                f"<tr class='{cls}'><td>{_html(check.get('result'))}</td>"
                f"<td>{_html(check.get('template'))}</td>"
                f"<td>{_html(check.get('check'))}</td>"
                f"<td>{_html(check.get('severity'))}</td>"
                f"<td>{_html(check.get('matched_at'))}</td>"
                f"<td class='narrative'>{_html(check.get('evidence'))}</td>"
                f"<td class='narrative'>{_html(check.get('description'))}</td>"
                f"<td class='narrative'>{_html(check.get('risk'))}</td>"
                f"<td class='narrative'>{_html(check.get('remediation'))}</td></tr>"
            )
        if not detail_rows:
            detail_rows = "<tr><td colspan='9' class='muted'>No check rows available.</td></tr>"

        title = host.get("host", "")
        company = co.get("name") or "Unknown company"
        host_sections += f"""
<details class="host-card" {'open' if host.get('status') == 'affected' else ''}>
  <summary>
    <span class="host-title">{_html(title)}</span>
    <span class="badge {'bad' if host.get('status') == 'affected' else 'good'}">
      {_html(host.get('affected_checks', host.get('findings', 0)))} affected checks
    </span>
    <span class="badge neutral">{_html(host.get('finding_events', 0))} events</span>
    <span class="muted">{_html(company)}</span>
  </summary>
  <div class="company">
    <div><strong>Entity:</strong> {_html(co.get('entity'))}</div>
    <div><strong>NACE:</strong> {_html(co.get('nace'))}</div>
    <div><strong>Sector:</strong> {_html(co.get('sector'))}</div>
    <div><strong>KBO:</strong> <a href="{_html(co.get('kbo_url'))}">{_html(co.get('kbo_url'))}</a></div>
  </div>
  <table>
    <thead><tr><th>Result</th><th>Template</th><th>Check</th><th>Severity</th><th>Matched at</th><th>Evidence</th><th>Description / Executive Summary</th><th>Risk</th><th>Remediation</th></tr></thead>
    <tbody>{detail_rows}</tbody>
  </table>
</details>"""

    if not host_sections:
        host_sections = "<p class='muted'>No hosts were available in this report.</p>"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>NIS2 Scan Results</title>
<style>
:root {{ --bg:#f5f7fb; --card:#fff; --text:#1f2937; --muted:#6b7280; --line:#e5e7eb; --bad:#b91c1c; --good:#047857; }}
body {{ margin:0; font-family:Segoe UI,Arial,sans-serif; background:var(--bg); color:var(--text); }}
header {{ background:#111827; color:white; padding:24px 32px; }}
h1 {{ margin:0 0 6px; font-size:26px; }}
main {{ padding:24px 32px 40px; }}
.cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:14px; margin-bottom:22px; }}
.card,.panel,.host-card {{ background:var(--card); border:1px solid var(--line); border-radius:12px; box-shadow:0 1px 2px rgba(0,0,0,.04); }}
.card {{ padding:16px; }}
.card strong {{ display:block; font-size:26px; }}
.muted {{ color:var(--muted); }}
.grid {{ display:grid; grid-template-columns: minmax(240px,360px) 1fr; gap:18px; align-items:start; }}
.panel {{ padding:16px; overflow:auto; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th,td {{ border-bottom:1px solid var(--line); padding:8px 10px; text-align:left; vertical-align:top; }}
th {{ background:#f9fafb; font-weight:600; }}
.num {{ text-align:right; }}
.narrative {{ white-space:pre-wrap; min-width:180px; }}
.finding td {{ background:#fff5f5; }}
.clean td {{ background:#f4fbf7; }}
.host-card {{ margin-top:14px; padding:0; overflow:hidden; }}
.host-card summary {{ cursor:pointer; padding:14px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
.host-title {{ font-weight:700; }}
.badge {{ border-radius:999px; padding:3px 9px; color:white; font-size:12px; }}
.badge.bad {{ background:var(--bad); }}
.badge.good {{ background:var(--good); }}
.badge.neutral {{ background:#4b5563; }}
.company {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:8px; padding:0 16px 12px; color:var(--muted); font-size:13px; }}
.note {{ margin:-8px 0 20px; }}
a {{ color:#2563eb; }}
@media (max-width: 900px) {{ .grid {{ grid-template-columns:1fr; }} main,header {{ padding-left:16px; padding-right:16px; }} }}
</style>
</head>
<body>
<header>
  <h1>NIS2 Scan Results</h1>
  <div class="muted">Generated {_html(summary.get('generated_at'))} from {_html(summary.get('results_file'))}</div>
</header>
<main>
  <section class="cards">
    <div class="card"><strong>{_html(summary.get('hosts_scanned', 0))}</strong><span class="muted">URLs scanned</span></div>
    <div class="card"><strong>{_html(summary.get('hosts_affected', 0))}</strong><span class="muted">URLs with findings</span></div>
    <div class="card"><strong>{_html(summary.get('findings_total', 0))}</strong><span class="muted">Nuclei finding events</span></div>
    <div class="card"><strong>{_html(summary.get('finding_checks', 0))}</strong><span class="muted">Affected checks</span></div>
    <div class="card"><strong>{_html(summary.get('checks_defined', 0))}</strong><span class="muted">Checks defined</span></div>
    <div class="card"><strong>{_html(summary.get('clean_checks', 0))}</strong><span class="muted">Clean check evaluations</span></div>
  </section>
  <p class="muted note">Finding events are raw nuclei JSON rows. Affected checks are deduplicated by URL and check name, so this is the number to compare with clean check evaluations.</p>
  <section>
    <h2>Findings per URL</h2>
    {finding_url_sections}
  </section>
  <section class="grid">
    <div class="panel">
      <h2>Severity breakdown</h2>
      <table><tbody>{severity_rows}</tbody></table>
    </div>
    <div class="panel">
      <h2>Checks summary</h2>
      <table>
        <thead><tr><th>Template</th><th>Check</th><th>Severity</th><th>Findings</th><th>Clean</th><th>% Clean</th></tr></thead>
        <tbody>{check_rows}</tbody>
      </table>
    </div>
  </section>
  <section>
    <h2>Full URL check matrix</h2>
    {host_sections}
  </section>
</main>
</body>
</html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)

def save_readable_scan_reports(matrix: dict,
    all_checks: List[dict],
    findings: List[dict],
    scanned_hosts: List[str],
    lookup: dict,
    hostname_index: dict,
    output_dir: str,
    nuclei_output: str) -> dict:
    out_dir = Path(output_dir)
    report, csv_rows = _build_readable_scan_report(
        matrix, all_checks, findings, scanned_hosts,
        lookup, hostname_index, nuclei_output)
    paths = {
        "json": out_dir / SCAN_RESULTS_JSON,
        "csv": out_dir / SCAN_RESULTS_CSV,
        "html": out_dir / SCAN_RESULTS_HTML,
    }
    try:
        _write_readable_scan_json(report, paths["json"])
        ok(f"Readable JSON: {paths['json']}")
    except OSError as e:
        warn(f"Could not write readable JSON report: {e}")
    try:
        _write_readable_scan_csv(csv_rows, paths["csv"])
        ok(f"Readable CSV : {paths['csv']}")
    except OSError as e:
        warn(f"Could not write readable CSV report: {e}")
    try:
        _write_readable_scan_html(report, paths["html"])
        ok(f"Readable HTML: {paths['html']}")
    except OSError as e:
        warn(f"Could not write readable HTML report: {e}")
    return {k: str(v) for k, v in paths.items()}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Post-scan summary

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def print_scan_summary(nuclei_output: str,
    output_dir: Optional[str] = None,
    template_paths: Optional[List[str]] = None,
    template_checks: Optional[Dict[str, dict]] = None,
    scanned_hosts: Optional[List[str]] = None,
    export_xlsx: bool = False) -> None:
    p = Path(nuclei_output)
    valid, msg = validate_results_file(p)
    empty_results = (p.exists() and p.stat().st_size == 0)
    if not valid and not empty_results:
        warn(f"Results file unusable: {msg}")
        return

    findings: List[dict] = [] if empty_results else list(stream_findings(p))
    if empty_results:
        ok("Scan complete — no findings (results file empty).")

    lookup, hostname_index = (load_url_lookup(output_dir)
                              if output_dir else ({}, {}))

    if template_checks is None:
        template_checks = parse_template_checks(template_paths or [])

    if scanned_hosts is None:
        tf = Path(output_dir or ".") / "targets.txt"
        if tf.exists():
            scanned_hosts = load_targets_from_file(str(tf))
        else:
            scanned_hosts = list(
                {f.get("host", "").rstrip("/") for f in findings}
            )
    scanned_hosts = [h.rstrip("/") for h in scanned_hosts]

    matrix, all_checks_list = build_coverage_matrix(
        findings, scanned_hosts, template_checks)

    run_start = datetime.now()
    header("SCAN SUMMARY")
    ok(f"Findings loaded : {len(findings):,}")
    ok(f"URLs scanned    : {len(scanned_hosts):,}")
    ok(f"Templates parsed: {len(template_checks)}")

    if not findings and not template_checks:
        info("No findings and no named template checks were parsed.")

    # ━━ 1. Severity breakdown ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    subhead("Findings by Severity")
    sev_counts = Counter(
        f.get("info", {}).get("severity", "unknown").lower()
        for f in findings
    )
    for s in SEV_ORDER:
        cnt = sev_counts.get(s, 0)
        if cnt:
            label = _c(f"{s:<10}", SEV_COLORS.get(s, Fore.WHITE))
            print(f"  {label}  {cnt:>5}  {_sev_bar(cnt)}")
    if not findings:
        print("  (no findings)")
    print()

    # ━━ 2. Per-template breakdown ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    subhead("Findings by Template")
    tmpl_sev: Dict[str, Counter] = defaultdict(Counter)
    for f in findings:
        tmpl_sev[f.get("template-id", "unknown")][
            f.get("info", {}).get("severity", "unknown").lower()] += 1
    for tid in template_checks:
        tmpl_sev.setdefault(tid, Counter())

    cols = ["Template", "Total"] + [s.capitalize() for s in SEV_ORDER]
    rows = [[tid, sum(sc.values())] + [sc.get(s, 0) for s in SEV_ORDER]
            for tid, sc in sorted(tmpl_sev.items(),
                                  key=lambda x: -sum(x[1].values()))]
    row_colors = [Fore.RED if r[1] else Fore.GREEN for r in rows]
    summary_table(rows, cols, row_colors=row_colors)
    print()

    # ━━ 3. All checks: findings vs clean ━━━━━━━━━━━━━━━━━━━━━━━━━━━
    subhead("All Checks – Affected URLs vs Clean  (deduplicated by URL)")
    fired_events: Counter = Counter()
    fired_hosts_by_key: Dict[str, set] = defaultdict(set)
    for f in findings:
        host = _finding_host(f)
        if not host:
            continue
        key = _finding_check_key(f)
        fired_events[key] += 1
        fired_hosts_by_key[key].add(host)

    n_hosts = len(matrix)
    rows, row_colors = [], []
    for ch in sorted(all_checks_list, key=lambda x: x["name"]):
        key = ch["key"]
        fired_n = len(fired_hosts_by_key.get(key, set()))
        event_n = fired_events.get(key, 0)
        clean_n = max(n_hosts - fired_n, 0)
        pct_c   = f"{clean_n / max(n_hosts, 1) * 100:.0f}%"
        has_f   = fired_n > 0
        rows.append([ch["template"], ch["name"],
                     ch["severity"].upper(), fired_n, event_n, clean_n, pct_c,
                     "✗ FINDINGS" if has_f else "✓ ALL CLEAN"])
        row_colors.append(Fore.RED if has_f else Fore.GREEN)
    rows.sort(key=lambda r: (-r[3], r[1]))
    summary_table(rows,
                  ["Template", "Check", "Sev",
                   "Affected URLs", "Events", "Clean", "% Clean", "Status"],
                  row_colors=row_colors)
    print()

    # ━━ 4. Per-sector breakdown ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    if lookup:
        subhead("Findings by NIS2 Sector")
        sector_counts: Counter            = Counter()
        sector_sev:    Dict[str, Counter] = defaultdict(Counter)
        for f in findings:
            co     = resolve_company(_finding_host(f),
                                     lookup, hostname_index)
            sector = co["sector"] if co else "Unknown"
            sev    = f.get("info", {}).get("severity", "unknown").lower()
            sector_counts[sector] += 1
            sector_sev[sector][sev] += 1
        cols = ["Sector", "Total"] + [s.capitalize() for s in SEV_ORDER]
        rows = [[sec, tot] + [sector_sev[sec].get(s, 0) for s in SEV_ORDER]
                for sec, tot in sorted(sector_counts.items(),
                                       key=lambda x: -x[1])]
        summary_table(rows, cols)
        print()

    # ━━ 5. Findings per URL ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    subhead("Findings per URL  (top 20)")
    host_findings: Dict[str, list] = defaultdict(list)
    for f in findings:
        host = _finding_host(f)
        if host:
            host_findings[host].append(f)
    for h in matrix:
        host_findings.setdefault(h, [])

    total_checks_per_host = len(all_checks_list) or 1
    cols = ["#", "URL", "Affected Checks", "Events", "Clean Checks",
            "Company", "NACE", "Sector", "KBO"]
    top_rows, row_colors = [], []
    def _affected_check_count(host: str) -> int:
        return sum(
            1 for cell in matrix.get(host, {}).values()
            if cell.get("status") == "FINDING"
        )

    sorted_hosts = sorted(
        host_findings,
        key=lambda h: (-_affected_check_count(h), -len(host_findings[h]), h)
    )
    for rank, host in enumerate(sorted_hosts[:20], 1):
        flist = host_findings[host]
        co = resolve_company(host, lookup, hostname_index)
        f  = _affected_check_count(host)
        c  = max(total_checks_per_host - f, 0)
        top_rows.append([rank, host, f, len(flist), c,
                         (co["name"]    or "–") if co else "–",
                         (co["nace"]    or "–") if co else "–",
                         (co["sector"]  or "–") if co else "–",
                         (co["kbo_url"] or "–") if co else "–"])
        row_colors.append(Fore.RED if f else Fore.GREEN)
    summary_table(top_rows, cols, row_colors=row_colors)
    print()

    # ━━ 6. Per-URL detail ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    subhead("Per-URL Finding Detail  (top 10 URLs)")
    top_hosts = [r[1] for r in top_rows[:10]]

    findings_by_host_check: Dict[Tuple[str, str], List[dict]] = defaultdict(list)
    for finding in findings:
        host = _finding_host(finding)
        if not host:
            continue
        findings_by_host_check[(host, _finding_check_key(finding))].append(
            _finding_summary(finding))

    for host in top_hosts:
        host_key = host.rstrip("/")
        co       = resolve_company(host, lookup, hostname_index) or {}
        name     = co.get("name", "")
        label    = host + (f"  │  {name}" if name else "")
        print(_c(f"\n  ▸ {label}", Fore.WHITE))
        if co.get("entity"):
            print(_c(f"      Entity : {co['entity']}   "
                     f"NACE: {co.get('nace', '')}   "
                     f"Sector: {co.get('sector', '')}", Fore.CYAN))
        if co.get("kbo_url"):
            print(_c(f"      KBO    : {co['kbo_url']}", Fore.CYAN))

        host_matrix = matrix.get(host_key, {})
        fnd = [(ch, host_matrix[ch["key"]]) for ch in all_checks_list
               if host_matrix.get(ch["key"], {}).get("status") == "FINDING"]
        cln = [(ch, host_matrix[ch["key"]]) for ch in all_checks_list
               if host_matrix.get(ch["key"], {}).get("status") == "CLEAN"]

        print(_c(f"      Checks: {len(all_checks_list)}   "
                 f"Findings: {len(fnd)}   Clean: {len(cln)}", Fore.WHITE))

        if fnd:
            print(_c("      ── FINDINGS ──", Fore.RED))
            for ch, cell in fnd:
                sev_col  = SEV_COLORS.get(cell["severity"], Fore.WHITE)
                # FIX 1: pre-compute the padded name column so the format spec
                # :<45 never appears inside an f-string function-call argument,
                # which causes a SyntaxError in Python < 3.12.
                name_col = f"{_trunc(ch['name'], 45):<45}"
                instances = findings_by_host_check.get((host_key, ch["key"]), [])
                first = instances[0] if instances else {}
                print(f"        {_c('✗', Fore.RED)}  "
                      f"{_c(name_col, Fore.YELLOW)}  "
                      f"{_c(cell['severity'].upper()[:8], sev_col)}  "
                      f"events={len(instances) or 1}")
                if first.get("matched_at"):
                    print(f"           matched: {_trunc(first['matched_at'], 90)}")
                if first.get("evidence"):
                    print(f"           evidence: {_trunc(first['evidence'], 90)}")
        if cln:
            print(_c("      ── CLEAN ──", Fore.GREEN))
            for ch, _ in cln:
                print(f"        {_c('✓', Fore.GREEN)}  "
                      f"{_c(_trunc(ch['name'], 45), Fore.WHITE)}")
    print()

    # ━━ 7. Export outputs ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    if output_dir:
        save_coverage_csv(matrix, all_checks_list,
                          lookup, hostname_index, output_dir)
        save_readable_scan_reports(matrix, all_checks_list,
                                   findings, scanned_hosts,
                                   lookup, hostname_index,
                                   output_dir, nuclei_output)
        if export_xlsx:
            save_coverage_xlsx(matrix, all_checks_list,
                               lookup, hostname_index, output_dir)

    # ━━ 8. Totals ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    total_findings = len(findings)
    total_clean    = sum(1 for hm in matrix.values()
                         for v in hm.values() if v["status"] == "CLEAN")
    affected       = sum(
        1 for host, hm in matrix.items()
        if any(v["status"] == "FINDING" for v in hm.values())
        or host_findings.get(host)
    )
    clean_hosts    = len(matrix) - affected
    subhead("Totals")
    ok(f"Nuclei events    : {total_findings:,}")
    ok(f"Affected checks  : "
       f"{sum(_affected_check_count(h) for h in matrix):,}")
    ok(f"Clean checks     : {total_clean:,}  (ran, no issue found)")
    ok(f"URLs affected    : {affected:,}  /  {len(matrix):,} reported  "
       f"({clean_hosts:,} fully clean)")
    ok(f"Templates run    : {len(template_checks)}")
    ok(f"Checks defined   : {len(all_checks_list)}")
    info(f"Results file     : {nuclei_output}")
    info(f"Summary duration : {datetime.now() - run_start}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# ██████████████████  CONTACT INTELLIGENCE ENGINE  ████████████████████

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# 

# Standalone OR post-scan enrichment.

# Sources (in order):

# 1  KBO Public Search     → mandataries, registered phone

# 2  Company website       → staff/team/board pages, emails, phones

# 3  Belgisch Staatsblad   → historical mandatary publications

# 4  Google SERP           → indexed emails, LinkedIn URLs, mentions

# 5  Bing SERP             → cross-check

# 6  DuckDuckGo SERP       → fallback (no JS needed)

# 7  LinkedIn public page  → role / employer via JSON-LD

# 8  Hunter.io API         → verified emails + domain pattern

# 9  SMTP RCPT-TO probe    → confirm deliverability of inferred addresses

# ─────────────────────────────────────────────────────────────────────

try:
    import requests as _requests
    from bs4 import BeautifulSoup as _BS
    from requests.adapters import HTTPAdapter as _HTTPAdapter
    try:
        from urllib3.util.retry import Retry as _Retry
    except Exception:
        _Retry = None
    _CI_HTTP = True
except ImportError:
    _CI_HTTP = False
    _HTTPAdapter = None
    _Retry = None

try:
    import dns.resolver as _dns_resolver
    _CI_DNS = True
except ImportError:
    _CI_DNS = False

import smtplib as _smtplib
from dataclasses import dataclass as _dataclass, field as _field
from urllib.parse import quote_plus as _qp

# ── Constants ─────────────────────────────────────────────────────────

_CI_KBO_URL      = ("https://kbopub.economie.fgov.be/kbopub/"
"toonondernemingps.html")
_CI_STAATSBLAD   = "https://www.ejustice.just.fgov.be/cgi_tsv/list.pl"
_CI_HUNTER_API   = "https://api.hunter.io/v2"
_CI_APOLLO_API   = "https://api.apollo.io/v1"
_CI_EMAILFMT_URL = "https://www.email-format.com/d"
_CI_INFOBEL_URL  = "https://www.infobel.com/fr/belgium"

_CI_UA_POOL = [
("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
"(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
("Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 "
"(KHTML, like Gecko) Version/17.4 Safari/605.1.15"),
"Mozilla/5.0 (X11; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0",
]
_CI_UA_IDX = 0
_CI_UA_LOCK = threading.Lock()
_CI_TLS = threading.local()

# Role keyword → buyer-likelihood score (cybersec assessment context)

_CI_ROLE_SCORES: Dict[str, int] = {
"coordinat": 95, "kringcoordinat": 95,
"directeur": 90, "director": 90, "ceo": 90,
"cto": 92,       "ciso": 98,
"it manager": 88, "ict manager": 88, "it coordinator": 87,
"digital": 82,   "informatica": 85,
"stafmedewerker": 68, "staf": 65, "databeheer": 72,
"project": 60,   "financiën": 58, "finance": 58,
"penningmeester": 55,
"administratief": 22, "admin": 22, "secretaris": 18,
"voorzitter": 52, "president": 52,
"ondervoorzitter": 22, "vice": 22,
"bestuurder": 18,
"huisarts": 10, "arts": 10, "dr.": 10, "gp": 10,
}

# ── Data models ───────────────────────────────────────────────────────

@_dataclass
class CIContact:
    name:               str
    role:               str  = ""
    email:              str  = ""
    email_status:       str  = ""   # confirmed | smtp-ok | smtp-reject | inferred
    phone:              str  = ""
    phone_type:         str  = ""   # org | practice | direct
    linkedin_url:       str  = ""   # confirmed profile URL
    linkedin_search_url: str = ""   # always-populated manual fallback search URL
    linkedin_role:      str  = ""
    sources:            list = _field(default_factory=list)
    score:              int  = 0
    notes:              str  = ""

@_dataclass
class CIOrgProfile:
    kbo:           str
    name:          str  = ""
    domain:        str  = ""
    address:       str  = ""
    org_phone:     str  = ""
    org_email:     str  = ""
    email_pattern: str  = ""
    functional_emails: list = _field(default_factory=list)
    security_contact: str = ""
    security_policy:  str = ""
    network_abuse:    str = ""
    contacts:      list = _field(default_factory=list)

# Functional/role mailboxes, in priority order for security & NIS2 outreach.
# These are not personal data about an identifiable individual, which makes
# them the preferred contact channel over an inferred person's address.
_FUNCTIONAL_PREFIXES = [
    "security", "abuse", "soc", "cert",           # security reporting
    "privacy", "dpo", "gdpr", "compliance",       # data protection
    "it", "ict", "helpdesk", "support",           # technical
    "contact", "info", "hello", "onthaal",        # general
    "secretariaat", "administratie", "office",    # admin (NL/BE)
]
_FUNCTIONAL_RANK = {p: i for i, p in enumerate(_FUNCTIONAL_PREFIXES)}

def _local_part(email: str) -> str:
    return email.split("@", 1)[0].strip().lower() if "@" in email else ""

def _is_functional(email: str) -> bool:
    lp = _local_part(email)
    return any(lp == p or lp.startswith(p + ".") or lp.startswith(p + "-")
               for p in _FUNCTIONAL_PREFIXES)

def _rank_functional(emails: Iterable[str]) -> List[str]:
    def key(e):
        lp = _local_part(e)
        for p in _FUNCTIONAL_PREFIXES:
            if lp == p or lp.startswith(p + ".") or lp.startswith(p + "-"):
                return (_FUNCTIONAL_RANK[p], e)
        return (len(_FUNCTIONAL_PREFIXES), e)
    seen, out = set(), []
    for e in sorted(set(emails), key=key):
        el = e.lower()
        if el not in seen:
            seen.add(el)
            out.append(e)
    return out

    # ── HTTP helper ───────────────────────────────────────────────────────

def _ci_ua() -> dict:
    global _CI_UA_IDX
    with _CI_UA_LOCK:
        ua = _CI_UA_POOL[_CI_UA_IDX % len(_CI_UA_POOL)]
        _CI_UA_IDX += 1
    h = {
        "User-Agent": ua,
        "Accept-Language": "nl-BE,nl;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    }
    return h

def _ci_session():
    """Thread-local requests session with connection pooling."""
    if not _CI_HTTP:
        return None
    sess = getattr(_CI_TLS, "session", None)
    if sess is not None:
        return sess

    sess = _requests.Session()
    retry_cfg = 0
    if _Retry is not None:
        retry_cfg = _Retry(
            total=2,
            connect=2,
            read=2,
            backoff_factor=0.2,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=frozenset(["GET", "POST", "HEAD"]),
        )
    adapter = _HTTPAdapter(
        pool_connections=32,
        pool_maxsize=32,
        max_retries=retry_cfg,
    )
    sess.mount("http://", adapter)
    sess.mount("https://", adapter)
    _CI_TLS.session = sess
    return sess

def _ci_get(url: str, proxies: dict, delay: float,
    params=None, extra_headers: dict = None,
    timeout: int = 14) -> Optional[object]:
    if not _CI_HTTP:
        return None
    if delay > 0:
        time.sleep(delay)
    headers = _ci_ua()
    if extra_headers:
        headers.update(extra_headers)
    session = _ci_session()
    if session is None:
        return None
    try:
        r = session.get(
            url,
            headers=headers,
            params=params,
            proxies=proxies,
            timeout=timeout,
            allow_redirects=True,
        )
        return r if r.status_code == 200 else None
    except Exception:
        return None

def _ci_soup(r) -> object:
    return _BS(r.text, "html.parser")

    # ── Text helpers ──────────────────────────────────────────────────────

def _ci_norm_phone(raw: str) -> str:
    digits = re.sub(r"[^\d]", "", raw)
    if digits.startswith("32") and len(digits) >= 10:
        return "+" + digits
    if digits.startswith("0") and len(digits) >= 9:
        return "+32" + digits[1:]
    return digits

def _ci_emails_from(text: str, domain: str = "") -> List[str]:
    found = {
        e.lower()
        for e in re.findall(r"[\w.+%-]{2,}@[\w.-]+\.[a-zA-Z]{2,}", text or "")
    }
    if domain:
        domain_l = domain.lower().strip()
        found = {e for e in found if e.split("@", 1)[-1].endswith(domain_l)}
    return sorted(found)

def _ci_phones_from(text: str) -> List[str]:
    """Extract Belgian phone numbers from plain text."""
    raw = re.findall(r"(?:\+32|0032|0)[\s./-]?\d(?:[\s./-]?\d){7,10}", text or "")
    results: List[str] = []
    seen = set()
    for p in raw:
        normed = _ci_norm_phone(p)
        digits = re.sub(r"\D", "", normed)
        if 9 <= len(digits) <= 12 and normed not in seen:
            seen.add(normed)
            results.append(normed)
    return results

def _ci_phones_from_soup(soup, kbo_digits: str = "") -> List[str]:
    """
    Extract phones from HTML structure — 3 layers:
    1. <a href="tel:..."> links  (most reliable on modern sites)
    2. JSON-LD telephone field   (Organization / LocalBusiness schema)
    3. Plain-text regex fallback
    Filters out any number whose digits match kbo_digits.
    """
    found: List[str] = []

    def _accept(normed: str) -> bool:
        d = re.sub(r"\D", "", normed)
        return (9 <= len(d) <= 12
                and d != kbo_digits
                and normed not in found)

    # 1. tel: links
    for a in soup.find_all("a", href=re.compile(r"^tel:", re.I)):
        raw = a["href"][4:].strip()
        n   = _ci_norm_phone(raw)
        if _accept(n):
            found.append(n)

    # 2. JSON-LD structured data
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "{}")
            if isinstance(data, list):
                data = next((d for d in data if d.get("@type") in (
                    "Organization", "LocalBusiness",
                    "MedicalOrganization", "Hospital", "GovernmentOrganization"
                )), {})
            for phone in ([data.get("telephone")] +
                          data.get("contactPoint", [{}]) if isinstance(
                              data.get("contactPoint"), list) else []):
                if isinstance(phone, dict):
                    phone = phone.get("telephone", "")
                if phone:
                    n = _ci_norm_phone(str(phone))
                    if _accept(n):
                        found.append(n)
        except (json.JSONDecodeError, AttributeError, TypeError):
            pass

    # 3. Plain-text fallback
    for p in _ci_phones_from(soup.get_text(" ", strip=True)):
        if _accept(p):
            found.append(p)

    return found

def _ci_infer_email(first: str, last: str, domain: str, pattern: str) -> str:
    f = re.sub(r"[^a-z]", "", first.lower())
    l = re.sub(r"[^a-z]", "", last.lower())
    fi = f[0] if f else ""
    if "v.naam" in pattern:
        return f"{fi}.{l}@{domain}"
    if pattern.startswith("voornaam@"):
        return f"{f}@{domain}"
    return f"{f}.{l}@{domain}"   # default: voornaam.naam

def _ci_name_key(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())

    # ── 1. KBO profile fetch ──────────────────────────────────────────────

def ci_fetch_kbo(kbo: str, proxies: dict, delay: float) -> dict:
    """Fetch KBO page, return mandataries + org contact info."""
    clean = re.sub(r"[^0-9]", "", kbo)
    r     = _ci_get(_CI_KBO_URL, proxies, delay,
    params={"ondernemingsnummer": clean})
    out   = {"kbo": kbo, "name": "", "phone": "", "email": "",
    "address": "", "domain": "", "mandataries": []}
    if not r:
        return out

    soup = _ci_soup(r)
    board_roles = {
        "Bestuurder", "Voorzitter", "Gedelegeerd bestuurder",
        "Ondervoorzitter", "Secretaris", "Penningmeester",
        "Zaakvoerder", "Administrateur", "Directeur",
    }
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if not cells:
            continue
        c0 = cells[0].get_text(" ", strip=True)
        c1 = cells[1].get_text(" ", strip=True) if len(cells) > 1 else ""
        c2 = cells[2].get_text(" ", strip=True) if len(cells) > 2 else ""

        if "Naam:" in c0 and not out["name"]:
            # KBO appends " Naam in het Nederlands, sinds DD maand YYYY" —
            # strip everything from " Naam " onward to get the bare company name.
            raw_name = c1
            raw_name = re.sub(
                r"\s+Naam\s+in\s+het\s+\w+.*$", "", raw_name,
                flags=re.IGNORECASE).strip()
            raw_name = re.sub(
                r"\s*,\s*sinds\s+\d.*$", "", raw_name,
                flags=re.IGNORECASE).strip()
            out["name"] = raw_name
        if "Telefoonnummer" in c0:
            kbo_digits = re.sub(r"\D", "", kbo)
            # Try tel: links in this row first (more reliable)
            tel_links = row.find_all("a", href=re.compile(r"^tel:", re.I))
            if tel_links:
                raw    = tel_links[0]["href"][4:].strip()
                normed = _ci_norm_phone(raw)
                digits = re.sub(r"\D", "", normed)
                if 9 <= len(digits) <= 12 and digits != kbo_digits:
                    out["phone"] = normed
            else:
                # Fall back to regex on cell text
                phs = [p for p in re.findall(r"0\d[\d\s]{7,10}", c1)
                       if re.sub(r"\D", "", p) != kbo_digits]
                if phs:
                    normed = _ci_norm_phone(phs[0])
                    if len(re.sub(r"\D", "", normed)) >= 9:
                        out["phone"] = normed
        if "E-mail" in c0:
            em = re.search(r"[\w.+%-]+@[\w.-]+\.\w+", c1)
            if em:
                out["email"] = em.group().lower()
        if "Webadres" in c0:
            for a in row.find_all("a", href=True):
                h = a["href"]
                if h.startswith("http") and "kbo" not in h:
                    out["domain"] = urlparse(h).netloc.replace("www.", "")
        if "Adres van de zetel" in c0:
            out["address"] = c1
        if c0 in board_roles:
            parts = c1.split(",")
            name  = (f"{parts[1].strip()} {parts[0].strip()}"
                     if len(parts) == 2 else c1)
            # Skip legal entities acting as mandataries — their "name" in KBO
            # is a KBO number (0XXX.XXX.XXX) or contains NV/BV/VZW/SA etc.
            is_legal_entity = bool(
                re.match(r"^0\d{3}[\s.]?\d{3}[\s.]?\d{3}$", name.strip()) or
                re.search(r"\b(NV|BV|VZW|BVBA|SA|ASBL|SRL|CV|SCS|SNC)\b",
                          name, re.IGNORECASE)
            )
            if not is_legal_entity:
                out["mandataries"].append(
                    {"name": name, "role": c0, "since": c2})

    info(f"[CI-KBO] {out['name']} | {len(out['mandataries'])} mandataries")
    return out

    # ── 2. Website scraper ────────────────────────────────────────────────

def ci_scrape_website(domain: str, proxies: dict, delay: float) -> dict:
    out = {"staff": [], "board": [], "emails": [], "phones": [], "pattern": "",
           "functional": []}
    if not _CI_HTTP:
        return out

    base = f"https://{domain}"
    # Homepage always; also probe /contact since that's where phones live.
    # We deliberately keep this list short — the old 14-path approach
    # burned time on 404s and was removed. /contact is the one exception
    # worth the extra request because it almost always has phone numbers.
    urls_to_try = [base, base + "/contact"]

    for url in urls_to_try:
        r = _ci_get(url, proxies, delay * 0.4)
        if not r:
            continue
        soup = _ci_soup(r)
        text = soup.get_text(" ", strip=True)

        for e in _ci_emails_from(text, domain):
            if e not in out["emails"]:
                out["emails"].append(e)
        for a in soup.find_all("a", href=re.compile(r"mailto:", re.I)):
            raw = a["href"].replace("mailto:", "").split("?")[0].strip().lower()
            if "@" in raw and domain in raw and raw not in out["emails"]:
                out["emails"].append(raw)

        # Use structured soup extraction — far more reliable than regex on text
        for ph in _ci_phones_from_soup(soup):
            if ph not in out["phones"]:
                out["phones"].append(ph)

        for heading in soup.find_all(["h3", "h4", "strong", "b", "p"]):
            name_raw = heading.get_text(" ", strip=True)
            words    = name_raw.split()
            if not (2 <= len(words) <= 4):
                continue
            if not all(w[0].isupper() for w in words if w and w[0].isalpha()):
                continue
            role_el  = heading.find_next_sibling()
            role_raw = role_el.get_text(" ", strip=True) if role_el else ""
            entry    = {"name": name_raw, "role": role_raw}
            rl       = role_raw.lower()
            if any(k in rl for k in ("coordinat", "staf", "medewerker",
                                     "directeur", "director", "manager")):
                out["staff"].append(entry)
            elif any(k in rl for k in ("voorzitter", "bestuurder", "secretaris",
                                       "penningmeester", "ondervoorzitter")):
                out["board"].append(entry)

    personal = [e for e in out["emails"]
                if "." in e.split("@")[0]
                and not _is_functional(e)]
    out["pattern"] = f"voornaam.naam@{domain}" if personal else ""
    out["functional"] = _rank_functional(
        e for e in out["emails"] if _is_functional(e))
    dedup = lambda lst: list({e["name"]: e for e in lst}.values())
    out["staff"] = dedup(out["staff"])
    out["board"]  = dedup(out["board"])
    info(f"[CI-WEB] {len(out['emails'])} emails "
         f"({len(out['functional'])} functional), {len(out['phones'])} phones, "
         f"{len(out['staff'])} staff, {len(out['board'])} board entries")
    return out

    # ── 3. Belgisch Staatsblad ────────────────────────────────────────────

def ci_fetch_staatsblad(kbo: str, proxies: dict, delay: float) -> List[dict]:
    clean = re.sub(r"[^0-9]", "", kbo)
    r = _ci_get(_CI_STAATSBLAD, proxies, delay,
    params={"language": "nl", "btw": clean, "page": 1})
    if not r:
        return []
    soup  = _ci_soup(r)
    names = []
    for row in soup.find_all("tr"):
        text = row.get_text(" ", strip=True)
        if re.search(r"\b(bestuurder|mandataris|voorzitter|secretaris)\b",
        text, re.I):
            found = re.findall(
            r"\b([A-ZÀÁÂ][a-zàáâãäåæç]+(?:\s+[A-ZÀÁÂ][a-zàáâãäåæç]+)+)\b",
            text)
            for n in found:
                if len(n.split()) >= 2:
                    names.append({"name": n, "source": "Staatsblad"})
    dedup = list({e["name"]: e for e in names}.values())
    if dedup:
        info(f"[CI-SB] {len(dedup)} name mentions")
    return dedup

                # ── 4-6. Search engine SERPs ─────────────────────────────────────────

def _ci_serp_links(soup) -> List[str]:
    links = []
    seen = set()
    for a in soup.find_all("a", href=True):
        h = a["href"]
        if "/url?q=" in h:
            h = re.sub(r"^.*?/url\?q=([^&]+).*$", r"\1", h)
        if (h.startswith("http") and "google.com" not in h
            and "bing.com" not in h and "duckduckgo.com" not in h):
            if h not in seen:
                seen.add(h)
                links.append(h)
    return links

def _ci_serp_text(soup) -> str:
    for s in soup(["script", "style", "nav", "header", "footer"]):
        s.decompose()
    return soup.get_text(" ", strip=True)

def _ci_google(query: str, proxies: dict, delay: float):
    url = f"https://www.google.com/search?q={_qp(query)}&hl=nl&num=10"
    r = _ci_get(url, proxies, delay, extra_headers={"Accept-Language": "nl-BE"})
    if not r:
        return [], ""
    soup = _ci_soup(r)
    return _ci_serp_links(soup), _ci_serp_text(soup)

def _ci_bing(query: str, proxies: dict, delay: float):
    r = _ci_get("https://www.bing.com/search", proxies, delay,
    params={"q": query, "setlang": "nl-BE"})
    if not r:
        return [], ""
    soup = _ci_soup(r)
    return _ci_serp_links(soup), _ci_serp_text(soup)

def _ci_ddg(query: str, proxies: dict, delay: float):
    r = _ci_get("https://html.duckduckgo.com/html/", proxies, delay,
    params={"q": query, "kl": "be-nl"})
    if not r:
        return [], ""
    soup = _ci_soup(r)
    return _ci_serp_links(soup), _ci_serp_text(soup)

def _ci_linkedin_search_url(name: str, org: str) -> str:
    """Always-available manual LinkedIn search URL for the analyst."""
    return (f"https://www.linkedin.com/search/results/people/"
    f"?keywords={_qp(name + ' ' + org)}")

def ci_try_linkedin_pubdir(first: str, last: str,
    proxies: dict, delay: float) -> str:
    """
    Try LinkedIn pub/dir — a public legacy endpoint that lists matching
    profiles without requiring login. Returns first matching profile URL
    or empty string.
    """
    if not _CI_HTTP or not first or not last:
        return ""
    f_slug = re.sub(r"[^a-z]", "", first.lower())
    l_slug = re.sub(r"[^a-z]", "", last.lower())
    url = f"https://www.linkedin.com/pub/dir/{f_slug}/{l_slug}"
    r = _ci_get(url, proxies, delay * 0.5, extra_headers={
    "Referer": "https://www.google.com/"})
    if not r:
        return ""
    soup = _ci_soup(r)
    for a in soup.find_all("a", href=re.compile(r"linkedin.com/in/")):
        href = a.get("href", "").split("?")[0].rstrip("/")
        if "/in/" in href:
            info(f"[CI-LI] pub/dir hit: {href}")
            return href
    return ""

def ci_fetch_linkedin_company(domain: str, proxies: dict, delay: float) -> str:
    """
    Fetch LinkedIn company page — returns the about/description text.
    Slug is guessed from the root domain label (lotusbakeries → lotusbakeries).
    """
    if not _CI_HTTP:
        return ""
    slug = domain.split(".")[0].lower()
    slug = re.sub(r"[^a-z0-9-]", "", slug)
    url  = f"https://www.linkedin.com/company/{slug}"
    r = _ci_get(url, proxies, delay * 0.5, extra_headers={
    "Referer": "https://www.google.com/"})
    if not r:
        return ""
    soup = _ci_soup(r)
    # og:description often has "X followers · Industry · About" text
    og = soup.find("meta", property="og:description")
    if og and og.get("content"):
        return og["content"][:300]
    return ""

def ci_run_serps(name: str, org: str, domain: str,
    proxies: dict, delay: float) -> dict:
    result = {"linkedin_url": "", "emails": [], "phones": [],
    "mentions": [], "sources": []}

    # LinkedIn-specific query runs on ALL three engines — not rotating.
    # Google blocks most, but Bing and DDG often return results.
    li_query   = f'"{name}" "{org}" site:linkedin.com/in'
    mail_query = f'"{name}" "{org}" email'
    all_links: List[str] = []
    all_text:  str = ""

    for engine_fn, engine_name in [(_ci_ddg, "DDG"),
                                   (_ci_bing, "Bing"),
                                   (_ci_google, "Google")]:
        time.sleep(delay * 0.3)
        links, text = engine_fn(li_query, proxies, 0)
        all_links.extend(links)
        all_text += " " + text
        if links:
            result["sources"].append(engine_name)
        # Stop as soon as we find a LinkedIn URL
        if any("linkedin.com/in/" in l for l in links):
            break

    # One email/mention pass using DDG (most permissive)
    time.sleep(delay * 0.3)
    links2, text2 = _ci_ddg(mail_query, proxies, 0)
    all_links.extend(links2)
    all_text += " " + text2

    for l in all_links:
        if "linkedin.com/in/" in l:
            result["linkedin_url"] = l.split("?")[0].rstrip("/")
            break

    result["emails"] = _ci_emails_from(all_text, domain)
    result["phones"] = _ci_phones_from(all_text)

    _SERP_NOISE = re.compile(
        r"please click here|send feedback|privacy policy|terms of service"
        r"|cookie|translate this page|more results|did you mean"
        r"|people also ask|related searches",
        re.IGNORECASE)

    first_word = name.split()[0] if name.strip() else ""
    if first_word:
        for m in re.finditer(re.escape(first_word), all_text, re.I):
            s    = max(0, m.start() - 40)
            e    = min(len(all_text), m.end() + 120)
            snip = all_text[s:e].strip()
            if _SERP_NOISE.search(snip):
                continue
            if org.lower()[:4] in snip.lower() or domain in snip.lower():
                result["mentions"].append(snip[:200])
                if len(result["mentions"]) >= 3:
                    break

    return result

# ── 7. LinkedIn public profile ────────────────────────────────────────

def ci_fetch_linkedin(url: str, proxies: dict, delay: float) -> dict:
    if not url or not _CI_HTTP:
        return {}
    r = _ci_get(url, proxies, delay,
    extra_headers={"Referer": "https://www.google.com/"})
    if not r:
        return {}
    soup   = _ci_soup(r)
    result = {"url": url, "role": "", "company": ""}

    og = soup.find("meta", property="og:title")
    if og and og.get("content"):
        m = re.match(r"^.+?\s*[-–]\s*(.+?)\s*(?:at|bij|@|\|)", og["content"])
        if m:
            result["role"] = m.group(1).strip()
        m2 = re.search(r"(?:at|bij|@)\s+(.+?)(?:\s*\||\s*$)", og["content"])
        if m2:
            result["company"] = m2.group(1).strip()

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "{}")
            if isinstance(data, list):
                data = next((d for d in data if d.get("@type") == "Person"), {})
            if data.get("@type") == "Person":
                result["role"]    = result["role"]    or data.get("jobTitle", "")
                wf = data.get("worksFor", {})
                result["company"] = result["company"] or (
                    wf.get("name", "") if isinstance(wf, dict) else "")
        except (json.JSONDecodeError, AttributeError):
            pass
    return result

    # ── 8a. EmailFormat.com (free, no key) ───────────────────────────────

def ci_query_emailformat(domain: str, proxies: dict, delay: float) -> dict:
    """
    Scrape email-format.com to discover the company's standard email pattern
    and any indexed individual addresses. Completely free, no API key.
    Returns {"pattern": str, "emails": [str]}
    """
    if not _CI_HTTP:
        return {}
    r = _ci_get(f"{_CI_EMAILFMT_URL}/{domain}/", proxies, delay * 0.4)
    if not r:
        return {}
    soup    = _ci_soup(r)
    result  = {"pattern": "", "emails": []}
    # Pattern is shown in a <span class="email-format"> or similar element
    for el in soup.find_all(string=re.compile(r"[{(]first[*.]?name[})]|[{(]f[})]",
    re.I)):
        text = el.strip()
        if "@" in text and domain in text:
            result["pattern"] = text.strip()
            break
    # Also grab any individual emails listed
    for em in re.findall(r"[\w.+-]{2,}@" + re.escape(domain), soup.get_text()):
        em_l = em.lower()
        if em_l not in result["emails"]:
            result["emails"].append(em_l)
    if result["pattern"] or result["emails"]:
        info(f"[CI-EF] pattern={result['pattern'] or 'n/a'} "
             f"| {len(result['emails'])} emails")
    return result

                # ── 8b. Infobel.be (free Belgian business directory) ─────────────────

def ci_query_infobel(org_name: str, proxies: dict, delay: float) -> dict:
    """
    Search Infobel Belgium for the company — returns phone and address
    if listed. Completely free, no API key.
    """
    if not _CI_HTTP:
        return {}
    query = re.sub(r"\s+", "+", org_name.strip()[:40])
    r = _ci_get(f"{_CI_INFOBEL_URL}/{query}", proxies, delay * 0.5)
    if not r:
        return {}
    soup   = _ci_soup(r)
    result = {"phone": "", "address": ""}
    # Infobel shows phone in tel: links
    tel = soup.find("a", href=re.compile(r"tel:"))
    if tel:
        raw = tel["href"].replace("tel:", "").strip()
        normed = _ci_norm_phone(raw)
        if len(re.sub(r"\D", "", normed)) >= 9:
            result["phone"] = normed

    # Address in structured data
    for sd in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(sd.string or "{}")
        except (json.JSONDecodeError, TypeError):
            continue
        candidates = data if isinstance(data, list) else [data]
        for item in candidates:
            if not isinstance(item, dict):
                continue
            if item.get("@type") not in ("LocalBusiness", "Organization"):
                continue
            addr = item.get("address", {})
            if isinstance(addr, dict):
                assembled = (f"{addr.get('streetAddress', '')} "
                             f"{addr.get('postalCode', '')} "
                             f"{addr.get('addressLocality', '')}").strip()
                if assembled:
                    result["address"] = assembled
                    break
        if result["address"]:
            break

    if result["phone"] or result["address"]:
        info(f"[CI-IB] phone={result['phone'] or 'n/a'} "
             f"address={result['address'][:40] or 'n/a'}")
    return result

                # ── 8b2. Gouden Gids / Pages d'Or ────────────────────────────────────

def ci_query_goudengids(org_name: str, proxies: dict, delay: float) -> dict:
    """
    Gouden Gids (goudengids.be) / Pages d'Or — Belgian Yellow Pages.
    Authoritative for most NIS2-regulated Belgian companies.
    Free, no API key.
    """
    if not _CI_HTTP:
        return {}
    result = {"phone": "", "address": ""}
    for base_url in ("https://www.goudengids.be/nl/zoeken/",
    "https://www.pagesdor.be/fr/recherche/"):
        r = _ci_get(base_url, proxies, delay * 0.5,
        params={"q": org_name[:50]})
        if not r:
            continue
        soup = _ci_soup(r)
        for a in soup.find_all("a", href=re.compile(r"^tel:", re.I)):
            raw = a["href"][4:].strip()
            normed = _ci_norm_phone(raw)
            if len(re.sub(r"\D", "", normed)) >= 9:
                result["phone"] = normed
                break

        if not result["address"]:
            for el in soup.select("[class*='address'],[itemprop='address'],"
                                  "[itemprop='streetAddress']"):
                text = el.get_text(" ", strip=True)
                if text and len(text) > 5:
                    result["address"] = text[:120]
                    break

        for sd in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(sd.string or "{}")
            except (json.JSONDecodeError, TypeError):
                continue
            candidates = data if isinstance(data, list) else [data]
            for item in candidates:
                if not isinstance(item, dict):
                    continue
                if not result["phone"]:
                    ph = item.get("telephone", "")
                    if ph:
                        normed = _ci_norm_phone(str(ph))
                        if len(re.sub(r"\D", "", normed)) >= 9:
                            result["phone"] = normed
                if not result["address"]:
                    addr = item.get("address", {})
                    if isinstance(addr, dict):
                        assembled = (f"{addr.get('streetAddress', '')} "
                                     f"{addr.get('postalCode', '')} "
                                     f"{addr.get('addressLocality', '')}").strip()
                        if assembled:
                            result["address"] = assembled[:120]
            if result["phone"] and result["address"]:
                break
        if result["phone"] and result["address"]:
            break

    if result["phone"] or result["address"]:
        info(f"[CI-GG] phone={result['phone'] or 'n/a'}  "
             f"address={result['address'][:40] or 'n/a'}")
    return result

                    # ── 8b3. RIZIV/INAMI — Belgian healthcare provider registry ───────────

def ci_query_riziv(org_name: str, kbo: str,
    proxies: dict, delay: float) -> dict:
    """
    RIZIV/INAMI public institution register — hospitals, care homes,
    GP practices, pharmacies. Authoritative for Health sector NIS2.
    Free, no key.
    """
    if not _CI_HTTP:
        return {}
    result = {"phone": "", "address": "", "riziv_number": ""}
    clean_kbo = re.sub(r"\D", "", kbo)

    r = _ci_get(
        "https://ondpanon.riziv.fgov.be/Home/InstitutionSearch",
        proxies, delay * 0.5,
        params={"enterpriseNumber": clean_kbo})
    if r:
        soup = _ci_soup(r)
        for a in soup.find_all("a", href=re.compile(r"^tel:", re.I)):
            raw = a["href"][4:].strip()
            normed = _ci_norm_phone(raw)
            if len(re.sub(r"\D", "", normed)) >= 9:
                result["phone"] = normed
                break
        for el in soup.find_all(string=re.compile(r"\b\d{7}\b")):
            m = re.search(r"\b(\d{7})\b", el)
            if m:
                result["riziv_number"] = m.group(1)
                break

    if result["phone"]:
        info(f"[CI-RIZIV] phone={result['phone']}  "
             f"RIZIV#={result['riziv_number'] or 'n/a'}")
    return result

# ── 8b4. VREG — Flemish energy regulator ──────────────────────────────

def ci_query_vreg(org_name: str, kbo: str,
    proxies: dict, delay: float) -> dict:
    """
    VREG public register of licensed energy suppliers, DSOs and producers.
    Relevant for Energy sector NIS2 targets. Free, no key.
    """
    if not _CI_HTTP:
        return {}
    result = {"phone": "", "license_type": ""}
    clean_kbo = re.sub(r"\D", "", kbo)

    r = _ci_get(
        "https://www.vreg.be/nl/vergunningenregister",
        proxies, delay * 0.5,
        params={"ondernemingsnummer": clean_kbo, "naam": org_name[:40]})
    if r:
        soup = _ci_soup(r)
        for a in soup.find_all("a", href=re.compile(r"^tel:", re.I)):
            raw = a["href"][4:].strip()
            normed = _ci_norm_phone(raw)
            if len(re.sub(r"\D", "", normed)) >= 9:
                result["phone"] = normed
                break
        for el in soup.select("[class*='license'],[class*='vergunning']"):
            text = el.get_text(" ", strip=True)
            if text:
                result["license_type"] = text[:80]
                break

    # Fallback: VREG open data CSV (quarterly published)
    if not result["phone"]:
        try:
            r2 = _ci_get(
                "https://www.vreg.be/sites/default/files/"
                "vergunningenregister.csv",
                proxies, delay * 0.3)
            if r2:
                for line in r2.text.splitlines()[1:300]:
                    if clean_kbo in line or org_name[:10].lower() in line.lower():
                        for part in [p.strip().strip('"') for p in line.split(";")]:
                            if re.match(r"0\d{8,9}$|^\+32", part):
                                result["phone"] = _ci_norm_phone(part)
                                break
                    if result["phone"]:
                        break
        except Exception:
            pass

    if result["phone"]:
        info(f"[CI-VREG] phone={result['phone']}  "
             f"license={result['license_type'][:30] or 'n/a'}")
    return result

# ── 8b5. BIPT — Belgian telecom & digital infra regulator ─────────────

def ci_query_bipt(org_name: str, kbo: str,
    proxies: dict, delay: float) -> dict:
    """
    BIPT public operator register — licensed telecom, postal and digital
    infrastructure providers. Relevant for Digital infrastructure and
    Postal & courier NIS2 sectors. Free, no key.
    """
    if not _CI_HTTP:
        return {}
    result = {"phone": "", "operator_type": ""}
    clean_kbo = re.sub(r"\D", "", kbo)

    for search_url in (
        "https://www.bipt.be/operators/publication/list-of-operators",
        "https://www.ibpt.be/operateurs/publication/liste-des-operateurs",
    ):
        r = _ci_get(search_url, proxies, delay * 0.5)
        if not r:
            continue
        soup = _ci_soup(r)
        for row in soup.find_all("tr"):
            text = row.get_text(" ", strip=True)
            if clean_kbo in text or org_name[:12].lower() in text.lower():
                for a in row.find_all("a", href=re.compile(r"^tel:", re.I)):
                    raw = a["href"][4:].strip()
                    normed = _ci_norm_phone(raw)
                    if len(re.sub(r"\D", "", normed)) >= 9:
                        result["phone"] = normed
                        break
                cells = row.find_all("td")
                if cells:
                    result["operator_type"] = cells[0].get_text(
                        " ", strip=True)[:60]
                if result["phone"]:
                    break
        if result["phone"]:
            break

    if result["phone"]:
        info(f"[CI-BIPT] phone={result['phone']}  "
             f"type={result['operator_type'][:30] or 'n/a'}")
    return result

# ── 8c. Apollo.io (optional API key, 150 free credits/month) ──────────

def ci_query_apollo(domain: str, api_key: str) -> dict:
    """
    Apollo.io people search by domain — returns verified emails, names
    and titles. Free tier: 150 credits/month. Requires API key via
    –apollo-key flag.
    """
    if not api_key or not _CI_HTTP:
        return {}
    session = _ci_session()
    if session is None:
        return {}
    try:
        r = session.post(
            f"{_CI_APOLLO_API}/mixed_people/search",
            headers={
                "Content-Type": "application/json",
                "Cache-Control": "no-cache",
                "X-Api-Key": api_key,
            },
            json={"q_organization_domains": domain, "page": 1, "per_page": 10},
            timeout=15,
        )
        data = r.json()
        people = data.get("people", [])
        info(f"[CI-APOLLO] {len(people)} people found for {domain}")
        return {
        "emails": [
        {"email":    p.get("email", ""),
        "first":    p.get("first_name", ""),
        "last":     p.get("last_name", ""),
        "position": p.get("title", ""),
        "linkedin": p.get("linkedin_url", "")}
        for p in people if p.get("email")
        ]
        }
    except Exception as exc:
        warn(f"[CI-APOLLO] {exc}")
        return {}

    # ── 8d. Hunter.io ─────────────────────────────────────────────────────

def ci_query_hunter(domain: str, api_key: str) -> dict:
    if not api_key or not _CI_HTTP:
        return {}
    session = _ci_session()
    if session is None:
        return {}
    try:
        r = session.get(
            f"{_CI_HUNTER_API}/domain-search",
            params={"domain": domain, "api_key": api_key, "limit": 20},
            timeout=12,
        )
        data = r.json().get("data", {})
        info(f"[CI-HUNTER] pattern={data.get('pattern','')} "
        f"| {len(data.get('emails',[]))} emails")
        return {
        "pattern": data.get("pattern", ""),
        "emails":  [
        {"email":    e.get("value", ""),
        "first":    e.get("first_name", ""),
        "last":     e.get("last_name", ""),
        "position": e.get("position", ""),
        "confidence": e.get("confidence", 0)}
        for e in data.get("emails", [])
        ],
        }
    except Exception as exc:
        warn(f"[CI-HUNTER] {exc}")
        return {}

    # ── 9. SMTP verification ──────────────────────────────────────────────

def _ci_mx_for(domain: str) -> Optional[str]:
    if not _CI_DNS:
        return None
    try:
        records = sorted(_dns_resolver.resolve(domain, "MX"),
        key=lambda r: r.preference)
        return str(records[0].exchange).rstrip(".")
    except Exception:
        return None

def ci_smtp_verify(email: str, mx_host: Optional[str] = None) -> str:
    """RCPT-TO probe — no message sent. Returns smtp-ok/smtp-reject/smtp-unknown."""
    if not email or "@" not in email:
        return "unknown"
    domain = email.split("@")[1]
    mx     = mx_host or _ci_mx_for(domain)
    if not mx:
        return "smtp-unknown"
    try:
        with _smtplib.SMTP(timeout=8) as s:
            s.connect(mx, 25)
            s.ehlo("contact-intel.local")
            s.mail("probe@contact-intel.local")
            code, _ = s.rcpt(email)
            if code == 250:     return "smtp-ok"
            if code in (550, 551, 553): return "smtp-reject"
            return "smtp-unknown"
    except _smtplib.SMTPRecipientsRefused:
        return "smtp-reject"
    except Exception:
        return "smtp-unknown"

    # ── Functional mailbox discovery ──────────────────────────────────────

_SECURITY_TXT_PATHS = ["/.well-known/security.txt", "/security.txt"]

# Dedicated pages that commonly publish a functional/compliance mailbox.
_CONTACT_PAGES = ["/contact", "/contacts", "/contact-us", "/privacy",
                  "/privacybeleid", "/privacy-policy", "/gdpr", "/avg",
                  "/dpo", "/klokkenluider", "/security", "/legal",
                  "/over-ons", "/about"]

def _emails_on_domain(text: str, domain: str) -> List[str]:
    out = []
    for e in re.findall(r"[\w.+%-]+@[\w.-]+\.\w+", text or ""):
        el = e.lower()
        if el.endswith("@" + domain) or el.endswith("." + domain):
            if el not in out:
                out.append(el)
    return out

def soa_rname_email(domain: str) -> str:
    """The SOA record's responsible-person mailbox (RNAME), dots-to-@ decoded."""
    if not _CI_DNS or not domain:
        return ""
    try:
        ans = _dns_resolver.resolve(domain, "SOA")
        rname = str(ans[0].rname).rstrip(".")
    except Exception:
        return ""
    # First unescaped dot separates local-part from domain.
    parts = re.split(r"(?<!\\)\.", rname, maxsplit=1)
    if len(parts) == 2:
        local = parts[0].replace("\\.", ".")
        return f"{local}@{parts[1]}".lower()
    return ""

def dmarc_report_emails(domain: str) -> List[str]:
    """Mailboxes from the _dmarc TXT record's rua/ruf report addresses."""
    if not _CI_DNS or not domain:
        return []
    try:
        ans = _dns_resolver.resolve(f"_dmarc.{domain}", "TXT")
    except Exception:
        return []
    txt = " ".join(b.decode(errors="replace") if isinstance(b, bytes) else str(b)
                   for r in ans for b in r.strings)
    out = []
    for m in re.findall(r"(?:rua|ruf)=([^;]+)", txt, re.I):
        for addr in m.split(","):
            addr = addr.strip()
            if addr.lower().startswith("mailto:"):
                e = addr[7:].strip().lower()
                if "@" in e and e not in out:
                    out.append(e)
    return out

def tls_cert_emails(domain: str) -> List[str]:
    """Any emailAddress fields in the site's TLS certificate subject/SANs."""
    if not domain:
        return []
    import ssl as _ssl
    ctx = _ssl.create_default_context()
    out = []
    try:
        with socket.create_connection((domain, 443), timeout=6) as sock:
            with ctx.wrap_socket(sock, server_hostname=domain) as ss:
                cert = ss.getpeercert()
    except Exception:
        return []
    for field in cert.get("subject", ()) + cert.get("issuer", ()):
        for k, v in field:
            if k == "emailAddress" and "@" in v:
                out.append(v.lower())
    for typ, val in cert.get("subjectAltName", ()):
        if typ.lower() == "email" and "@" in val:
            out.append(val.lower())
    return list(dict.fromkeys(out))

def whois_abuse_mailbox(domain: str) -> str:
    """RIPE abuse-c mailbox for the IP the domain resolves to.

    This is the *hosting network's* abuse desk (mandated in the RIPE region),
    correct for network-abuse reports but NOT the organisation's own inbox.
    """
    try:
        ip = socket.gethostbyname(domain)
    except Exception:
        return ""
    def _q(server, query):
        try:
            with socket.create_connection((server, 43), timeout=6) as s:
                s.sendall((query + "\r\n").encode())
                buf = b""
                while True:
                    d = s.recv(4096)
                    if not d:
                        break
                    buf += d
                return buf.decode(errors="replace")
        except Exception:
            return ""
    # RIPE returns abuse-mailbox directly with the -b flag.
    text = _q("whois.ripe.net", f"-b {ip}")
    m = re.search(r"abuse-mailbox:\s*([\w.+%-]+@[\w.-]+\.\w+)", text, re.I)
    if m:
        return m.group(1).lower()
    m = re.search(r"%\s*Abuse contact for.*?is\s*'([\w.+%-]+@[\w.-]+\.\w+)'",
                  text, re.I | re.S)
    return m.group(1).lower() if m else ""

def scrape_contact_pages(domain: str, proxies: dict, delay: float) -> List[str]:
    """Functional mailboxes published on dedicated contact/privacy/DPO pages."""
    if not _CI_HTTP or not domain:
        return []
    out: List[str] = []
    for path in _CONTACT_PAGES:
        if len([e for e in out if _is_functional(e)]) >= 3:
            break  # enough signal; don't fetch every page
        r = _ci_get(f"https://{domain}{path}", proxies, delay * 0.25)
        if not r or getattr(r, "status_code", 0) != 200:
            continue
        soup = _ci_soup(r)
        for a in soup.find_all("a", href=re.compile(r"mailto:", re.I)):
            raw = a["href"].replace("mailto:", "").split("?")[0].strip().lower()
            if raw.endswith("@" + domain) and raw not in out:
                out.append(raw)
        for e in _emails_on_domain(soup.get_text(" ", strip=True), domain):
            if e not in out:
                out.append(e)
    # Keep only functional addresses; discard scraped personal ones here.
    return [e for e in out if _is_functional(e)]

def fetch_security_txt(domain: str, proxies: dict, delay: float) -> dict:
    """Fetch and parse RFC 9116 security.txt — the canonical, purpose-built
    place an organisation declares where to report vulnerabilities.

    Returns {"contacts": [emails...], "policy": url, "found": bool}. Contacts
    are the authoritative vulnerability-report addresses when present.
    """
    out = {"contacts": [], "policy": "", "found": False}
    if not _CI_HTTP or not domain:
        return out
    for scheme in ("https", "http"):
        for path in _SECURITY_TXT_PATHS:
            r = _ci_get(f"{scheme}://{domain}{path}", proxies, delay * 0.3)
            if not r or getattr(r, "status_code", 0) != 200:
                continue
            body = r.text or ""
            # Guard against soft-404 HTML pages served for missing files.
            if "<html" in body[:200].lower():
                continue
            for line in body.splitlines():
                line = line.strip()
                low = line.lower()
                if low.startswith("contact:"):
                    val = line.split(":", 1)[1].strip()
                    if val.lower().startswith("mailto:"):
                        val = val[7:].strip()
                    if "@" in val and " " not in val:
                        e = val.lower()
                        if e not in out["contacts"]:
                            out["contacts"].append(e)
                elif low.startswith("policy:") and not out["policy"]:
                    out["policy"] = line.split(":", 1)[1].strip()
            if out["contacts"] or out["policy"]:
                out["found"] = True
                return out
    return out

def discover_functional_mailboxes(domain: str,
    scraped: Optional[List[str]] = None,
    no_smtp: bool = False,
    extra: Optional[List[str]] = None) -> List[str]:
    """Return ranked functional/role mailboxes for a domain.

    Order of authority: security.txt-declared contacts (passed as `extra`)
    first, then addresses found on the site, then SMTP-confirmed standard
    aliases. RFC 2142 mandates abuse@ and postmaster@ exist for a domain, so
    those are probed alongside security@/privacy@. With --no-smtp, only
    scraped + declared addresses are returned plus a couple of proposed
    general mailboxes, so nothing unseen is asserted.
    """
    domain = (domain or "").strip().lower().replace("www.", "")
    if not domain:
        return []

    ordered: List[str] = []
    def _add(e):
        el = (e or "").strip().lower()
        if el and "@" in el and el not in ordered:
            ordered.append(el)

    # 1. Authoritative security.txt contacts on the same domain, first.
    for e in (extra or []):
        if e.split("@")[-1].endswith(domain) or e.endswith("@" + domain):
            _add(e)
    for e in (extra or []):
        _add(e)  # keep off-domain declared contacts too (e.g. a CSIRT alias)

    # 2. Functional addresses actually scraped from the site.
    for e in _rank_functional(x for x in (scraped or []) if _is_functional(x)):
        _add(e)
    have_local = {_local_part(e) for e in ordered}

    if no_smtp:
        for p in ("contact", "info"):
            if p not in have_local:
                _add(f"{p}@{domain}")
        return ordered

    mx = _ci_mx_for(domain)
    if not mx:
        return ordered
    # RFC 2142 mandatory aliases (abuse@, postmaster@) + security/privacy.
    for p in ("security", "abuse", "postmaster", "privacy", "dpo",
              "contact", "info"):
        if p in have_local:
            continue
        if ci_smtp_verify(f"{p}@{domain}", mx) == "smtp-ok":
            _add(f"{p}@{domain}")
            have_local.add(p)
    return ordered

def _pick_security_contact(emails: List[str], sectxt_contacts: List[str]) -> str:
    """Best vulnerability-report address: a security.txt contact if any,
    else a security/abuse/soc/cert mailbox, else empty."""
    for e in sectxt_contacts:
        if "@" in e:
            return e.lower()
    for e in emails:
        if _local_part(e) in ("security", "abuse", "soc", "cert", "psirt"):
            return e
    return ""

    # ── Scoring ───────────────────────────────────────────────────────────

def ci_score(c: CIContact) -> int:
    s  = 0
    rl = (c.role + " " + c.linkedin_role).lower()
    for kw, w in _CI_ROLE_SCORES.items():
        if kw in rl:
            s = max(s, w)

    if c.email_status == "smtp-ok":
        s = min(100, s + 5)
    elif c.email_status in ("confirmed", "confirmed (Hunter)", "confirmed (SERP)"):
        s = min(100, s + 4)
    elif c.email_status == "inferred":
        s = max(0, s - 2)

    if c.linkedin_url:
        s = min(100, s + 3)
    if c.phone_type == "practice":
        s = max(0, s - 8)
    return s

    # ── Orchestrator ──────────────────────────────────────────────────────

def ci_run_single(kbo: str, domain: str,
    hunter_key: str = "",
    apollo_key: str = "",
    delay: float    = 1.8,
    no_smtp: bool   = False,
    proxies: dict   = None) -> CIOrgProfile:
    """
    Run all 9 contact intelligence sources for one company.
    Returns a CIOrgProfile with ranked CIContact list.
    """
    if proxies is None:
        proxies = {"http": None, "https": None}
    # Strip www. — email addresses must use the root domain, not the subdomain,
    # otherwise MX lookup and SMTP verify will fail and inferred emails are wrong.
    domain = domain.replace("https://", "").replace("http://", "").strip("/")
    domain = re.sub(r"^www\.", "", domain)
    org    = CIOrgProfile(kbo=kbo, domain=domain)

    # 1 KBO
    header(f"[CI] Step 1/9  KBO  →  {kbo}")
    kbo_data       = ci_fetch_kbo(kbo, proxies, delay)
    org.name       = kbo_data.get("name", "")
    org.org_phone  = kbo_data.get("phone", "")
    org.org_email  = kbo_data.get("email", "")
    org.address    = kbo_data.get("address", "")
    if kbo_data.get("domain") and not domain:
        domain     = kbo_data["domain"]
        org.domain = domain

    # 2 Website
    header(f"[CI] Step 2/9  Website  →  {domain}")
    web_data          = ci_scrape_website(domain, proxies, delay)
    org.email_pattern = web_data.get("pattern") or f"voornaam.naam@{domain}"
    if not org.org_phone and web_data["phones"]:
        org.org_phone = web_data["phones"][0]

    # Functional / role mailboxes — preferred, non-personal contact channel.
    # Aggregate sources by authority: security.txt (declared) is authoritative;
    # DMARC rua, SOA RNAME, TLS cert and dedicated pages add same-domain
    # mailboxes; RIPE abuse-c is the hosting network's desk, kept separate.
    sectxt = fetch_security_txt(domain, proxies, delay)
    authoritative = list(sectxt.get("contacts", []))
    same_domain: List[str] = []
    for src in (dmarc_report_emails(domain),
                tls_cert_emails(domain),
                scrape_contact_pages(domain, proxies, delay)):
        for e in src:
            if e.endswith("@" + domain) and e not in same_domain:
                same_domain.append(e)
    soa = soa_rname_email(domain)
    if soa and soa.endswith("@" + domain) and soa not in same_domain:
        same_domain.append(soa)

    org.functional_emails = discover_functional_mailboxes(
        domain, web_data.get("functional", []), no_smtp,
        extra=authoritative + same_domain)
    org.security_contact = _pick_security_contact(
        org.functional_emails, sectxt.get("contacts", []))
    org.security_policy = sectxt.get("policy", "")
    org.network_abuse = whois_abuse_mailbox(domain)
    if org.functional_emails and not org.org_email:
        org.org_email = org.functional_emails[0]
    if not org.org_email:
        org.org_email = next(
            (e for e in web_data["emails"] if _is_functional(e)), "")
    if org.security_contact:
        detail(f"security contact: {org.security_contact}"
               + ("  (security.txt)" if sectxt.get("found") else ""))
    if org.network_abuse:
        detail(f"network abuse desk (host): {org.network_abuse}")
    if org.functional_emails:
        detail("functional mailboxes: " + ", ".join(org.functional_emails[:6]))

    # 3 Staatsblad
    header("[CI] Step 3/9  Staatsblad")
    sb_names = ci_fetch_staatsblad(kbo, proxies, delay)

    # Seed contact pool
    pool: Dict[str, CIContact] = {}

    def _upsert(name: str, role: str, source: str,
                phone: str = "", phone_type: str = ""):
        k = _ci_name_key(name)
        if k not in pool:
            pool[k] = CIContact(name=name)
        c = pool[k]
        if role and not c.role:
            c.role = role
        if phone and not c.phone:
            c.phone      = phone
            c.phone_type = phone_type
        if source not in c.sources:
            c.sources.append(source)
        return c

    for m in kbo_data.get("mandataries", []):
        _upsert(m["name"], m["role"], "KBO",
                phone=org.org_phone, phone_type="org")
    for e in web_data.get("staff", []) + web_data.get("board", []):
        _upsert(e["name"], e["role"], "website")
    for e in sb_names:
        _upsert(e["name"], "", "Staatsblad")

    # 4-7 SERP + LinkedIn per contact
    header(f"[CI] Step 4-7/9  SERP search + LinkedIn lookup  ({len(pool)} contact(s))")
    mx_host = _ci_mx_for(domain) if _CI_DNS else None

    # Fetch LinkedIn company page once for the org (not per-contact)
    li_company_about = ci_fetch_linkedin_company(domain, proxies, delay)
    if li_company_about:
        detail(f"company LinkedIn: {li_company_about[:70]}")

    if not pool:
        detail("no contacts to enrich for this company")

    for k, c in pool.items():
        detail(f"{c.name} — searching…")
        # Always generate a search URL — analyst can use it even if automation fails
        c.linkedin_search_url = _ci_linkedin_search_url(c.name, org.name)

        serp = ci_run_serps(c.name, org.name, domain, proxies, delay)
        if serp["linkedin_url"] and not c.linkedin_url:
            c.linkedin_url = serp["linkedin_url"]
        if serp["emails"] and not c.email:
            c.email        = serp["emails"][0]
            c.email_status = "confirmed (SERP)"
        if serp["phones"] and not c.phone:
            c.phone      = serp["phones"][0]
            c.phone_type = "direct"
        for s in serp["sources"]:
            if s not in c.sources:
                c.sources.append(s)
        if serp["mentions"] and not c.notes:
            c.notes = serp["mentions"][0]

        # If SERP found nothing, try pub/dir directly
        if not c.linkedin_url:
            parts = c.name.split()
            if len(parts) >= 2:
                c.linkedin_url = ci_try_linkedin_pubdir(
                    parts[0], parts[-1], proxies, delay)

        if c.linkedin_url:
            li = ci_fetch_linkedin(c.linkedin_url, proxies, delay * 0.8)
            if li.get("role"):
                c.linkedin_role = li["role"]
            if li.get("company"):
                c.notes = (c.notes + f" | LI: {li['role']} @ {li['company']}"
                           ).strip(" |")

        if not c.email:
            parts = c.name.split()
            if len(parts) >= 2:
                c.email        = _ci_infer_email(parts[0], parts[-1],
                                                 domain, org.email_pattern)
                c.email_status = "inferred"

        # Per-contact result: what each source produced
        serp_hits = []
        if serp.get("linkedin_url"):
            serp_hits.append("linkedin")
        if serp.get("emails"):
            serp_hits.append("email")
        if serp.get("phones"):
            serp_hits.append("phone")
        serp_summary = "+".join(serp_hits) if serp_hits else "no hits"
        if c.linkedin_url:
            li_role = getattr(c, "linkedin_role", "")
            li_summary = f"profile ({li_role})" if li_role else "profile"
        else:
            li_summary = "search-url only"
        email_summary = c.email_status or ("none" if not c.email else "found")
        detail(f"{c.name:<26}  SERP:{serp_summary}  LinkedIn:{li_summary}  "
               f"email:{email_summary}")

    # 8 External email/contact databases
    header("[CI] Step 8/9  External databases  "
           "(EmailFormat · Infobel · Gouden Gids · RIZIV · VREG · BIPT"
           " · Apollo · Hunter)")

    # 8a EmailFormat — free, no key
    ef = ci_query_emailformat(domain, proxies, delay)
    if ef.get("pattern") and not org.email_pattern:
        org.email_pattern = ef["pattern"]
    for em in ef.get("emails", []):
        k = _ci_name_key(em.split("@")[0].replace(".", " "))
        if k in pool and not pool[k].email:
            pool[k].email        = em
            pool[k].email_status = "confirmed (EmailFormat)"
            if "EmailFormat" not in pool[k].sources:
                pool[k].sources.append("EmailFormat")

    # 8b Infobel — free Belgian directory
    ib = ci_query_infobel(org.name, proxies, delay)
    if ib.get("phone") and not org.org_phone:
        org.org_phone = ib["phone"]
    if ib.get("address") and not org.address:
        org.address = ib["address"]

    # 8b2 Gouden Gids — Belgian Yellow Pages (always run)
    gg = ci_query_goudengids(org.name, proxies, delay)
    if gg.get("phone") and not org.org_phone:
        org.org_phone = gg["phone"]
    if gg.get("address") and not org.address:
        org.address = gg["address"]

    # 8b3 RIZIV — healthcare provider registry
    # (run for all: KBO alone doesn't tell us the NIS2 sector here,
    #  so we try it and let the endpoint return empty for non-health orgs)
    rz = ci_query_riziv(org.name, kbo, proxies, delay)
    if rz.get("phone") and not org.org_phone:
        org.org_phone = rz["phone"]

    # 8b4 VREG — Flemish energy regulator
    vr = ci_query_vreg(org.name, kbo, proxies, delay)
    if vr.get("phone") and not org.org_phone:
        org.org_phone = vr["phone"]

    # 8b5 BIPT — telecom / digital infrastructure regulator
    bp = ci_query_bipt(org.name, kbo, proxies, delay)
    if bp.get("phone") and not org.org_phone:
        org.org_phone = bp["phone"]

    # Propagate org phone to contacts that still have none
    if org.org_phone:
        for c in pool.values():
            if not c.phone:
                c.phone      = org.org_phone
                c.phone_type = "org"

    # 8c Apollo — optional key, 150 free credits/month
    apollo_data = ci_query_apollo(domain, apollo_key)
    for ap in apollo_data.get("emails", []):
        name = f"{ap['first']} {ap['last']}".strip()
        k    = _ci_name_key(name)
        if k in pool:
            if ap["email"] and not pool[k].email:
                pool[k].email        = ap["email"]
                pool[k].email_status = "confirmed (Apollo)"
            if ap.get("linkedin") and not pool[k].linkedin_url:
                pool[k].linkedin_url = ap["linkedin"]
            if ap.get("position") and not pool[k].role:
                pool[k].role = ap["position"]
            if "Apollo" not in pool[k].sources:
                pool[k].sources.append("Apollo")
        else:
            nc = CIContact(name=name, role=ap.get("position", ""),
                           email=ap["email"],
                           email_status="confirmed (Apollo)",
                           linkedin_url=ap.get("linkedin", ""),
                           sources=["Apollo"])
            pool[k] = nc

    # 8d Hunter.io — optional key
    hunter = ci_query_hunter(domain, hunter_key)
    if hunter.get("pattern"):
        org.email_pattern = hunter["pattern"] + "@" + domain
    for he in hunter.get("emails", []):
        name = f"{he['first']} {he['last']}".strip()
        k    = _ci_name_key(name)
        if k in pool:
            pool[k].email        = he["email"]
            pool[k].email_status = "confirmed (Hunter)"
            if he.get("position") and not pool[k].role:
                pool[k].role = he["position"]
            if "Hunter" not in pool[k].sources:
                pool[k].sources.append("Hunter")
        else:
            nc = CIContact(name=name, role=he.get("position", ""),
                           email=he["email"],
                           email_status="confirmed (Hunter)",
                           sources=["Hunter.io"])
            pool[k] = nc

    # 9 SMTP
    header(f"[CI] Step 9/9  SMTP verify  (MX: {mx_host or 'n/a'})")
    for c in pool.values():
        if no_smtp:
            break
        if c.email and c.email_status in ("inferred", ""):
            c.email_status = ci_smtp_verify(c.email, mx_host)
            info(f"  {c.email:<42} → {c.email_status}")
            time.sleep(delay * 0.25)

    for c in pool.values():
        c.score = ci_score(c)

    org.contacts = sorted(pool.values(), key=lambda x: x.score, reverse=True)
    return org

    # ── Contact output helpers ────────────────────────────────────────────

def ci_print_report(org: CIOrgProfile) -> None:
    header(f"CONTACT INTELLIGENCE  —  {org.name or org.kbo}")
    bullet(f"KBO      : {org.kbo}")
    bullet(f"Domain   : {org.domain}")
    bullet(f"Phone    : {org.org_phone}")
    bullet(f"Email    : {org.org_email}")
    bullet(f"Pattern  : {org.email_pattern}")
    bullet(f"Address  : {org.address}")
    print()
    for i, c in enumerate(org.contacts, 1):
        bar   = "█" * (c.score // 10) + "░" * (10 - c.score // 10)
        stars = "★" * round(c.score / 20) + "☆" * (5 - round(c.score / 20))
        print(f"  [{i:02d}] {stars} {c.score:>3}%  {bar}  {c.name}")
        if c.role:         bullet(f"       Role    : {c.role}")
        if c.linkedin_role and c.linkedin_role != c.role:
            bullet(f"       LI role : {c.linkedin_role}")
        bullet(f"       Email   : {c.email or '—'}  [{c.email_status}]")
        bullet(f"       Phone   : {c.phone or '—'}  ({c.phone_type})")
        if c.linkedin_url:
            bullet(f"       LinkedIn: {c.linkedin_url}")
        elif c.linkedin_search_url:
            bullet(f"       LI Search: {c.linkedin_search_url}")
        if c.sources:
            bullet(f"       Sources : {', '.join(c.sources)}")
        if c.notes:
            bullet(f"       Notes   : {c.notes[:120]}")
        print()

def ci_export_csv(org: CIOrgProfile, path: str) -> None:
    fields = ["score", "kbo", "org_name", "domain", "org_functional_emails",
    "name", "role",
    "linkedin_role", "email", "email_status",
    "phone", "phone_type", "linkedin_url", "linkedin_search_url",
    "sources", "notes"]
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for c in org.contacts:
                w.writerow({
                "score":               c.score,
                "kbo":                 org.kbo,
                "org_name":            org.name,
                "domain":              org.domain,
                "org_functional_emails": " | ".join(org.functional_emails),
                "name":                c.name,
                "role":                c.role,
                "linkedin_role":       c.linkedin_role,
                "email":               c.email,
                "email_status":        c.email_status,
                "phone":               c.phone,
                "phone_type":          c.phone_type,
                "linkedin_url":        c.linkedin_url,
                "linkedin_search_url": c.linkedin_search_url,
                "sources":             " | ".join(c.sources),
                "notes":               c.notes,
                })
                ok(f"Contact CSV : {path}")
    except OSError as e:
        warn(f"Could not write contact CSV: {e}")

def ci_export_json(org: CIOrgProfile, path: str) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
            "org": {"kbo": org.kbo, "name": org.name, "domain": org.domain,
            "phone": org.org_phone, "email": org.org_email,
            "email_pattern": org.email_pattern,
            "functional_emails": org.functional_emails,
            "security_contact": org.security_contact,
            "security_policy": org.security_policy,
            "network_abuse": org.network_abuse,
            "address": org.address},
            "contacts": [
            {"score": c.score, "name": c.name, "role": c.role,
            "linkedin_role": c.linkedin_role, "email": c.email,
            "email_status": c.email_status, "phone": c.phone,
            "phone_type": c.phone_type, "linkedin_url": c.linkedin_url,
            "linkedin_search_url": c.linkedin_search_url,
            "sources": c.sources, "notes": c.notes}
            for c in org.contacts
            ],
            }, f, indent=2, ensure_ascii=False)
            ok(f"Contact JSON: {path}")
    except OSError as e:
        warn(f"Could not write contact JSON: {e}")

def ci_export_html(orgs: list, findings_path: str, output_path: str) -> None:
    """
    Generate a self-contained HTML report with:
    - Per-company contact cards (score bar, email/phone/LinkedIn)
    - Nuclei findings table with severity badges
    - Printable, no external dependencies
    """

    # Finding descriptions and remediations come from the canonical
    # module-level FINDING_MATCHER_* tables via _finding_remediation()
    # and _finding_summary(); no local copy is kept here.
    # Load findings grouped by host
    findings_by_host: Dict[str, list] = {}
    p = Path(findings_path)
    if p.exists() and p.stat().st_size > 0:
        for f in stream_findings(p):
            host = f.get("host", "").rstrip("/")
            findings_by_host.setdefault(host, []).append(f)

    SEV_COLOR = {
        "critical": "#c0392b", "high": "#e74c3c",
        "medium":   "#e67e22", "low":  "#3498db",
        "info":     "#95a5a6", "unknown": "#bdc3c7",
    }

    def badge(sev: str) -> str:
        col = SEV_COLOR.get(sev.lower(), "#bdc3c7")
        return (f'<span style="background:{col};color:#fff;padding:2px 8px;'
                f'border-radius:3px;font-size:11px;font-weight:700;'
                f'text-transform:uppercase">{sev}</span>')

    severity_rank = {s: i for i, s in enumerate(SEV_ORDER)}

    def remediation_action(check: str, sev: str) -> str:
        return _finding_remediation(check, sev)

    def score_bar(score: int) -> str:
        pct   = max(0, min(100, score))
        color = ("#27ae60" if pct >= 70 else
                 "#e67e22" if pct >= 40 else "#e74c3c")
        return (f'<div style="display:inline-block;width:80px;height:8px;'
                f'background:#eee;border-radius:4px;vertical-align:middle">'
                f'<div style="width:{pct}%;height:100%;background:{color};'
                f'border-radius:4px"></div></div> '
                f'<small style="color:#666">{pct}%</small>')

    def mailto(email: str, status: str) -> str:
        if not email:
            return '—'
        col = ("#27ae60" if "ok" in status or "confirmed" in status
               else "#e67e22" if status == "inferred"
               else "#e74c3c" if "reject" in status else "#666")
        return (f'<a href="mailto:{email}" style="color:{col}">{email}</a>'
                f' <small style="color:#aaa">({status})</small>')

    def li_link(url: str, search_url: str) -> str:
        if url:
            return f'<a href="{url}" target="_blank">Profile ↗</a>'
        if search_url:
            return f'<a href="{search_url}" target="_blank" style="color:#aaa">Search ↗</a>'
        return '—'

    # Build org cards HTML
    cards_html = ""
    total_findings = sum(len(v) for v in findings_by_host.values())

    for org_data in orgs:
        name    = org_data.get("name", org_data.get("kbo", "?"))
        kbo     = org_data.get("kbo", "")
        domain  = org_data.get("domain", "")
        phone   = org_data.get("org_phone", org_data.get("phone", ""))
        email   = org_data.get("org_email", org_data.get("email", ""))
        address = org_data.get("address", "")
        pattern = org_data.get("email_pattern", "")
        kbo_url_str = (f"https://kbopub.economie.fgov.be/kbopub/"
                       f"toonondernemingps.html?ondernemingsnummer={kbo}")

        # Match findings to this org by domain
        org_findings = []
        for host, flist in findings_by_host.items():
            if domain and domain in host:
                org_findings.extend(flist)

        findings_count = len(org_findings)
        fcount_col = ("#c0392b" if findings_count >= 10 else
                      "#e67e22" if findings_count >= 3 else
                      "#27ae60" if findings_count == 0 else "#3498db")

        # Findings table
        findings_html = ""
        if org_findings:
            rows = ""
            action_rollup: Dict[str, dict] = {}
            for f in sorted(org_findings,
                            key=lambda x: ["critical","high","medium","low","info"]
                            .index(x.get("info",{}).get("severity","info").lower())
                            if x.get("info",{}).get("severity","info").lower()
                            in ["critical","high","medium","low","info"] else 99):
                summary  = _finding_summary(f)
                sev      = summary.get("severity", "info")
                check    = summary.get("check") or f.get("template-id", "?")
                host     = f.get("host", "")
                evidence = summary.get("evidence", "")
                desc     = summary.get("description", "")
                risk     = summary.get("risk", "")
                action   = (summary.get("remediation")
                            or remediation_action(check, sev))

                existing = action_rollup.get(action)
                if not existing:
                    action_rollup[action] = {"count": 1, "severity": sev}
                else:
                    existing["count"] += 1
                    if severity_rank.get(sev, 99) < severity_rank.get(existing["severity"], 99):
                        existing["severity"] = sev

                rows += (f"<tr><td>{badge(sev)}</td>"
                         f"<td style='font-family:monospace;font-size:12px'>{check}</td>"
                         f"<td style='color:#666;font-size:12px'>{host}</td>"
                         f"<td style='color:#666;font-size:11px;white-space:pre-wrap'>{evidence}</td>"
                         f"<td style='color:#444;font-size:11px;white-space:pre-wrap'>{desc}</td>"
                         f"<td style='color:#333;font-size:11px'>{risk}</td>"
                         f"<td style='color:#333;font-size:11px'>{action}</td></tr>")

            action_items = ""
            for action, meta in sorted(
                action_rollup.items(),
                key=lambda kv: (
                    severity_rank.get(kv[1]["severity"], 99),
                    -kv[1]["count"],
                    kv[0],
                ),
            )[:6]:
                action_items += (
                    "<li style='margin:4px 0'>"
                    f"{badge(meta['severity'])} "
                    f"{action} "
                    f"<small style='color:#888'>(matched {meta['count']} finding(s))</small>"
                    "</li>"
                )

            remediation_html = ""
            if action_items:
                remediation_html = f"""
    <h4 style="margin:14px 0 8px;color:#333">Prioritized Remediation Actions</h4>
    <ul style="margin:0 0 8px 18px;padding:0;line-height:1.45">
    {action_items}
    </ul>"""

            findings_html = f"""
    <h4 style="margin:16px 0 8px;color:#333">
    Nuclei Findings
    <span style="color:{fcount_col};font-weight:700">{findings_count}</span>
    </h4>
    {remediation_html}
    <table style="width:100%;border-collapse:collapse;font-size:13px">
    <thead><tr style="background:#f8f9fa">
    <th style="padding:6px 8px;text-align:left;width:90px">Severity</th>
    <th style="padding:6px 8px;text-align:left">Check</th>
    <th style="padding:6px 8px;text-align:left">Host</th>
    <th style="padding:6px 8px;text-align:left">Evidence</th>
    <th style="padding:6px 8px;text-align:left">Description / Executive Summary</th>
    <th style="padding:6px 8px;text-align:left">Risk</th>
    <th style="padding:6px 8px;text-align:left">Remediation Action</th>
    </tr></thead>
    <tbody>{rows}</tbody>
    </table>"""
        else:
            findings_html = (
                '<p style="color:#27ae60;margin:8px 0">✓ No findings</p>'
                if p.exists() else
                '<p style="color:#aaa;margin:8px 0;font-style:italic">'
                'Scan not yet run</p>')

        # Contact rows
        contacts = org_data.get("contacts", [])
        contact_rows = ""
        for c in sorted(contacts, key=lambda x: -x.get("score", 0)):
            cname  = c.get("name", "")
            role   = c.get("role", "") or c.get("linkedin_role", "")
            em     = c.get("email", "")
            estat  = c.get("email_status", "")
            ph     = c.get("phone", "")
            li_url = c.get("linkedin_url", "")
            li_srch= c.get("linkedin_search_url", "")
            sc     = c.get("score", 0)
            contact_rows += f"""
    <tr style="border-bottom:1px solid #f0f0f0">
    <td style="padding:8px">{score_bar(sc)}</td>
    <td style="padding:8px;font-weight:600">{cname}</td>
    <td style="padding:8px;color:#666;font-size:12px">{role}</td>
    <td style="padding:8px;font-size:13px">{mailto(em, estat)}</td>
    <td style="padding:8px;font-size:13px">{ph or '—'}</td>
    <td style="padding:8px;font-size:13px">{li_link(li_url, li_srch)}</td>
    </tr>"""

        contacts_html = ""
        if contact_rows:
            contacts_html = f"""
    <h4 style="margin:20px 0 8px;color:#333">Contacts</h4>
    <table style="width:100%;border-collapse:collapse">
    <thead><tr style="background:#f8f9fa;font-size:12px;color:#666">
    <th style="padding:6px 8px;text-align:left;width:110px">Score</th>
    <th style="padding:6px 8px;text-align:left">Name</th>
    <th style="padding:6px 8px;text-align:left">Role</th>
    <th style="padding:6px 8px;text-align:left">Email</th>
    <th style="padding:6px 8px;text-align:left">Phone</th>
    <th style="padding:6px 8px;text-align:left">LinkedIn</th>
    </tr></thead>
    <tbody>{contact_rows}</tbody>
    </table>"""

        cards_html += f"""
    <div style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
    padding:20px;margin-bottom:24px;box-shadow:0 1px 3px rgba(0,0,0,.06)">
    <div style="display:flex;justify-content:space-between;align-items:flex-start">
    <div>
    <h3 style="margin:0 0 4px;font-size:18px">{name}</h3>
    <div style="color:#666;font-size:13px">
    <a href="https://{domain}" target="_blank">{domain}</a> ·
    <a href="{kbo_url_str}" target="_blank"
    style="color:#666">KBO {kbo}</a>
    {(' · ' + address) if address else ''}
    </div>
    <div style="margin-top:6px;font-size:13px;color:#555">
    {('📞 ' + phone + ' &nbsp;') if phone else ''}
    {('✉ ' + email + ' &nbsp;') if email else ''}
    {('📧 Pattern: <code>' + pattern + '</code>') if pattern else ''}
    </div>
    </div>
    <div style="text-align:right;flex-shrink:0;margin-left:16px">
    <div style="font-size:28px;font-weight:700;color:{fcount_col}">
    {findings_count}
    </div>
    <div style="font-size:11px;color:#aaa;text-transform:uppercase">
    findings
    </div>
    </div>
    </div>
    {findings_html}
    {contacts_html}
    </div>"""

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!DOCTYPE html>

    <html lang="en">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <title>TEXCEL NIS2 Report — {now_str}</title>
    <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #f5f6fa; color: #222; line-height: 1.5; }}
    .header {{ background: #1a1a2e; color: #fff; padding: 24px 32px;
    display: flex; justify-content: space-between; align-items: center; }}
    .header h1 {{ font-size: 20px; font-weight: 600; letter-spacing: .5px; }}
    .header .meta {{ font-size: 12px; color: #aaa; text-align: right; }}
    .summary {{ background: #fff; border-bottom: 1px solid #e0e0e0;
    padding: 16px 32px; display: flex; gap: 32px; }}
    .stat {{ text-align: center; }}
    .stat .val {{ font-size: 24px; font-weight: 700; color: #1a1a2e; }}
    .stat .lbl {{ font-size: 11px; color: #aaa; text-transform: uppercase; }}
    .content {{ max-width: 1100px; margin: 24px auto; padding: 0 24px; }}
    a {{ color: #2980b9; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    @media print {{
    body {{ background: #fff; }}
    .header {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    }}
    </style>
    </head>
    <body>
    <div class="header">
    <h1>TEXCEL Solutions — NIS2 Prospect Report</h1>
    <div class="meta">
    Generated: {now_str}<br>
    {len(orgs)} companies · {total_findings} findings
    </div>
    </div>
    <div class="summary">
    <div class="stat">
    <div class="val">{len(orgs)}</div>
    <div class="lbl">Companies</div>
    </div>
    <div class="stat">
    <div class="val">{total_findings}</div>
    <div class="lbl">Total Findings</div>
    </div>
    <div class="stat">
    <div class="val">{sum(len(o.get("contacts",[])) for o in orgs)}</div>
    <div class="lbl">Contacts</div>
    </div>
    <div class="stat">
    <div class="val">{sum(1 for o in orgs
    for c in o.get("contacts",[])
    if "smtp-ok" in c.get("email_status","")
    or "confirmed" in c.get("email_status",""))}</div>
    <div class="lbl">Verified Emails</div>
    </div>
    </div>
    <div class="content">
    {cards_html}
    </div>
    </body>
    </html>"""

    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        ok(f"HTML report    : {output_path}")
    except OSError as e:
        warn(f"Could not write HTML report: {e}")

def _split_recipients(raw: str) -> List[str]:
    if not raw:
        return []
    items: List[str] = []
    for tok in re.split(r"[;,]", raw):
        addr = tok.strip()
        if addr and addr not in items:
            items.append(addr)
    return items

def _build_report_summary(nuclei_output: str,
    scanned_hosts: Optional[List[str]] = None,
    orgs: Optional[list] = None) -> dict:
    findings: List[dict] = []
    p = Path(nuclei_output)
    if p.exists() and p.stat().st_size > 0:
        findings = list(stream_findings(p))

    sev = Counter(
        f.get("info", {}).get("severity", "unknown").lower()
        for f in findings
    )
    host_counts = Counter(f.get("host", "").rstrip("/") for f in findings
                          if f.get("host"))
    scanned = [h.rstrip("/") for h in (scanned_hosts or []) if h]
    scanned_count = len(scanned) if scanned else len(host_counts)
    affected = sum(1 for _, cnt in host_counts.items() if cnt > 0)
    contacts_total = 0
    verified_emails = 0
    company_count = 0

    if orgs:
        company_count = len(orgs)
        for org in orgs:
            contacts = org.get("contacts", [])
            contacts_total += len(contacts)
            for c in contacts:
                st = str(c.get("email_status", "")).lower()
                if ("smtp-ok" in st) or ("confirmed" in st):
                    verified_emails += 1

    top_hosts = [{"host": h, "findings": c}
                 for h, c in host_counts.most_common(10)]

    return {
        "companies": company_count,
        "hosts_scanned": scanned_count,
        "hosts_affected": affected,
        "findings_total": len(findings),
        "severity": {k: sev.get(k, 0) for k in SEV_ORDER},
        "contacts_total": contacts_total,
        "verified_emails": verified_emails,
        "top_hosts": top_hosts,
    }

def _write_scan_brief_html(summary: dict, output_path: str) -> str:
    sev = summary.get("severity", {})
    top_rows = ""
    for row in summary.get("top_hosts", []):
        top_rows += (
            f"<tr><td style='padding:6px 8px'>{row.get('host','')}</td>"
            f"<td style='padding:6px 8px;text-align:right'>{row.get('findings',0)}</td></tr>"
        )
    if not top_rows:
        top_rows = ("<tr><td style='padding:6px 8px;color:#777' colspan='2'>"
                    "No findings</td></tr>")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>NIS2 Executive Report</title></head>
<body style="font-family:Segoe UI,Arial,sans-serif;background:#f5f6fa;color:#222;margin:0">
<div style="background:#1a1a2e;color:#fff;padding:18px 24px">
<h2 style="margin:0">NIS2 Executive Report</h2>
<div style="font-size:12px;color:#cfd3e3">Generated: {now}</div></div>
<div style="padding:20px 24px">
<div style="display:flex;gap:22px;flex-wrap:wrap;margin-bottom:14px">
<div><strong>{summary.get('hosts_scanned',0)}</strong><div style="font-size:12px;color:#777">Hosts scanned</div></div>
<div><strong>{summary.get('hosts_affected',0)}</strong><div style="font-size:12px;color:#777">Hosts affected</div></div>
<div><strong>{summary.get('findings_total',0)}</strong><div style="font-size:12px;color:#777">Findings</div></div>
<div><strong>{summary.get('contacts_total',0)}</strong><div style="font-size:12px;color:#777">Contacts</div></div>
<div><strong>{summary.get('verified_emails',0)}</strong><div style="font-size:12px;color:#777">Verified emails</div></div>
</div>
<h3 style="margin:16px 0 8px">Severity breakdown</h3>
<ul style="line-height:1.6">
<li>Critical: {sev.get('critical',0)}</li>
<li>High: {sev.get('high',0)}</li>
<li>Medium: {sev.get('medium',0)}</li>
<li>Low: {sev.get('low',0)}</li>
<li>Info: {sev.get('info',0)}</li>
</ul>
<h3 style="margin:18px 0 8px">Top affected hosts</h3>
<table style="border-collapse:collapse;width:100%;max-width:860px;background:#fff">
<thead><tr style="background:#f1f3f6"><th style="padding:6px 8px;text-align:left">Host</th>
<th style="padding:6px 8px;text-align:right">Findings</th></tr></thead>
<tbody>{top_rows}</tbody></table></div></body></html>"""
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        ok(f"Brief HTML report: {output_path}")
        return output_path
    except OSError as e:
        warn(f"Could not write brief HTML report: {e}")
        return ""

def _attachment_payload(path: Path, max_bytes: int = 5 * 1024 * 1024) -> Optional[dict]:
    if not path.exists() or not path.is_file():
        return None
    size = path.stat().st_size
    if size > max_bytes:
        warn(f"Skipping attachment > {max_bytes} bytes: {path.name}")
        return {
            "name": path.name,
            "path": str(path.resolve()),
            "size_bytes": size,
            "skipped": "too-large",
        }
    try:
        with open(path, "rb") as f:
            blob = f.read()
        mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        return {
            "name": path.name,
            "path": str(path.resolve()),
            "size_bytes": size,
            "mime_type": mime,
            "content_base64": base64.b64encode(blob).decode("ascii"),
        }
    except OSError as e:
        warn(f"Could not read attachment {path}: {e}")
        return None

def send_report_via_power_automate(webhook_url: str,
    summary: dict,
    report_path: str,
    subject: str = "",
    recipients_raw: str = "",
    attachment_paths: Optional[List[str]] = None,
    timeout: int = 20) -> bool:
    if not webhook_url:
        return False

    recipients = _split_recipients(recipients_raw)
    report_html = ""
    rp = Path(report_path) if report_path else None
    if rp and rp.exists():
        try:
            report_html = rp.read_text(encoding="utf-8", errors="replace")
        except OSError as e:
            warn(f"Could not read report HTML ({rp}): {e}")

    attachments = []
    for ap in (attachment_paths or []):
        item = _attachment_payload(Path(ap))
        if item:
            attachments.append(item)

    final_subject = (subject.strip() if subject.strip() else
                     f"NIS2 report {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    payload = {
        "source": "nis2_scanner",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "subject": final_subject,
        "to": recipients,
        "summary": summary,
        "report": {
            "path": str(rp.resolve()) if rp and rp.exists() else "",
            "name": rp.name if rp else "",
            "html": report_html,
        },
        "attachments": attachments,
    }

    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        webhook_url.strip(),
        data=body,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=max(5, int(timeout))) as resp:
            status = getattr(resp, "status", resp.getcode())
            resp_text = resp.read().decode("utf-8", errors="replace")[:240]
        if 200 <= status < 300:
            ok("Power Automate webhook accepted report payload.")
            if resp_text:
                info(f"Webhook response: {resp_text}")
            return True
        warn(f"Power Automate webhook returned status {status}")
        return False
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode("utf-8", errors="replace")[:240]
        except Exception:
            err_body = ""
        warn(f"Power Automate HTTP error {e.code}: {err_body}")
        return False
    except Exception as e:
        warn(f"Power Automate delivery failed: {e}")
        return False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SharePoint upload (Microsoft Graph, client-credentials flow)
#
# Credentials are read from the environment, never from CLI flags:
#   SP_TENANT_ID        Entra directory (tenant) ID
#   SP_CLIENT_ID        app registration (client) ID
#   SP_CLIENT_SECRET    client secret value
#   SP_HOSTNAME         e.g. contoso.sharepoint.com
#   SP_SITE_PATH        e.g. /sites/NIS2
#   SP_TARGET_FOLDER    optional, default "NIS2-Scans"
#
# The app needs the Graph application permission Sites.Selected, plus a
# per-site "write" grant on the target site. Sites.Selected grants nothing
# until that per-site authorisation is added, so it is the least-privilege
# option here — do not substitute Sites.ReadWrite.All.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
_SP_SIMPLE_PUT_LIMIT = 4 * 1024 * 1024          # 4 MiB
_SP_CHUNK = 8 * 320 * 1024                       # 2.5 MiB, multiple of 320 KiB


class _SharePointClient:
    """Minimal Graph client using only the standard library."""

    def __init__(self, tenant_id, client_id, client_secret, hostname, site_path):
        self.tenant_id = tenant_id
        self.client_id = client_id
        self.client_secret = client_secret
        # Tolerate a pasted scheme/trailing slash: Graph needs the bare host
        # (e.g. contoso.sharepoint.com), not https://contoso.sharepoint.com/.
        host = str(hostname).strip()
        host = re.sub(r"^https?://", "", host, flags=re.IGNORECASE)
        host = host.split("/", 1)[0].rstrip("/").strip()
        self.hostname = host
        self.site_path = site_path if site_path.startswith("/") else "/" + site_path
        self._token = None
        self._token_expiry = 0.0
        self._site_id = None
        self._drive_id = None

    # -- auth ----------------------------------------------------------------
    def _access_token(self):
        if self._token and time.time() < self._token_expiry - 60:
            return self._token
        body = urllib.parse.urlencode({
            "grant_type": "client_credentials",
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "scope": "https://graph.microsoft.com/.default",
        }).encode()
        url = (f"https://login.microsoftonline.com/{self.tenant_id}"
               "/oauth2/v2.0/token")
        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("Content-Type", "application/x-www-form-urlencoded")
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                data = json.loads(r.read())
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")
            raise RuntimeError(f"token request failed ({e.code}): {detail}") from None
        self._token = data["access_token"]
        self._token_expiry = time.time() + int(data.get("expires_in", 3600))
        return self._token

    # -- low-level request ---------------------------------------------------
    def _graph(self, method, url, *, body=None, content_type=None, raw=False,
               use_bearer=True):
        req = urllib.request.Request(url, data=body, method=method)
        if use_bearer:
            req.add_header("Authorization", f"Bearer {self._access_token()}")
        if content_type:
            req.add_header("Content-Type", content_type)
        try:
            with urllib.request.urlopen(req, timeout=120) as r:
                payload = r.read()
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")
            if e.code == 403:
                raise PermissionError(
                    "Graph returned 403. The token is valid but the app is not "
                    "authorised on this site — add a Sites.Selected 'write' grant "
                    "for the app on the target site."
                ) from None
            raise RuntimeError(f"{method} {url} failed ({e.code}): {detail}") from None
        if raw:
            return payload
        return json.loads(payload) if payload else {}

    # -- resource resolution -------------------------------------------------
    @property
    def site_id(self):
        if self._site_id is None:
            url = f"{_GRAPH_ROOT}/sites/{self.hostname}:{self.site_path}"
            self._site_id = self._graph("GET", url)["id"]
        return self._site_id

    @property
    def drive_id(self):
        if self._drive_id is None:
            url = f"{_GRAPH_ROOT}/sites/{self.site_id}/drive"
            self._drive_id = self._graph("GET", url)["id"]
        return self._drive_id

    # -- upload --------------------------------------------------------------
    def upload_file(self, local_path, remote_path):
        local_path = Path(local_path)
        size = local_path.stat().st_size
        enc_remote = urllib.parse.quote(remote_path)
        if size <= _SP_SIMPLE_PUT_LIMIT:
            url = f"{_GRAPH_ROOT}/drives/{self.drive_id}/root:/{enc_remote}:/content"
            with open(local_path, "rb") as fh:
                item = self._graph("PUT", url, body=fh.read(),
                                   content_type="application/octet-stream")
            return item.get("webUrl", "")
        return self._upload_large(local_path, enc_remote, size)

    def _upload_large(self, local_path, enc_remote, size):
        session_url = (f"{_GRAPH_ROOT}/drives/{self.drive_id}/root:/"
                       f"{enc_remote}:/createUploadSession")
        session = self._graph(
            "POST", session_url,
            body=json.dumps({"item": {"@microsoft.graph.conflictBehavior": "replace"}}).encode(),
            content_type="application/json",
        )
        upload_url = session["uploadUrl"]
        web_url = ""
        with open(local_path, "rb") as fh:
            offset = 0
            while offset < size:
                chunk = fh.read(_SP_CHUNK)
                end = offset + len(chunk) - 1
                # The upload session URL is pre-authorised; no bearer token here.
                req = urllib.request.Request(upload_url, data=chunk, method="PUT")
                req.add_header("Content-Length", str(len(chunk)))
                req.add_header("Content-Range", f"bytes {offset}-{end}/{size}")
                with urllib.request.urlopen(req, timeout=120) as r:
                    if r.status in (200, 201):
                        web_url = json.loads(r.read()).get("webUrl", "")
                offset = end + 1
        return web_url


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Scan ledger — global record of scanned targets, campaign-independent.
# Deduplicates so a target is never scanned twice across campaigns. The
# local file is the source of truth for the check; a copy is pushed to a
# FIXED SharePoint path (overwrite) each run.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SCAN_LEDGER_DEFAULT = "scan_ledger.csv"
_LEDGER_FIELDS = ["domain", "entity", "company", "sector",
                  "first_scanned", "last_scanned", "scan_count",
                  "last_campaign", "last_findings"]

def _ledger_key(url_or_host: str) -> str:
    s = str(url_or_host or "").strip().lower()
    if not s:
        return ""
    netloc = urlparse(s if "://" in s else "http://" + s).netloc or s
    return netloc.replace("www.", "").strip("/")

def load_scan_ledger(path: str) -> Dict[str, dict]:
    p = Path(path)
    if not p.exists():
        return {}
    out: Dict[str, dict] = {}
    try:
        with open(p, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                k = (row.get("domain") or "").strip().lower()
                if k:
                    out[k] = row
    except Exception as e:  # noqa: BLE001
        warn(f"Could not read scan ledger {path}: {e}")
    return out

def filter_unscanned(urls: List[str], ledger: Dict[str, dict]
                     ) -> Tuple[List[str], List[str]]:
    to_scan, skipped = [], []
    seen = set()
    for u in urls:
        k = _ledger_key(u)
        if k and k in ledger:
            skipped.append(u)
        elif k and k in seen:
            continue
        else:
            seen.add(k)
            to_scan.append(u)
    return to_scan, skipped

def _findings_count_by_domain(nuclei_output: str) -> Counter:
    counts: Counter = Counter()
    p = Path(nuclei_output)
    if p.exists() and p.stat().st_size > 0:
        for finding in stream_findings(p):
            counts[_ledger_key(_finding_host(finding))] += 1
    return counts

def update_scan_ledger(path: str,
    scanned_urls: List[str],
    lookup: dict,
    hostname_index: dict,
    campaign: str,
    nuclei_output: str) -> str:
    """Append/refresh ledger rows for the targets scanned this run, then
    rewrite the file atomically. Idempotent per domain (scan_count increments)."""
    ledger = load_scan_ledger(path)
    counts = _findings_count_by_domain(nuclei_output)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    for u in scanned_urls:
        k = _ledger_key(u)
        if not k:
            continue
        co = _company_record(u, lookup, hostname_index)
        row = ledger.get(k, {"domain": k, "first_scanned": now, "scan_count": "0"})
        row["entity"]  = co.get("entity") or row.get("entity", "")
        row["company"] = co.get("name") or row.get("company", "")
        row["sector"]  = co.get("sector") or row.get("sector", "")
        row["last_scanned"] = now
        try:
            row["scan_count"] = str(int(row.get("scan_count", "0") or 0) + 1)
        except ValueError:
            row["scan_count"] = "1"
        row["last_campaign"] = campaign
        row["last_findings"] = str(counts.get(k, 0))
        row.setdefault("first_scanned", now)
        ledger[k] = row

    dest = Path(path)
    if dest.parent and not dest.parent.exists():
        dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".tmp")
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_LEDGER_FIELDS)
        w.writeheader()
        for k in sorted(ledger):
            w.writerow({fld: ledger[k].get(fld, "") for fld in _LEDGER_FIELDS})
    tmp.replace(dest)
    ok(f"Scan ledger updated: {dest}  ({len(ledger)} target(s) tracked)")
    return str(dest)

def upload_scan_ledger_to_sharepoint(ledger_path: str,
    enabled: bool,
    target_folder: str = "") -> bool:
    """Upload the ledger to a FIXED SharePoint path so it overwrites each run.

    Lives at <folder>/scan_ledger.csv (no timestamp), unlike scan artifacts
    which go under a per-run timestamped folder.
    """
    if not enabled:
        return False
    if not Path(ledger_path).exists():
        return False
    tenant_id = os.environ.get("SP_TENANT_ID", "").strip()
    client_id = os.environ.get("SP_CLIENT_ID", "").strip()
    client_secret = os.environ.get("SP_CLIENT_SECRET", "").strip()
    hostname = os.environ.get("SP_HOSTNAME", "").strip()
    site_path = os.environ.get("SP_SITE_PATH", "").strip()
    if not all((tenant_id, client_id, client_secret, hostname, site_path)):
        return False
    folder = (target_folder
              or os.environ.get("SP_TARGET_FOLDER", "").strip()
              or "NIS2-Scans")
    remote = f"{folder.strip('/')}/scan_ledger.csv"
    try:
        client = _SharePointClient(tenant_id, client_id, client_secret,
                                   hostname, site_path)
        client.upload_file(ledger_path, remote)  # PUT /content overwrites
        ok(f"Scan ledger uploaded → {remote}")
        return True
    except PermissionError as e:
        warn(f"Ledger upload not authorised: {e}")
        return False
    except Exception as e:  # noqa: BLE001
        warn(f"Ledger upload failed: {e}")
        return False

def maybe_upload_to_sharepoint(output_dir: str,
    nuclei_output: str,
    enabled: bool,
    target_folder: str = "",
    scanned_hosts: Optional[List[str]] = None,
    preferred_report_path: str = "",
    include_report_files: bool = False,
    extra_paths: Optional[List[str]] = None,
    only_extra_paths: bool = False) -> bool:
    """
    Upload scan artifacts to a SharePoint document library via Graph.
    No-op unless `enabled` is set and the SP_* environment variables are present.
    Files land under <target_folder>/<UTC timestamp>/ so runs never overwrite.
    """
    if not enabled:
        return False

    tenant_id = os.environ.get("SP_TENANT_ID", "").strip()
    client_id = os.environ.get("SP_CLIENT_ID", "").strip()
    client_secret = os.environ.get("SP_CLIENT_SECRET", "").strip()
    hostname = os.environ.get("SP_HOSTNAME", "").strip()
    site_path = os.environ.get("SP_SITE_PATH", "").strip()

    missing = [name for name, val in (
        ("SP_TENANT_ID", tenant_id), ("SP_CLIENT_ID", client_id),
        ("SP_CLIENT_SECRET", client_secret), ("SP_HOSTNAME", hostname),
        ("SP_SITE_PATH", site_path),
    ) if not val]
    if missing:
        warn(f"SharePoint upload requested but missing env var(s): "
             f"{', '.join(missing)} — skipping upload.")
        return False

    folder = (target_folder
              or os.environ.get("SP_TARGET_FOLDER", "").strip()
              or "NIS2-Scans")

    # Collect the same artifact set the Power Automate delivery uses.
    out_dir = Path(output_dir or ".")
    candidates: List[Path] = []
    if only_extra_paths:
        # Per-company split mode: deliver ONLY the by_company/ tree so no
        # uploaded file spans multiple companies. Skip the HTML report, the
        # aggregates, and the raw nuclei JSONL — all of which are multi-company.
        for extra in (extra_paths or []):
            if extra:
                candidates.append(Path(extra))
    else:
        if preferred_report_path:
            candidates.append(Path(preferred_report_path))
        for name in (SCAN_RESULTS_HTML, "nis2_report.html", "nis2_summary_brief.html"):
            candidates.append(out_dir / name)
        candidates.extend(sorted(out_dir.glob("report_*.html")))
        if include_report_files:
            for name in (
                SCAN_RESULTS_JSON, SCAN_RESULTS_CSV, SCAN_RESULTS_HTML,
                "full_coverage_report.csv", "full_coverage_report.xlsx",
                "combined_contacts.csv", "combined_contacts.json",
                "contact_enrichment.csv", "contact_enrichment.json",
                "nis2_companies_manifest.csv", "step_timings.json",
            ):
                candidates.append(out_dir / name)
        if nuclei_output:
            candidates.append(Path(nuclei_output))
        for extra in (extra_paths or []):
            if extra:
                candidates.append(Path(extra))

    # De-duplicate, keep only existing non-empty files.
    seen = set()
    files: List[Path] = []
    for c in candidates:
        try:
            rp = c.resolve()
        except OSError:
            rp = c
        if rp in seen:
            continue
        seen.add(rp)
        if c.exists() and c.is_file() and c.stat().st_size > 0:
            files.append(c)

    if not files:
        warn("SharePoint upload: no artifacts found to upload — skipping.")
        return False

    header("SHAREPOINT UPLOAD")
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%SZ")
    remote_dir = f"{folder.strip('/')}/{stamp}"

    try:
        client = _SharePointClient(tenant_id, client_id, client_secret,
                                   hostname, site_path)
        bullet(f"Site   : {client.hostname}{client.site_path}")
        bullet(f"Folder : {remote_dir}")
        info(f"Uploading {len(files)} artifact(s)…")

        uploaded = 0
        uploaded_bytes = 0
        first_web_url = ""
        failures: List[str] = []
        out_root = Path(output_dir or ".").resolve()
        for f in files:
            # Preserve any per-company subfolder structure under output_dir so
            # each company's files stay grouped and same-named files (e.g.
            # findings_*.txt) never collide in one flat folder.
            try:
                rel = Path(f).resolve().relative_to(out_root)
                suffix = str(rel).replace("\\", "/")
            except (ValueError, OSError):
                suffix = f.name
            remote_path = f"{remote_dir}/{suffix}"
            try:
                web_url = client.upload_file(f, remote_path)
                uploaded += 1
                uploaded_bytes += f.stat().st_size
                first_web_url = first_web_url or web_url
                ok(f"  {f.name}")
            except PermissionError as e:
                # A 403/401 on the first file is systemic (auth/grant), not a
                # per-file problem — report it once and stop rather than
                # repeating the same error for every remaining artifact.
                error("Upload not authorised — stopping.")
                bullet(str(e))
                return False
            except Exception as e:  # noqa: BLE001 — one bad file must not abort the rest
                failures.append(f"{f.name}: {e}")

        # Verdict
        if uploaded == 0:
            error(f"Upload failed — 0/{len(files)} artifact(s) sent.")
            for msg in failures[:3]:
                bullet(msg)
            if len(failures) > 3:
                bullet(f"…and {len(failures) - 3} more.")
            return False

        size_mb = uploaded_bytes / (1024 * 1024)
        ok(f"Uploaded {uploaded}/{len(files)} artifact(s) "
           f"({size_mb:.1f} MB) to SharePoint.")
        if failures:
            warn(f"{len(failures)} artifact(s) failed:")
            for msg in failures[:3]:
                bullet(msg)
            if len(failures) > 3:
                bullet(f"…and {len(failures) - 3} more.")
        if first_web_url:
            bullet(f"View: {first_web_url.rsplit('/', 1)[0]}")
        return True
    except PermissionError as e:
        error("Upload not authorised.")
        bullet(str(e))
        return False
    except Exception as e:  # noqa: BLE001
        error(f"SharePoint upload failed: {e}")
        return False


def maybe_deliver_report(output_dir: str,
    nuclei_output: str,
    webhook_url: str,
    outlook_to: str = "",
    outlook_subject: str = "",
    scanned_hosts: Optional[List[str]] = None,
    orgs: Optional[list] = None,
    preferred_report_path: str = "",
    include_report_files: bool = False,
    extra_attachment_paths: Optional[List[str]] = None,
    timeout: int = 20) -> bool:
    if not webhook_url:
        return False

    out_dir = Path(output_dir or ".")
    report_candidates: List[Path] = []
    if preferred_report_path:
        report_candidates.append(Path(preferred_report_path))
    report_candidates.append(out_dir / SCAN_RESULTS_HTML)
    report_candidates.append(out_dir / "nis2_report.html")
    report_candidates.extend(sorted(out_dir.glob("report_*.html")))

    report_path = ""
    for cand in report_candidates:
        if cand.exists():
            report_path = str(cand)
            break

    summary = _build_report_summary(nuclei_output, scanned_hosts=scanned_hosts, orgs=orgs)
    if not report_path:
        report_path = _write_scan_brief_html(summary, str(out_dir / "nis2_summary_brief.html"))
        if not report_path:
            warn("No HTML report available for delivery.")
            return False

    attachment_paths = [report_path]
    if include_report_files:
        for name in (
            SCAN_RESULTS_JSON,
            SCAN_RESULTS_CSV,
            SCAN_RESULTS_HTML,
            "full_coverage_report.csv",
            "full_coverage_report.xlsx",
            "combined_contacts.csv",
            "combined_contacts.json",
            "step_timings.json",
        ):
            p = out_dir / name
            if p.exists():
                attachment_paths.append(str(p))
    for extra in (extra_attachment_paths or []):
        if extra:
            attachment_paths.append(extra)

    # Preserve order, remove duplicates
    uniq = list(dict.fromkeys(attachment_paths))
    return send_report_via_power_automate(
        webhook_url=webhook_url,
        summary=summary,
        report_path=report_path,
        subject=outlook_subject,
        recipients_raw=outlook_to,
        attachment_paths=uniq,
        timeout=timeout,
    )

    # ── Post-scan enrichment ──────────────────────────────────────────────

def ci_enrich_from_scan(output_dir: str,
    nuclei_output: str,
    contact_limit: int,
    hunter_key: str,
    apollo_key: str,
    serp_delay: float,
    no_smtp: bool,
    workers: int,
    proxies: dict) -> None:
    """
    After a NIS2 scan, load the manifest + findings and run contact intel
    on the top `contact_limit` companies ordered by finding count.
    """
    header("POST-SCAN CONTACT ENRICHMENT")

    manifest_path = Path(output_dir) / "nis2_companies_manifest.csv"
    combined_path = Path(output_dir) / "contact_enrichment.csv"
    json_path     = Path(output_dir) / "contact_enrichment.json"
    if not manifest_path.exists():
        warn(f"Manifest not found: {manifest_path} – run a scan first.")
        return

    def _norm_kbo(value: str) -> str:
        return str(value or "").replace(".", "").strip()

    # Existing enrichment outputs are treated as "already done" so re-runs
    # only select companies that were not enriched before.
    all_orgs: dict = {}
    done_kbos: set = set()

    def _load_done_kbos(path: Path, include_orgs: bool = False) -> Tuple[set, dict]:
        found: set = set()
        loaded_orgs: dict = {}
        try:
            with open(path, encoding="utf-8") as f:
                prev_orgs = json.load(f)
            if not isinstance(prev_orgs, list):
                return found, loaded_orgs
            for org in prev_orgs:
                if not isinstance(org, dict):
                    continue
                kbo = _norm_kbo(org.get("kbo", ""))
                if not kbo:
                    continue
                found.add(kbo)
                if include_orgs:
                    loaded_orgs[kbo] = {**org, "kbo": kbo}
        except (json.JSONDecodeError, OSError) as e:
            warn(f"Could not load contact enrichment JSON '{path}': {e}")
        return found, loaded_orgs

    if json_path.exists():
        local_done, local_orgs = _load_done_kbos(json_path, include_orgs=True)
        done_kbos.update(local_done)
        all_orgs.update(local_orgs)

    # Also inspect child folders of the current working directory so selection
    # can skip companies already enriched in sibling campaign folders.
    skip_dirs = {
        ".git", ".hg", ".svn", "__pycache__", ".pytest_cache", ".mypy_cache",
        ".venv", "venv", "node_modules",
    }
    extra_sources = 0
    extra_kbos = 0
    try:
        local_json_resolved = json_path.resolve()
    except OSError:
        local_json_resolved = json_path
    for root, dirs, files in os.walk(Path.cwd()):
        dirs[:] = [d for d in dirs if d not in skip_dirs and not d.startswith(".")]
        if "contact_enrichment.json" not in files:
            continue
        candidate = Path(root) / "contact_enrichment.json"
        try:
            candidate_resolved = candidate.resolve()
        except OSError:
            candidate_resolved = candidate
        if candidate_resolved == local_json_resolved:
            continue
        found, _ = _load_done_kbos(candidate, include_orgs=False)
        if not found:
            continue
        before = len(done_kbos)
        done_kbos.update(found)
        added = len(done_kbos) - before
        if added:
            extra_sources += 1
            extra_kbos += added

    if done_kbos:
        bullet(f"Selection filter : skip {len(done_kbos):,} previously "
               "enriched companies")
    if extra_sources:
        bullet(f"Selection scope  : +{extra_kbos:,} from {extra_sources:,} "
               "child folder enrichment file(s)")

    all_contacts: List[dict] = []
    if combined_path.exists():
        try:
            with open(combined_path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    if not row:
                        continue
                    # Keep numeric sort deterministic when re-writing the file.
                    for num_key in ("findings", "score"):
                        try:
                            row[num_key] = int(row.get(num_key, 0))
                        except (TypeError, ValueError):
                            row[num_key] = 0
                    all_contacts.append(row)
        except (OSError, csv.Error) as e:
            warn(f"Could not load existing combined contacts CSV: {e}")

    # Load manifest: EntityNumber, CompanyName, NaceCode, NIS2_Sector,
    #                Website, KBO_URL
    try:
        manifest_df = pd.read_csv(manifest_path, dtype=str)
    except Exception as e:
        warn(f"Could not read manifest: {e}"); return

    # Count findings per host
    host_findings: Counter = Counter()
    p = Path(nuclei_output)
    if p.exists():
        for finding in stream_findings(p):
            host_findings[finding.get("host", "").rstrip("/")] += 1
    else:
        warn("No nuclei results file found – enriching all manifest companies.")

    # Load URL→company lookup
    lookup, hostname_index = load_url_lookup(output_dir)

    # Rank companies by finding count
    ranked: List[dict] = []
    for host, cnt in sorted(host_findings.items(), key=lambda x: -x[1]):
        co = resolve_company(host, lookup, hostname_index)
        if co:
            if _norm_kbo(co.get("entity", "")) in done_kbos:
                continue
            ranked.append({**co, "findings": cnt, "host": host})
    # Also add companies with 0 findings if not enough ranked.
    # itertuples is ~10x faster than iterrows and breaks early once
    # contact_limit is reached — avoids scanning the full manifest.
    seen_entities_norm = {_norm_kbo(r["entity"]) for r in ranked}
    if len(ranked) < contact_limit:
        needed_cols = ["EntityNumber", "CompanyName", "NaceCode",
                       "NIS2_Sector", "KBO_URL", "Website"]
        present_cols = [c for c in needed_cols if c in manifest_df.columns]
        sub = manifest_df[present_cols].copy()
        sub["EntityNumber"] = sub["EntityNumber"].astype(str).str.strip()
        sub["_EntityNumberNorm"] = sub["EntityNumber"].map(_norm_kbo)
        sub = sub[~sub["_EntityNumberNorm"].isin(done_kbos | seen_entities_norm)]
        for row in sub.itertuples(index=False):
            ent     = getattr(row, "EntityNumber", "")
            ent_norm= _norm_kbo(ent)
            if not ent_norm or ent_norm in done_kbos:
                continue
            website = str(getattr(row, "Website", "")).strip()
            ranked.append({
                "entity":   ent,
                "name":     str(getattr(row, "CompanyName", "")),
                "nace":     str(getattr(row, "NaceCode",    "")),
                "sector":   str(getattr(row, "NIS2_Sector", "")),
                "kbo_url":  str(getattr(row, "KBO_URL",     "")),
                "findings": 0,
                "host":     website,
            })
            seen_entities_norm.add(ent_norm)
            if len(ranked) >= contact_limit:
                break

    targets = ranked[:contact_limit]
    ok(f"Enriching top {len(targets)} companies  "
       f"(of {len(ranked)} total, ordered by finding count)")

    # Deduplicate targets by KBO before starting — manifest bug can produce dupes
    seen_kbos: set = set()
    deduped_targets = []
    for co in targets:
        k = co.get("entity", "").replace(".", "")
        if k and k not in seen_kbos:
            seen_kbos.add(k)
            deduped_targets.append(co)
        elif k:
            warn(f"Duplicate KBO {k} in target list — skipping second occurrence.")
    targets = deduped_targets

    max_workers = max(1, int(workers or 1))
    info(f"Contact enrichment workers: {min(max_workers, len(targets))}")

    def _enrich_one(item: Tuple[int, dict]) -> dict:
        i, co = item
        kbo = co.get("entity", "").replace(".", "")
        domain = urlparse(co.get("host", "")).netloc.replace("www.", "")
        if not domain:
            raw = co.get("host", "")
            domain = raw.replace("https://", "").replace("http://", "").split("/")[0]
        if not kbo or not domain:
            return {
                "idx": i, "co": co, "kbo": kbo, "domain": domain,
                "org": None,
                "error": (f"Skipping {co.get('name','')} — missing KBO ({kbo!r}) "
                          f"or domain ({domain!r})"),
            }
        try:
            org = ci_run_single(
                kbo, domain,
                hunter_key=hunter_key,
                apollo_key=apollo_key,
                delay=serp_delay,
                no_smtp=no_smtp,
                proxies=proxies,
            )
            return {"idx": i, "co": co, "kbo": kbo, "domain": domain,
                    "org": org, "error": ""}
        except Exception as exc:
            return {"idx": i, "co": co, "kbo": kbo, "domain": domain,
                    "org": None, "error": str(exc)}

    jobs = list(enumerate(targets, 1))
    results: List[dict] = []
    try:
        if max_workers == 1 or len(jobs) <= 1:
            for job in jobs:
                results.append(_enrich_one(job))
        else:
            with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(max_workers, len(jobs))
            ) as ex:
                futures = {ex.submit(_enrich_one, job): job[0] for job in jobs}
                for fut in _pbar(
                    concurrent.futures.as_completed(futures),
                    total=len(futures),
                    unit="org",
                    desc="CI enrich",
                ):
                    try:
                        results.append(fut.result())
                    except Exception as exc:
                        warn(f"  CI worker failure: {exc}")
    except KeyboardInterrupt:
        warn("Enrichment interrupted.")

    for res in sorted(results, key=lambda x: x["idx"]):
        i = res["idx"]
        co = res["co"]
        kbo = res["kbo"]
        org = res["org"]
        if not org:
            if res["error"]:
                warn(f"  [{i}/{len(targets)}] {res['error']}")
            continue

        info(f"\n[CI {i}/{len(targets)}] {co.get('name','')}  "
             f"KBO={kbo}  domain={res['domain']}  findings={co.get('findings',0)}")
        ci_print_report(org)
        stem = re.sub(r"[^\w]", "_", org.name or kbo)[:40]
        ci_export_csv(org, str(Path(output_dir) / f"contacts_{stem}.csv"))
        ci_export_json(org, str(Path(output_dir) / f"contacts_{stem}.json"))

        for c in org.contacts:
            all_contacts.append({
                "findings":      co.get("findings", 0),
                "score":         c.score,
                "kbo":           kbo,
                "org_name":      org.name,
                "domain":        org.domain,
                "sector":        co.get("sector", ""),
                "name":          c.name,
                "role":          c.role,
                "linkedin_role": c.linkedin_role,
                "email":         c.email,
                "email_status":  c.email_status,
                "phone":         c.phone,
                "phone_type":    c.phone_type,
                "linkedin_url":  c.linkedin_url,
                "sources":       " | ".join(c.sources),
                "notes":         c.notes,
            })
        all_orgs[kbo] = {
            "kbo": kbo, "name": org.name, "domain": org.domain,
            "findings": co.get("findings", 0),
            "contacts": [
                {"score": c.score, "name": c.name, "role": c.role,
                 "email": c.email, "email_status": c.email_status,
                 "phone": c.phone, "linkedin_url": c.linkedin_url}
                for c in org.contacts[:5]
            ]
        }

    # Write combined outputs
    if all_contacts:
        fields = list(all_contacts[0].keys())
        try:
            with open(combined_path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields)
                w.writeheader()
                for row in sorted(all_contacts,
                                  key=lambda x: (-x["findings"], -x["score"])):
                    w.writerow(row)
            ok(f"Combined contacts : {combined_path}  "
               f"({len(all_contacts)} rows)")
        except OSError as e:
            warn(f"Could not write combined CSV: {e}")

    if all_orgs:
        try:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(list(all_orgs.values()), f, indent=2, ensure_ascii=False)
            ok(f"Combined JSON     : {json_path}")
        except OSError as e:
            warn(f"Could not write combined JSON: {e}")

        # HTML report
        html_path = Path(output_dir) / "nis2_report.html"
        ci_export_html(list(all_orgs.values()), nuclei_output, str(html_path))

    header("CONTACT ENRICHMENT COMPLETE")
    ok(f"{len(targets)} companies enriched  |  "
       f"{len(all_contacts)} contacts ranked")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# CLI

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_args():
    p = argparse.ArgumentParser(
    description="NIS2 Belgian company scanner",
    formatter_class=argparse.RawDescriptionHelpFormatter)

    p.add_argument("--config",        metavar="FILE")
    p.add_argument("--init-config",   metavar="FILE", nargs="?",
                   const="campaign.yml")
    p.add_argument("--activity",      default=None, metavar="FILE")
    p.add_argument("--contact",       default=None, metavar="FILE")
    p.add_argument("--denomination",  default=None, metavar="FILE")
    p.add_argument("--output-dir",    default=None, metavar="DIR",
                   help=f"Output directory (default: {OUTPUT_DIR}_YYYYMMDD_HHMMSS)")

    fg = p.add_argument_group("Filtering")
    fg.add_argument("--sector",          action="append", metavar="NAME")
    fg.add_argument("--nace",            action="append", metavar="CODE")
    fg.add_argument("--limit",           type=int, metavar="N")
    fg.add_argument("--annex1-only",     action="store_true")
    fg.add_argument("--list-sectors",    action="store_true")

    tg = p.add_argument_group("Target processing")
    tg.add_argument("--resolve-dns",     action="store_true")
    tg.add_argument("--resolve-urls",    action="store_true")
    tg.add_argument("--exclude",         metavar="FILE")
    tg.add_argument("--per-sector-dirs", action="store_true")

    ng = p.add_argument_group("Nuclei")
    ng.add_argument("--templates",        action="append", metavar="PATH",
                    help="May be specified multiple times")
    ng.add_argument("--severity",         default=None, metavar="LIST")
    ng.add_argument("--rate",             type=int, default=None, metavar="N")
    ng.add_argument("--concur",           type=int, default=None, metavar="N")
    ng.add_argument("--timeout",          type=int, default=None, metavar="S")
    ng.add_argument("--proxy",            default=None, metavar="URL")
    ng.add_argument(
        "--run-mode",
        choices=["auto", "dryrun", "dryrun-skip-selected", "run-from-dryrun"],
        default="auto",
        metavar="MODE",
        help=("Execution mode: auto (default legacy behavior), dryrun "
              "(target generation only), dryrun-skip-selected (dry run + "
              "skip previously selected/scanned targets), run-from-dryrun "
              "(run nuclei using existing output-dir/targets.txt)"),
    )
    ng.add_argument("--update-nuclei",    action="store_true")
    ng.add_argument("--resume",           action="store_true")
    ng.add_argument("--clear-checkpoint", action="store_true")
    ng.add_argument("--schedule",         type=int, default=None, metavar="N")
    ng.add_argument("--no-retry",         action="store_true")
    ng.add_argument("--dry-run",          action="store_true")
    ng.add_argument("--verbose-nuclei",   action="store_true",
                    help="Pass -v to nuclei for per-request debug output")
    ng.add_argument("--force-refresh",    action="store_true",
                    help="Ignore cached targets and regenerate from scratch")
    ng.add_argument("--cache-minutes",    type=int,
                    default=TARGETS_MAX_AGE_MINUTES, metavar="N",
                    help=f"Max age in minutes before regenerating targets "
                         f"(default: {TARGETS_MAX_AGE_MINUTES})")

    p.add_argument("--summary-only",  action="store_true")
    p.add_argument("--export-xlsx",   action="store_true",
                   help="Write an Excel workbook alongside the coverage CSV")
    p.add_argument("--no-color",      action="store_true")
    p.add_argument("--quiet",         action="store_true",
                   help="Suppress info/bullet output; keep warnings, "
                        "errors and results")

    cg = p.add_argument_group("Contact intelligence")
    cg.add_argument("--contact-only",  action="store_true",
                    help="Skip the NIS2 scan; run contact intel on a single "
                         "target specified with --kbo and --domain")
    cg.add_argument("--kbo",           default="",  metavar="NUM",
                    help="KBO number for --contact-only  e.g. 0419649912")
    cg.add_argument("--domain",        default="",  metavar="DOMAIN",
                    help="Domain for --contact-only  e.g. bhak.be")
    cg.add_argument("--enrich-contacts", action="store_true",
                    help="After the NIS2 scan, run contact intel on the top "
                         "--contact-limit companies (ordered by finding count)")
    cg.add_argument("--contact-limit", type=int, default=10, metavar="N",
                    help="Max companies to enrich with contacts (default: 10)")
    cg.add_argument("--hunter-key",    default="",  metavar="KEY",
                    help="Hunter.io API key for email verification (optional)")
    cg.add_argument("--apollo-key",    default="",  metavar="KEY",
                    help="Apollo.io API key — 150 free credits/month "
                         "(https://app.apollo.io/settings/integrations/api)")
    cg.add_argument("--serp-delay",    type=float, default=1.8, metavar="S",
                    help="Seconds between HTTP requests in contact intel "
                         "(default: 1.8)")
    cg.add_argument("--contact-workers", type=int,
                    default=DEFAULT_CONTACT_WORKERS, metavar="N",
                    help="Parallel worker threads for contact enrichment "
                         f"(default: {DEFAULT_CONTACT_WORKERS})")
    cg.add_argument("--no-smtp",       action="store_true",
                    help="Skip SMTP email verification in contact intel")
    cg.add_argument("--contact-proxy", default="",  metavar="URL",
                    help="Proxy for contact intel HTTP (separate from nuclei "
                         "proxy)  e.g. http://user:pass@host:3128")

    dg = p.add_argument_group("Report delivery (Outlook / Power Automate)")
    dg.add_argument("--power-automate-webhook", default="", metavar="URL",
                    help="HTTP trigger URL from Power Automate flow")
    dg.add_argument("--outlook-to", default="", metavar="EMAILS",
                    help="Recipients passed to flow (comma/semicolon-separated)")
    dg.add_argument("--outlook-subject", default="", metavar="TEXT",
                    help="Custom email subject for Outlook delivery")
    dg.add_argument("--power-automate-timeout", type=int, default=None, metavar="S",
                    help="Webhook timeout in seconds (default: 20)")
    dg.add_argument("--attach-report-files", action="store_true",
                    help="Embed report artifacts as base64 attachments in payload")

    sp = p.add_argument_group("Report delivery (SharePoint / Microsoft Graph)")
    sp.add_argument("--sharepoint-upload", action="store_true",
                    help="Upload scan artifacts to SharePoint. Credentials come "
                         "from SP_TENANT_ID / SP_CLIENT_ID / SP_CLIENT_SECRET / "
                         "SP_HOSTNAME / SP_SITE_PATH env vars.")
    sp.add_argument("--sharepoint-folder", default="", metavar="NAME",
                    help="Target library folder (default: SP_TARGET_FOLDER env "
                         "var, else 'NIS2-Scans'). A UTC-timestamped subfolder "
                         "is created per run.")
    sp.add_argument("--intro-emails", action="store_true",
                    help="Write a neutral Dutch B2B service-introduction email "
                         "per company (no findings, no scan reference, with "
                         "opt-out). Sender from SENDER_* / UNSUB_URL env vars.")
    sp.add_argument("--split-by-company", action="store_true",
                    help="Split every aggregate output (scan CSV/JSON, coverage, "
                         "contacts, manifest, nuclei findings) into one file set "
                         "per company under by_company/. When set, delivery and "
                         "upload send ONLY the per-company tree, so no delivered "
                         "file spans multiple companies.")
    sp.add_argument("--ccb-disclosure", action=argparse.BooleanOptionalAction,
                    default=True,
                    help="Write a CCB-style coordinated vulnerability disclosure "
                         "report per company (addressed to the CCB/CSIRT and the "
                         "org's security mailbox; factual, no commercial content). "
                         "On by default; use --no-ccb-disclosure to skip. "
                         "Reporter identity from SENDER_* env vars.")
    sp.add_argument("--scan-ledger", default=SCAN_LEDGER_DEFAULT, metavar="PATH",
                    help=f"Global ledger of scanned targets, shared across "
                         f"campaigns (default: {SCAN_LEDGER_DEFAULT}). Targets "
                         f"already in it are skipped; it is updated every run and "
                         f"uploaded to a fixed SharePoint path (overwrite).")
    sp.add_argument("--ignore-ledger", action="store_true",
                    help="Scan targets even if they are already in the ledger "
                         "(the ledger is still updated afterwards).")

    return p.parse_args()

def resolve_run_mode(args) -> Tuple[str, bool, bool, bool]:
    """Resolve execution behavior from run-mode and compatibility flags."""
    if args.run_mode == "auto":
        if args.dry_run and args.resume:
            mode = "dryrun-skip-selected"
        elif args.dry_run:
            mode = "dryrun"
        else:
            mode = "auto"
    else:
        mode = args.run_mode
        if args.dry_run or args.resume:
            warn("--run-mode overrides legacy --dry-run/--resume flags.")

    effective_dry_run = mode in ("dryrun", "dryrun-skip-selected")
    effective_resume = (mode == "dryrun-skip-selected") if mode != "auto" else args.resume
    run_from_dryrun = mode == "run-from-dryrun"
    return mode, effective_dry_run, effective_resume, run_from_dryrun

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Main

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main():
    args = parse_args()

    global USE_COLOR, QUIET
    USE_COLOR = not args.no_color
    QUIET     = args.quiet

    RUN_START = time.monotonic()

    # ── Contact-only mode (no scan) ────────────────────────────────────
    if args.contact_only:
        if not args.kbo or not args.domain:
            error("--contact-only requires --kbo and --domain")
            sys.exit(1)
        proxies = ({"http": args.contact_proxy, "https": args.contact_proxy}
                   if args.contact_proxy else {"http": None, "https": None})
        org = ci_run_single(
            kbo        = args.kbo,
            domain     = args.domain,
            hunter_key = args.hunter_key,
            apollo_key = args.apollo_key,
            delay      = args.serp_delay,
            no_smtp    = args.no_smtp,
            proxies    = proxies,
        )
        out_dir = args.output_dir or "."
        os.makedirs(out_dir, exist_ok=True)
        stem = re.sub(r"[^\w]", "_", org.name or args.kbo)[:40]
        ci_print_report(org)
        csv_path = os.path.join(out_dir, f"contacts_{stem}.csv")
        json_path = os.path.join(out_dir, f"contacts_{stem}.json")
        ci_export_csv( org, csv_path)
        ci_export_json(org, json_path)
        # Build org dict for HTML report
        org_dict = {
            "kbo": org.kbo, "name": org.name, "domain": org.domain,
            "org_phone": org.org_phone, "email": org.org_email,
            "address": org.address, "email_pattern": org.email_pattern,
            "contacts": [
                {"score": c.score, "name": c.name, "role": c.role,
                 "linkedin_role": c.linkedin_role, "email": c.email,
                 "email_status": c.email_status, "phone": c.phone,
                 "linkedin_url": c.linkedin_url,
                 "linkedin_search_url": c.linkedin_search_url}
                for c in org.contacts
            ]
        }
        nuclei_out = os.path.join(out_dir, "nuclei_results.json")
        report_path = os.path.join(out_dir, f"report_{stem}.html")
        ci_export_html([org_dict], nuclei_out, report_path)
        maybe_deliver_report(
            output_dir=out_dir,
            nuclei_output=nuclei_out,
            webhook_url=args.power_automate_webhook,
            outlook_to=args.outlook_to,
            outlook_subject=args.outlook_subject,
            scanned_hosts=[args.domain],
            orgs=[org_dict],
            preferred_report_path=report_path,
            include_report_files=args.attach_report_files,
            extra_attachment_paths=[csv_path, json_path],
            timeout=args.power_automate_timeout,
        )
        maybe_upload_to_sharepoint(
            output_dir=out_dir,
            nuclei_output=nuclei_out,
            enabled=args.sharepoint_upload,
            target_folder=args.sharepoint_folder,
            scanned_hosts=[args.domain],
            preferred_report_path=report_path,
            include_report_files=args.attach_report_files,
            extra_paths=[csv_path, json_path],
        )
        ok(f"Done. {len(org.contacts)} contacts → {out_dir}/contacts_{stem}.*")
        sys.exit(0)

    if args.init_config:
        write_example_config(args.init_config); sys.exit(0)

    if args.config:
        if not Path(args.config).exists():
            warn(f"Config not found: {args.config}")
            write_example_config(args.config)
            bullet("Example config generated. "
                   "Edit it or continue with CLI flags.")
        else:
            apply_config_to_args(load_yaml_config(args.config), args)

    resolved_run_mode, effective_dry_run, effective_resume, run_from_dryrun = resolve_run_mode(args)

    args.activity     = args.activity     or ACTIVITY_FILE
    args.contact      = args.contact      or CONTACT_FILE
    args.denomination = args.denomination or DENOMINATION_FILE
    args.output_dir   = args.output_dir   or default_output_dir()
    args.templates    = args.templates    or DEFAULT_TEMPLATES
    args.severity     = args.severity     or DEFAULT_SEVERITY
    args.rate         = args.rate         or DEFAULT_RATE
    args.concur       = args.concur       or DEFAULT_CONCUR
    args.timeout      = args.timeout      or DEFAULT_TIMEOUT
    args.power_automate_timeout = args.power_automate_timeout or 20

    try:
        os.makedirs(args.output_dir, exist_ok=True)
    except OSError as e:
        error(f"Cannot create output dir '{args.output_dir}': {e}")
        sys.exit(1)

    targets_file  = os.path.join(args.output_dir, "targets.txt")
    nuclei_output = os.path.join(args.output_dir, "nuclei_results.json")
    manifest_file = os.path.join(args.output_dir, "nis2_companies_manifest.csv")

    if args.list_sectors:
        header("NIS2 Sectors")
        for sector, codes in NIS2_NACE_PREFIXES.items():
            annex = ("Annex I  – Essential" if sector in ANNEX_I_SECTORS
                     else "Annex II – Important")
            print(f"  {_c(annex, Fore.CYAN)}  {_c(sector, Fore.WHITE)}")
            print(f"    NACE: {', '.join(codes)}\n")
        sys.exit(0)

    step_start("Parse templates")
    header("STEP 0 – PARSE TEMPLATE CHECKS")
    template_checks = parse_template_checks(args.templates)
    total_checks    = sum(len(v["checks"]) for v in template_checks.values())
    ok(f"{len(template_checks)} template(s)  |  "
       f"{total_checks} named check(s) defined")
    step_end()

    if args.summary_only:
        print_scan_summary(nuclei_output,
                           output_dir=args.output_dir,
                           template_checks=template_checks,
                           export_xlsx=args.export_xlsx)
        sys.exit(0)

    if args.clear_checkpoint:
        clear_checkpoint(args.output_dir); sys.exit(0)

    sectors = list(args.sector or [])
    if args.annex1_only:
        sectors = (list(set(sectors) | set(ANNEX_I_SECTORS))
                   if sectors else list(ANNEX_I_SECTORS))

    header("RUN CONFIGURATION")
    bullet(f"Config          : {args.config           or '(none)'}")
    bullet(f"Sectors         : {sectors                or '(all NIS2)'}")
    bullet(f"NACE codes      : {args.nace              or '(all NIS2)'}")
    bullet(f"Limit           : {args.limit             or '(none)'}")
    bullet(f"Denomination    : {args.denomination}")
    bullet(f"Resolve DNS     : {args.resolve_dns}")
    bullet(f"Resolve URLs    : {args.resolve_urls}")
    bullet(f"Per-sector dirs : {args.per_sector_dirs}")
    bullet(f"Exclude list    : {args.exclude           or '(none)'}")
    bullet(f"Templates ({len(args.templates):>2})    :")
    for t in args.templates: bullet(f"    • {t}")
    bullet(f"Severity        : {args.severity}")
    bullet(f"Rate / Concur   : {args.rate} req/s / {args.concur} concurrent")
    bullet(f"Timeout         : {args.timeout}s")
    bullet(f"Proxy           : {args.proxy             or '(none)'}")
    bullet(f"Schedule        : "
           f"{f'{args.schedule}/day' if args.schedule else '(none)'}")
    bullet(f"Run mode        : {resolved_run_mode}")
    bullet(f"Resume          : {effective_resume}")
    bullet(f"No retry        : {args.no_retry}")
    bullet(f"Update nuclei   : {args.update_nuclei}")
    bullet(f"Dry run         : {effective_dry_run}")
    bullet(f"Force refresh   : {args.force_refresh}")
    bullet(f"Cache minutes   : {args.cache_minutes}")
    bullet(f"Export XLSX     : {args.export_xlsx}")
    bullet(f"Quiet           : {args.quiet}")
    bullet(f"Output dir      : {args.output_dir}")
    bullet(f"Enrich contacts : {args.enrich_contacts}")
    bullet(f"Power Automate  : {'(set)' if args.power_automate_webhook else '(none)'}")
    bullet(f"Outlook To      : {args.outlook_to or '(none)'}")
    bullet(f"Attach files    : {args.attach_report_files}")
    if args.enrich_contacts:
        bullet(f"Contact limit   : {args.contact_limit}")
        bullet(f"Contact workers : {args.contact_workers}")
        bullet(f"Hunter key      : {'(set)' if args.hunter_key else '(none)'}")
        bullet(f"SERP delay      : {args.serp_delay}s")
        bullet(f"No SMTP         : {args.no_smtp}")

    allowed         = codes_for_filter(sectors, args.nace or [])
    excludes        = load_exclude_list(args.exclude) if args.exclude else set()
    already_scanned = load_checkpoint(args.output_dir) if effective_resume else set()
    subdir_targets  = load_subdir_targets(args.output_dir)
    if subdir_targets:
        before = len(already_scanned)
        already_scanned.update(subdir_targets)
        added = len(already_scanned) - before
        if added:
            bullet(f"Skip source      : +{added:,} URL(s) from subdir targets")
    dead_urls_cache = load_dead_targets(args.output_dir)

    fresh, age_str = targets_are_fresh(targets_file, args.cache_minutes)
    if run_from_dryrun:
        if not Path(targets_file).exists() or Path(targets_file).stat().st_size == 0:
            error("--run-mode run-from-dryrun requires an existing non-empty "
                  f"targets file: {targets_file}")
            sys.exit(1)
        _, forced_age = targets_are_fresh(targets_file, max(args.cache_minutes, 10**9))
        fresh = True
        age_str = forced_age or "unknown age"
        if args.force_refresh:
            warn("--force-refresh is ignored when --run-mode run-from-dryrun is set.")
    elif args.force_refresh:
        fresh = False
        if Path(targets_file).exists():
            info("--force-refresh: ignoring cached targets file.")

    n_urls:       int       = 0
    final_urls:   List[str] = []
    sector_files: dict      = {}

    if fresh:
        header("TARGETS CACHE HIT — SKIPPING EXTRACTION")
        step_start("Load cached targets")
        final_urls   = load_targets_from_file(targets_file)
        if already_scanned:
            header("TARGETS CACHE RESUME FILTER")
            final_urls = apply_resume(final_urls, already_scanned)
            try:
                with open(targets_file, "w") as f:
                    f.write("\n".join(final_urls))
                ok(f"Updated cached targets after resume filter: {targets_file}  "
                   f"({len(final_urls):,} URLs)")
            except OSError as e:
                warn(f"Could not update targets file after resume filter: {e}")
        n_urls       = len(final_urls)
        sector_files = {}
        ok(f"Reusing {n_urls:,} targets from {targets_file}  "
           f"(generated {age_str})")
        bullet("Use --force-refresh to regenerate, or "
               "--cache-minutes to adjust TTL.")
        step_end()
    else:
        if not fresh and Path(targets_file).exists() and not effective_dry_run:
            warn(f"Targets file stale (>{args.cache_minutes}m) "
                 f"– regenerating.")

        entity_scan_goal = None
        if args.limit:
            entity_scan_goal = max(args.limit * 10, args.limit)

        while True:
            step_start("Filter NIS2 entities")
            header("STEP 1 – FILTER NIS2 ENTITIES")
            nis2_df, activity_reached_eof = load_nis2_entities(
                args.activity,
                allowed,
                stop_after_entities=entity_scan_goal,
            )
            entity_order = list(dict.fromkeys(
                e.strip() for e in nis2_df["EntityNumber"].astype(str)
            ))
            step_end()

            prefetch_candidates = len(entity_order)
            if args.limit:
                prefetch_candidates = min(max(args.limit * 10, args.limit), len(entity_order))

            need_more_entities = False
            while True:
                contact_target_entities = entity_order[:prefetch_candidates]
                contact_stop_after = prefetch_candidates if args.limit else None

                step_start("Load websites")
                header("STEP 2 – LOAD WEBSITES FOR TARGET SET")
                all_websites = load_websites_for_entities(
                    args.contact,
                    entity_numbers=set(contact_target_entities),
                    stop_after=contact_stop_after,
                )
                if not all_websites:
                    if args.limit and prefetch_candidates < len(entity_order):
                        next_prefetch = min(prefetch_candidates * 2, len(entity_order))
                        warn(f"No WEB entries yet in first {prefetch_candidates:,} entities; "
                             f"widening candidate pool to {next_prefetch:,}.")
                        prefetch_candidates = next_prefetch
                        step_end()
                        continue
                    if args.limit:
                        warn(f"No WEB entries in first {prefetch_candidates:,} entities "
                             f"from current STEP 1 window.")
                        need_more_entities = True
                        step_end()
                        break
                    error("No WEB entries found in contact.csv for target entities.")
                    sys.exit(1)
                step_end()

                step_start("Join entities ↔ websites")
                header("STEP 3 – JOIN ENTITIES ↔ WEBSITES")
                websites = {e: all_websites[e]
                            for e in contact_target_entities if e in all_websites}
                coverage = len(websites) / max(len(contact_target_entities), 1) * 100
                ok(f"Websites for NIS2 entities: {len(websites):,}  "
                   f"(coverage: {coverage:.1f}%)")
                if not websites:
                    if args.limit and prefetch_candidates < len(entity_order):
                        next_prefetch = min(prefetch_candidates * 2, len(entity_order))
                        warn(f"No websites joinable in first {prefetch_candidates:,} entities; "
                             f"widening candidate pool to {next_prefetch:,}.")
                        prefetch_candidates = next_prefetch
                        step_end()
                        continue
                    if args.limit:
                        warn(f"No websites joinable in first {prefetch_candidates:,} entities "
                             f"from current STEP 1 window.")
                        need_more_entities = True
                        step_end()
                        break
                    warn("No websites found for filtered entities.")
                    bullet(f"Sample (activity): {contact_target_entities[:5]}")
                    bullet(f"Sample (contact) : {list(all_websites.keys())[:5]}")
                    sys.exit(1)

                entity_numbers = set(websites.keys())
                if args.limit:
                    keep = set(e for e in contact_target_entities if e in websites)
                    nis2_candidate_df = nis2_df[
                        nis2_df["EntityNumber"].str.strip().isin(keep)]
                    websites = {e: websites[e] for e in keep}
                    entity_numbers = keep
                    bullet(f"Pre-fetching {len(keep):,} website candidates "
                           f"(from first {prefetch_candidates:,} entities) "
                           f"for {args.limit} live target(s)")
                else:
                    nis2_candidate_df = nis2_df
                step_end()

                step_start("Load company names")
                header("STEP 4 – COMPANY NAMES")
                denominations = load_denominations(args.denomination, entity_numbers)
                step_end()

                step_start("Process & write targets")
                header("STEP 5 – PROCESS & WRITE TARGETS")
                n_urls, sector_counts, final_urls, sector_files = save_outputs(
                    nis2_candidate_df, websites, denominations,
                    targets_file, manifest_file,
                    excludes, args.resolve_dns, args.resolve_urls,
                    already_scanned, args.per_sector_dirs, args.output_dir,
                    limit=args.limit, dead_urls_cache=dead_urls_cache,
                )
                info("Breakdown by NIS2 sector:\n")
                summary_table(sorted(sector_counts.items(), key=lambda x: -x[1]),
                              ["Sector", "Companies"])
                step_end()

                if not args.limit or n_urls > 0 or prefetch_candidates >= len(entity_order):
                    break

                next_prefetch = min(prefetch_candidates * 2, len(entity_order))
                warn(f"--limit {args.limit} requested {args.limit} live target(s), "
                     f"but current candidate pool yielded 0 after preflight. "
                     f"Widening from {prefetch_candidates:,} to {next_prefetch:,} entities.")
                prefetch_candidates = next_prefetch

            if not args.limit or n_urls > 0 or activity_reached_eof:
                break

            # Jump based on what we actually loaded (chunk granularity can
            # overshoot small goals), otherwise we can re-run the same window.
            next_entity_scan_goal = max(len(entity_order) * 2,
                                        (entity_scan_goal or args.limit))
            reason = ("no joinable WEB entries" if need_more_entities
                      else "0 live targets after preflight")
            warn(f"--limit {args.limit} requested {args.limit} live target(s), "
                 f"but first {len(entity_order):,} filtered entities yielded {reason}. "
                 f"Widening STEP 1 activity scan window to {next_entity_scan_goal:,} entities.")
            entity_scan_goal = next_entity_scan_goal

    if n_urls == 0:
        error("No valid URLs to scan."); sys.exit(1)

    # ── Scan ledger: skip targets already scanned in any prior campaign ──
    campaign = Path(args.output_dir).name or "default"
    if not args.ignore_ledger:
        _ledger = load_scan_ledger(args.scan_ledger)
        if _ledger:
            to_scan, skipped = filter_unscanned(final_urls, _ledger)
            if skipped:
                info(f"Scan ledger: skipping {len(skipped)} target(s) already "
                     f"scanned in a previous run.")
                final_urls = to_scan
                n_urls = len(final_urls)
                Path(targets_file).write_text("\n".join(final_urls),
                                              encoding="utf-8")
    if n_urls == 0:
        ok("All candidate targets are already in the scan ledger — nothing new "
           "to scan. Use --ignore-ledger to force a re-scan.")
        # Ledger unchanged; still refresh the SharePoint copy for consistency.
        upload_scan_ledger_to_sharepoint(
            args.scan_ledger, args.sharepoint_upload, args.sharepoint_folder)
        sys.exit(0)

    def do_dry_run(reason: str = "--dry-run flag set") -> None:
        header(f"DRY RUN  ({reason})")
        cmd = build_nuclei_cmd(targets_file, nuclei_output, args.templates,
                               args.rate, args.concur, args.timeout,
                               args.severity, args.proxy,
                               verbose=args.verbose_nuclei)
        bullet(f"Targets file  : {targets_file}  ({n_urls:,} URLs)")
        bullet(f"Manifest file : {manifest_file}")
        bullet(f"Results would : {nuclei_output}")
        bullet(f"Templates ({len(args.templates)}):")
        for t in args.templates: bullet(f"    • {t}")
        if sector_files:
            bullet("Per-sector targets:")
            for sector, sf in sector_files.items():
                bullet(f"  {sector:<40} → {sf}")
        print()
        info("nuclei command that WOULD run:\n")
        print("  " + " \\\n    ".join(cmd))
        print()
        ok("Dry run complete. Nothing sent to any server.")
        sys.exit(0)

    if effective_dry_run:
        reason = (f"--run-mode {resolved_run_mode}"
                  if resolved_run_mode != "auto"
                  else "--dry-run flag set")
        do_dry_run(reason=reason)

    header("STEP 6 – NUCLEI SCAN")
    if not check_nuclei():
        warn("nuclei not found – falling back to dry-run.")
        do_dry_run(reason="nuclei not installed")

    if args.update_nuclei:
        update_nuclei_templates()

    if args.schedule:
        step_start("Scheduled scan")
        run_scheduled(final_urls, args.output_dir, nuclei_output,
                      args.templates, args.rate, args.concur,
                      args.timeout, args.severity, args.proxy,
                      args.schedule, args.no_retry, already_scanned)
        step_end()
        _lk_s, _hx_s = load_url_lookup(args.output_dir)
        update_scan_ledger(args.scan_ledger, final_urls, _lk_s, _hx_s,
                           campaign, nuclei_output)
        upload_scan_ledger_to_sharepoint(
            args.scan_ledger, args.sharepoint_upload, args.sharepoint_folder)
        step_start("Generate summary")
        print_scan_summary(nuclei_output,
                           output_dir=args.output_dir,
                           template_checks=template_checks,
                           scanned_hosts=final_urls,
                           export_xlsx=args.export_xlsx)
        step_end()
        if args.enrich_contacts:
            ci_proxies = ({"http":  args.contact_proxy,
                           "https": args.contact_proxy}
                          if args.contact_proxy
                          else {"http": None, "https": None})
            ci_enrich_from_scan(
                args.output_dir, nuclei_output,
                args.contact_limit, args.hunter_key,
                args.apollo_key,
                args.serp_delay, args.no_smtp,
                args.contact_workers, ci_proxies)
        step_start("Per-company findings reports")
        _lookup, _hidx = load_url_lookup(args.output_dir)
        findings_reports = write_company_findings_reports(
            nuclei_output, args.output_dir, _lookup, _hidx)
        step_end()
        delivered_extra = list(findings_reports)
        if args.ccb_disclosure:
            step_start("CCB disclosure reports")
            delivered_extra += write_ccb_disclosure_reports(
                nuclei_output, args.output_dir, _lookup, _hidx)
            step_end()
        if args.intro_emails:
            step_start("Intro emails")
            delivered_extra += write_intro_emails(
                final_urls, args.output_dir, _lookup, _hidx)
            step_end()
        if args.split_by_company:
            step_start("Split outputs by company")
            delivered_extra += split_outputs_by_company(
                args.output_dir, nuclei_output, _lookup, _hidx)
            step_end()
        maybe_deliver_report(
            output_dir=args.output_dir,
            nuclei_output=nuclei_output,
            webhook_url=args.power_automate_webhook,
            outlook_to=args.outlook_to,
            outlook_subject=args.outlook_subject,
            scanned_hosts=final_urls,
            include_report_files=args.attach_report_files,
            extra_attachment_paths=delivered_extra,
            timeout=args.power_automate_timeout,
        )
        maybe_upload_to_sharepoint(
            output_dir=args.output_dir,
            nuclei_output=nuclei_output,
            enabled=args.sharepoint_upload,
            target_folder=args.sharepoint_folder,
            scanned_hosts=final_urls,
            include_report_files=args.attach_report_files,
            extra_paths=delivered_extra,
            only_extra_paths=args.split_by_company,
        )
        header("RUN COMPLETE")
        print_timings()
        save_timings(args.output_dir)
        sys.exit(0)

    step_start("Nuclei scan")
    rc = run_nuclei(targets_file, nuclei_output, args.templates,
                    args.rate, args.concur, args.timeout,
                    args.severity, args.proxy,
                    verbose=args.verbose_nuclei)
    step_end()

    if rc in (0, 130):
        save_checkpoint(args.output_dir,
                        list(already_scanned | set(final_urls)))
        _lk, _hx = load_url_lookup(args.output_dir)
        update_scan_ledger(args.scan_ledger, final_urls, _lk, _hx,
                           campaign, nuclei_output)
        upload_scan_ledger_to_sharepoint(
            args.scan_ledger, args.sharepoint_upload, args.sharepoint_folder)

    if rc in (0, 130) and not args.no_retry:
        step_start("Retry failed targets")
        retry_failed_targets(args.output_dir, nuclei_output, args.templates,
                             args.rate, args.concur, args.timeout,
                             args.severity, args.proxy)
        step_end()

    step_start("Generate summary")
    print_scan_summary(nuclei_output,
                       output_dir=args.output_dir,
                       template_checks=template_checks,
                       scanned_hosts=final_urls,
                       export_xlsx=args.export_xlsx)
    step_end()

    step_start("Per-company findings reports")
    _lookup, _hidx = load_url_lookup(args.output_dir)
    findings_reports = write_company_findings_reports(
        nuclei_output, args.output_dir, _lookup, _hidx)
    step_end()

    delivered_extra = list(findings_reports)
    if args.ccb_disclosure:
        step_start("CCB disclosure reports")
        delivered_extra += write_ccb_disclosure_reports(
            nuclei_output, args.output_dir, _lookup, _hidx)
        step_end()
    if args.intro_emails:
        step_start("Intro emails")
        delivered_extra += write_intro_emails(
            final_urls, args.output_dir, _lookup, _hidx)
        step_end()
    if args.split_by_company:
        step_start("Split outputs by company")
        delivered_extra += split_outputs_by_company(
            args.output_dir, nuclei_output, _lookup, _hidx)
        step_end()

    # ── Contact enrichment (post-scan) ─────────────────────────────────
    _results_exist = (Path(nuclei_output).exists()
                      and Path(nuclei_output).stat().st_size > 0)
    if args.enrich_contacts and rc in (0, 130):
        if rc == 130 and not _results_exist:
            warn("Scan interrupted before any results were written — "
                 "skipping contact enrichment. Re-run with --resume to "
                 "continue scanning, or use --contact-only for a single target.")
        else:
            step_start("Contact enrichment")
            ci_proxies = ({"http":  args.contact_proxy,
                           "https": args.contact_proxy}
                          if args.contact_proxy
                          else {"http": None, "https": None})
            ci_enrich_from_scan(
                args.output_dir, nuclei_output,
                args.contact_limit, args.hunter_key,
                args.apollo_key,
                args.serp_delay, args.no_smtp,
                args.contact_workers, ci_proxies)
            step_end()

    if rc in (0, 130):
        maybe_deliver_report(
            output_dir=args.output_dir,
            nuclei_output=nuclei_output,
            webhook_url=args.power_automate_webhook,
            outlook_to=args.outlook_to,
            outlook_subject=args.outlook_subject,
            scanned_hosts=final_urls,
            include_report_files=args.attach_report_files,
            extra_attachment_paths=delivered_extra,
            timeout=args.power_automate_timeout,
        )
        maybe_upload_to_sharepoint(
            output_dir=args.output_dir,
            nuclei_output=nuclei_output,
            enabled=args.sharepoint_upload,
            target_folder=args.sharepoint_folder,
            scanned_hosts=final_urls,
            include_report_files=args.attach_report_files,
            extra_paths=delivered_extra,
            only_extra_paths=args.split_by_company,
        )

    header("RUN COMPLETE")
    print_timings()
    save_timings(args.output_dir)
    total_elapsed = time.monotonic() - RUN_START
    ok(f"Total wall time : {total_elapsed:.1f}s  "
       f"({timedelta(seconds=int(total_elapsed))})")

    sys.exit(0 if rc in (0, 130) else rc)

if __name__ == "__main__":
    main()
